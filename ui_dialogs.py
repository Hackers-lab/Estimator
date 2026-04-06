"""
ui_dialogs.py
=============
All QDialog subclasses for ERP Estimate Generator.

Dialogs
-------
ProjectSetupDialog      NEW — project wizard shown on launch and via
                             "Project Settings" button. Captures project
                             type, subject, lat/long, division, circle,
                             UH toggle. Returns project_meta dict.

SearchDialog            — search materials / labour DB and add to estimate.
                          Unchanged from v4 except minor style tweaks.

SettingsDialog          — gateway to DB manager and Ruleset Manager.
                          Unchanged from v4.

DatabaseManagerDialog   — view, import, export the SQLite master DB.
                          Unchanged from v4.

RulesetManagerDialog    — full rule builder / simulator / editor.
                          Updated: TREE_DEF, FILTER_CHIPS, SIM_DEFAULTS
                          now imported from constants.py instead of being
                          hardcoded in the class body. SmartStructure and
                          SmartConsumer added throughout.
"""

import sqlite3
import json
import re
import openpyxl

import defaults
from app_config import APP_DISPLAY_NAME, APP_VERSION

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLineEdit, QListWidget, QPushButton, QCheckBox,
    QTabWidget, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QGroupBox, QComboBox,
    QSpinBox, QDoubleSpinBox, QHeaderView, QInputDialog,
    QWidget, QSplitter, QTreeWidget, QTreeWidgetItem,
    QLabel, QScrollArea, QDialogButtonBox, QFrame,
)
from PyQt6.QtCore import Qt

from constants import (
    PROPERTY_DATA, FORMULA_VARS,
    PROJECT_TYPES, SUPERVISION_RATES,
    SIM_DEFAULTS, TREE_DEF, FILTER_CHIPS,
)


class ClickableCard(QWidget):
    """Simple clickable container used by RulesetManager rule cards."""

    def __init__(self, on_click, parent=None):
        super().__init__(parent)
        self._on_click = on_click

    def mousePressEvent(self, event):
        if self._on_click:
            self._on_click()
        super().mousePressEvent(event)


# ─────────────────────────────────────────────────────────────────────────────
#  ProjectSetupDialog
# ─────────────────────────────────────────────────────────────────────────────

class ProjectSetupDialog(QDialog):
    """
    Project Setup Wizard — shown on first launch and via 'Project Settings'.

    Captures
    --------
    subject       : project name / description
    lat, long     : GPS coordinates
    division      : utility division name
    circle        : utility circle name
    project_type  : one of PROJECT_TYPES (drives supervision rate)
    use_uh        : bool — use UH (readymade) materials instead of raw steel
    supervision_rate : float — auto-derived from project_type

    Parameters
    ----------
    current_meta : dict  — pre-populate fields from existing project_meta
    parent       : QWidget
    first_run    : bool  — if True, shows a welcome banner; if False,
                           shows an "Edit Settings" heading instead
    """

    def __init__(self, current_meta: dict, parent=None, first_run: bool = True):
        super().__init__(parent)
        self._meta = dict(current_meta)
        self.setWindowTitle(
            "New Project Setup" if first_run else "Project Settings"
        )
        self.setMinimumWidth(480)
        self.setModal(True)

        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(16, 16, 16, 16)

        # ── Banner ────────────────────────────────────────────────────────
        if first_run:
            banner = QLabel(
                f"<b style='font-size:14px;'>{APP_DISPLAY_NAME} v{APP_VERSION}</b><br>"
                "<span style='color:#555;'>Set up your project before drawing.</span>"
            )
            banner.setStyleSheet(
                "background:#ddeeff; padding:10px; border-radius:5px;"
            )
            banner.setWordWrap(True)
            root.addWidget(banner)
        else:
            lbl = QLabel("<b>Edit Project Settings</b>")
            lbl.setStyleSheet("font-size:13px;")
            root.addWidget(lbl)

        # ── Form ──────────────────────────────────────────────────────────
        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Subject
        self._subject = QLineEdit(self._meta.get("subject", ""))
        self._subject.setPlaceholderText("e.g. GOCHIYA II LT Line Extension")
        form.addRow("Project Name:", self._subject)

        # Lat / Long side by side
        ll_w = QWidget()
        ll_l = QHBoxLayout(ll_w)
        ll_l.setContentsMargins(0, 0, 0, 0)
        ll_l.setSpacing(6)
        self._lat  = QLineEdit(self._meta.get("lat", ""))
        self._long = QLineEdit(self._meta.get("long", ""))
        self._lat.setPlaceholderText("Latitude")
        self._long.setPlaceholderText("Longitude")
        ll_l.addWidget(self._lat)
        ll_l.addWidget(self._long)
        form.addRow("Lat / Long:", ll_w)

        # Separator
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#ccc;")
        form.addRow(sep)

        # Project type (Shifting / Maintenance / Augmentation disabled — NSC only)
        self._proj_type = QComboBox()
        self._proj_type.addItems(PROJECT_TYPES)
        current_type = self._meta.get("project_type", "NSC")
        if current_type in PROJECT_TYPES:
            self._proj_type.setCurrentText(current_type)
        else:
            self._proj_type.setCurrentText("NSC")
        form.addRow("Project Type:", self._proj_type)

        # Supervision rate display (read-only)
        self._sup_lbl = QLabel()
        self._sup_lbl.setStyleSheet("color:#27ae60; font-weight:bold;")
        form.addRow("Supervision Rate:", self._sup_lbl)

        # Separator
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet("color:#ccc;")
        form.addRow(sep2)

        # UH toggle (NSC projects only)
        self._uh = QCheckBox(
            "Use UH (Readymade) Materials instead of Raw Steel"
        )
        self._uh.setStyleSheet("font-weight:bold; color:#107C41;")
        self._uh.setChecked(self._meta.get("use_uh", False))
        form.addRow(self._uh)

        # Now connect signal & trigger initial sync (after _uh exists)
        self._proj_type.currentTextChanged.connect(self._on_type_changed)
        self._on_type_changed(self._proj_type.currentText())

        root.addLayout(form)

        # ── Buttons ───────────────────────────────────────────────────────
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        ok_btn = btns.button(QDialogButtonBox.StandardButton.Ok)
        if ok_btn is not None:
            ok_btn.setText("✔ Continue" if first_run else "✔ Save")
            ok_btn.setStyleSheet(
                "background:#2980b9; color:white; font-weight:bold; padding:6px 16px;"
            )
        root.addWidget(btns)

    # ── Slots ─────────────────────────────────────────────────────────────────

    def _on_type_changed(self, proj_type: str):
        rate = SUPERVISION_RATES.get(proj_type, 0.10)
        self._sup_lbl.setText(f"{int(rate * 100)}%")
        self._meta["project_type"]     = proj_type
        self._meta["supervision_rate"] = rate
        self._sync_uh_visibility(proj_type)

    def _sync_uh_visibility(self, proj_type: str):
        is_nsc = proj_type == "NSC"
        self._uh.setEnabled(is_nsc)
        if not is_nsc:
            self._uh.setChecked(False)

    def _on_accept(self):
        subj = self._subject.text().strip()
        if not subj:
            QMessageBox.warning(
                self, "Required", "Please enter a Project Name."
            )
            return
        self._meta["subject"]  = subj
        self._meta["lat"]      = self._lat.text().strip()
        self._meta["long"]     = self._long.text().strip()
        self._meta["use_uh"]   = self._uh.isChecked()
        proj_type = self._proj_type.currentText()
        self._meta["project_type"]     = proj_type
        self._meta["supervision_rate"] = SUPERVISION_RATES.get(proj_type, 0.10)
        self.accept()

    def get_meta(self) -> dict:
        """Call after exec() == Accepted to retrieve the filled project_meta."""
        return dict(self._meta)


# ─────────────────────────────────────────────────────────────────────────────
#  SearchDialog
# ─────────────────────────────────────────────────────────────────────────────

class SearchDialog(QDialog):
    """
    Search the materials or labour database and pick an item to add
    to the live estimate as a custom (override) entry.
    """

    def __init__(self, db_type: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Search & Add {db_type}")
        self.setMinimumSize(600, 420)

        lay = QVBoxLayout(self)

        # Type badge
        badge_color = "#3498db" if db_type == "Material" else "#e67e22"
        badge = QLabel(f"  {db_type} Database  ")
        badge.setStyleSheet(
            f"background:{badge_color}; color:white; font-weight:bold;"
            "padding:4px 10px; border-radius:3px;"
        )
        badge.setFixedHeight(28)
        lay.addWidget(badge)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Type to search…")
        self._search.setStyleSheet("padding:6px; font-size:12px;")
        lay.addWidget(self._search)

        self._list = QListWidget()
        self._list.setStyleSheet("font-size:11px;")
        lay.addWidget(self._list)

        self._search.textChanged.connect(self._filter)
        self._list.itemDoubleClicked.connect(self.accept)

        add_btn = QPushButton(f"✔ Add Selected {db_type} to Estimate")
        add_btn.setStyleSheet(
            f"background:{badge_color}; color:white;"
            "font-weight:bold; padding:8px; font-size:12px;"
        )
        add_btn.clicked.connect(self.accept)
        lay.addWidget(add_btn)

        self._items_data: dict = {}
        self._load(db_type)

    def _load(self, db_type: str):
        conn   = sqlite3.connect("erp_master.db")
        cursor = conn.cursor()
        if db_type == "Material":
            cursor.execute(
                "SELECT item_code, item_name, unit, rate FROM materials "
                "ORDER BY item_name"
            )
        else:
            cursor.execute(
                "SELECT labor_code, task_name, unit, rate FROM labor "
                "ORDER BY task_name"
            )
        for row in cursor.fetchall():
            display = f"{row[1]}  ({row[2]})  —  Rs.{row[3]:.2f}"
            self._items_data[display] = {
                "code": row[0], "name": row[1],
                "unit": row[2], "rate": row[3],
                "type": db_type,
            }
            self._list.addItem(display)
        conn.close()

    def _filter(self, text: str):
        text = text.lower()
        for i in range(self._list.count()):
            item = self._list.item(i)
            if item is None:
                continue
            item.setHidden(text not in item.text().lower())

    def get_selected(self):
        sel = self._list.currentItem()
        return self._items_data.get(sel.text()) if sel else None


# ─────────────────────────────────────────────────────────────────────────────
#  SettingsDialog
# ─────────────────────────────────────────────────────────────────────────────

class SettingsDialog(QDialog):
    """Gateway dialog for advanced settings."""

    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("Advanced Settings")
        self.setFixedSize(320, 140)

        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(12, 12, 12, 12)

        db_btn = QPushButton("🗃️  Master Database (Excel Sync)")
        db_btn.clicked.connect(self.parent_app.open_db_manager)
        db_btn.setStyleSheet("padding:8px; font-size:12px;")
        lay.addWidget(db_btn)

        rule_btn = QPushButton("🧠  Ruleset Manager")
        rule_btn.clicked.connect(self.parent_app.open_rule_manager)
        rule_btn.setStyleSheet("padding:8px; font-size:12px;")
        lay.addWidget(rule_btn)

        lay.addStretch()


# ─────────────────────────────────────────────────────────────────────────────
#  PlacementDefaultsDialog
# ─────────────────────────────────────────────────────────────────────────────

class PlacementDefaultsDialog(QDialog):
    """
    Lets the user configure what values are applied when a new canvas
    object is placed (pole heights, conductor types, span lengths, etc.).
    Changes are persisted to ``defaults.json`` immediately on Save.
    """

    _HEIGHT_OPTIONS = {
        "PCC":    ["8MTR", "9MTR"],
        "STP":    ["9MTR", "9.5MTR", "11MTR"],
        "H-BEAM": ["13MTR"],
    }
    _LT_CONDUCTORS = ["AB Cable", "ACSR", "PVC Cable"]
    _HT_CONDUCTORS = ["ACSR", "AB Cable", "PVC Cable"]
    _CONDUCTOR_SIZES = {
        ("AB Cable",  "LT"): ["3CX50+1CX35", "3CX50+1CX16+1CX35", "3CX70+1CX16+1CX50"],
        ("AB Cable",  "HT"): ["3CX50+1CX150", "3CX95+1CX70"],
        ("ACSR",      "LT"): ["30SQMM", "50SQMM"],
        ("ACSR",      "HT"): ["30SQMM", "50SQMM"],
        ("PVC Cable", "LT"): ["10 SQMM", "16 SQMM", "25 SQMM", "50 SQMM", "95 SQMM", "120 SQMM"],
        ("PVC Cable", "HT"): ["10 SQMM", "16 SQMM", "25 SQMM", "50 SQMM", "95 SQMM", "120 SQMM"],
    }
    _SD_SIZES = ["10 SQMM", "16 SQMM", "25 SQMM", "50 SQMM"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Placement Defaults")
        self.setMinimumWidth(480)
        self.setModal(True)

        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(14, 14, 14, 14)

        tabs = QTabWidget()
        tabs.addTab(self._build_poles_tab(),  "⚡ Poles & Structures")
        tabs.addTab(self._build_spans_tab(),  "📏 Spans & Service Drop")
        root.addWidget(tabs)

        # ── Buttons ───────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        reset_btn = QPushButton("↩  Reset to Factory")
        reset_btn.clicked.connect(self._reset)
        btn_row.addWidget(reset_btn)
        btn_row.addStretch()

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._save)
        btns.rejected.connect(self.reject)
        btn_row.addWidget(btns)
        root.addLayout(btn_row)

    # ── Tab builders ──────────────────────────────────────────────────────

    def _build_poles_tab(self) -> QWidget:
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)

        d = defaults.current

        # ── LT Pole ───────────────────────────────────────────────────────
        lt_grp = QGroupBox("LT Pole")
        lt_frm = QFormLayout(lt_grp)

        self.lt_type2 = QComboBox()
        self.lt_type2.addItems(["PCC", "STP", "H-BEAM"])
        self.lt_type2.setCurrentText(d["lt_pole_type2"])
        lt_frm.addRow("Material:", self.lt_type2)

        self.lt_height = QComboBox()
        self._refresh_heights(self.lt_type2, self.lt_height, d["lt_height"])
        self.lt_type2.currentTextChanged.connect(
            lambda: self._refresh_heights(self.lt_type2, self.lt_height)
        )
        lt_frm.addRow("Height:", self.lt_height)

        self.lt_earth = QSpinBox()
        self.lt_earth.setRange(0, 20)
        self.lt_earth.setValue(d["lt_earth_count"])
        lt_frm.addRow("Earth Sets:", self.lt_earth)

        self.lt_stay = QSpinBox()
        self.lt_stay.setRange(0, 20)
        self.lt_stay.setValue(d["lt_stay_count"])
        lt_frm.addRow("Stay Sets:", self.lt_stay)

        lay.addWidget(lt_grp)

        # ── HT Pole ───────────────────────────────────────────────────────
        ht_grp = QGroupBox("HT Pole")
        ht_frm = QFormLayout(ht_grp)

        self.ht_type2 = QComboBox()
        self.ht_type2.addItems(["PCC", "STP", "H-BEAM"])
        self.ht_type2.setCurrentText(d["ht_pole_type2"])
        ht_frm.addRow("Material:", self.ht_type2)

        self.ht_height = QComboBox()
        self._refresh_heights(self.ht_type2, self.ht_height, d["ht_height"])
        self.ht_type2.currentTextChanged.connect(
            lambda: self._refresh_heights(self.ht_type2, self.ht_height)
        )
        ht_frm.addRow("Height:", self.ht_height)

        self.ht_earth = QSpinBox()
        self.ht_earth.setRange(0, 20)
        self.ht_earth.setValue(d["ht_earth_count"])
        ht_frm.addRow("Earth Sets:", self.ht_earth)

        self.ht_stay = QSpinBox()
        self.ht_stay.setRange(0, 20)
        self.ht_stay.setValue(d["ht_stay_count"])
        ht_frm.addRow("Stay Sets:", self.ht_stay)

        lay.addWidget(ht_grp)

        # ── Structure ─────────────────────────────────────────────────────
        st_grp = QGroupBox("Structure (DP / TP / 4P / DTR)")
        st_frm = QFormLayout(st_grp)

        self.st_type2 = QComboBox()
        self.st_type2.addItems(["PCC", "STP", "H-BEAM"])
        self.st_type2.setCurrentText(d["struct_pole_type2"])
        st_frm.addRow("Material:", self.st_type2)

        self.st_height = QComboBox()
        self._refresh_heights(self.st_type2, self.st_height, d["struct_height"])
        self.st_type2.currentTextChanged.connect(
            lambda: self._refresh_heights(self.st_type2, self.st_height)
        )
        st_frm.addRow("Height:", self.st_height)

        self.st_stay = QSpinBox()
        self.st_stay.setRange(0, 20)
        self.st_stay.setValue(d["struct_stay_count"])
        st_frm.addRow("Stay Sets:", self.st_stay)

        lay.addWidget(st_grp)

        # ── Shared ────────────────────────────────────────────────────────
        ext_grp = QGroupBox("Extension")
        ext_frm = QFormLayout(ext_grp)
        self.ext_ht = QDoubleSpinBox()
        self.ext_ht.setRange(1.0, 10.0)
        self.ext_ht.setSingleStep(0.5)
        self.ext_ht.setSuffix(" m")
        self.ext_ht.setValue(d["extension_height"])
        ext_frm.addRow("Default Ext. Height:", self.ext_ht)
        lay.addWidget(ext_grp)

        # ── Placement rules ──────────────────────────────────────────────
        pr_grp = QGroupBox("Placement Rules")
        pr_frm = QFormLayout(pr_grp)

        self.node_min_gap = QSpinBox()
        self.node_min_gap.setRange(5, 500)
        self.node_min_gap.setSuffix(" units")
        self.node_min_gap.setValue(int(d.get("node_min_gap", 36)))
        pr_frm.addRow("Min node spacing:", self.node_min_gap)
        lay.addWidget(pr_grp)

        lay.addStretch()
        return w

    def _build_spans_tab(self) -> QWidget:
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)

        d = defaults.current

        # ── LT Span ───────────────────────────────────────────────────────
        lt_grp = QGroupBox("LT Span")
        lt_frm = QFormLayout(lt_grp)

        self.lt_cond = QComboBox()
        self.lt_cond.addItems(self._LT_CONDUCTORS)
        self.lt_cond.setCurrentText(d["lt_conductor"])
        lt_frm.addRow("Conductor:", self.lt_cond)

        self.lt_size = QComboBox()
        self.lt_size.setEditable(True)
        self._refresh_sizes(self.lt_cond, self.lt_size, "LT", d["lt_conductor_size"])
        self.lt_cond.currentTextChanged.connect(
            lambda: self._refresh_sizes(self.lt_cond, self.lt_size, "LT")
        )
        lt_frm.addRow("Conductor Size:", self.lt_size)

        self.lt_len = QSpinBox()
        self.lt_len.setRange(1, 5000)
        self.lt_len.setSuffix(" m")
        self.lt_len.setValue(d["lt_span_length"])
        lt_frm.addRow("Default Length:", self.lt_len)

        self.lt_wires = QComboBox()
        self.lt_wires.addItems(["1", "2", "3", "4"])
        self.lt_wires.setCurrentText(str(d["lt_wire_count"]))
        lt_frm.addRow("Wire Count:", self.lt_wires)

        lay.addWidget(lt_grp)

        # ── HT Span ───────────────────────────────────────────────────────
        ht_grp = QGroupBox("HT Span")
        ht_frm = QFormLayout(ht_grp)

        self.ht_cond = QComboBox()
        self.ht_cond.addItems(self._HT_CONDUCTORS)
        self.ht_cond.setCurrentText(d["ht_conductor"])
        ht_frm.addRow("Conductor:", self.ht_cond)

        self.ht_size = QComboBox()
        self.ht_size.setEditable(True)
        self._refresh_sizes(self.ht_cond, self.ht_size, "HT", d["ht_conductor_size"])
        self.ht_cond.currentTextChanged.connect(
            lambda: self._refresh_sizes(self.ht_cond, self.ht_size, "HT")
        )
        ht_frm.addRow("Conductor Size:", self.ht_size)

        self.ht_len = QSpinBox()
        self.ht_len.setRange(1, 5000)
        self.ht_len.setSuffix(" m")
        self.ht_len.setValue(d["ht_span_length"])
        ht_frm.addRow("Default Length:", self.ht_len)

        self.ht_wires = QComboBox()
        self.ht_wires.addItems(["1", "2", "3", "4"])
        self.ht_wires.setCurrentText(str(d["ht_wire_count"]))
        ht_frm.addRow("Wire Count:", self.ht_wires)

        lay.addWidget(ht_grp)

        # ── Service Drop ──────────────────────────────────────────────────
        sd_grp = QGroupBox("Service Drop")
        sd_frm = QFormLayout(sd_grp)

        self.sd_size = QComboBox()
        self.sd_size.addItems(self._SD_SIZES)
        self.sd_size.setCurrentText(d["sd_conductor_size"])
        sd_frm.addRow("Cable Size:", self.sd_size)

        self.sd_len = QSpinBox()
        self.sd_len.setRange(1, 500)
        self.sd_len.setSuffix(" m")
        self.sd_len.setValue(d["sd_length"])
        sd_frm.addRow("Default Length:", self.sd_len)

        self.sd_phase = QComboBox()
        self.sd_phase.addItems(["1 Phase", "3 Phase"])
        self.sd_phase.setCurrentText(d["sd_phase"])
        sd_frm.addRow("Phase:", self.sd_phase)

        lay.addWidget(sd_grp)

        lay.addStretch()
        return w

    # ── Cascade helpers ───────────────────────────────────────────────────

    def _refresh_heights(self, type2_cb: QComboBox, height_cb: QComboBox,
                         current_val: str = "") -> None:
        opts = self._HEIGHT_OPTIONS.get(type2_cb.currentText(), ["8MTR", "9MTR"])
        height_cb.blockSignals(True)
        height_cb.clear()
        height_cb.addItems(opts)
        if current_val in opts:
            height_cb.setCurrentText(current_val)
        height_cb.blockSignals(False)

    def _refresh_sizes(self, cond_cb: QComboBox, size_cb: QComboBox,
                       voltage: str, current_val: str = "") -> None:
        key  = (cond_cb.currentText(), voltage)
        opts = self._CONDUCTOR_SIZES.get(key, ["10 SQMM"])
        size_cb.blockSignals(True)
        size_cb.clear()
        size_cb.addItems(opts)
        if current_val and current_val not in opts:
            size_cb.addItem(current_val)   # preserve user-typed custom size
        if current_val:
            size_cb.setCurrentText(current_val)
        size_cb.blockSignals(False)

    # ── Actions ───────────────────────────────────────────────────────────

    def _collect(self) -> dict:
        return {
            "lt_pole_type2":     self.lt_type2.currentText(),
            "lt_height":         self.lt_height.currentText(),
            "lt_earth_count":    self.lt_earth.value(),
            "lt_stay_count":     self.lt_stay.value(),
            "ht_pole_type2":     self.ht_type2.currentText(),
            "ht_height":         self.ht_height.currentText(),
            "ht_earth_count":    self.ht_earth.value(),
            "ht_stay_count":     self.ht_stay.value(),
            "struct_pole_type2": self.st_type2.currentText(),
            "struct_height":     self.st_height.currentText(),
            "struct_stay_count": self.st_stay.value(),
            "extension_height":  self.ext_ht.value(),
            "node_min_gap":      self.node_min_gap.value(),
            "lt_conductor":      self.lt_cond.currentText(),
            "lt_conductor_size": self.lt_size.currentText(),
            "lt_span_length":    self.lt_len.value(),
            "lt_wire_count":     self.lt_wires.currentText(),
            "ht_conductor":      self.ht_cond.currentText(),
            "ht_conductor_size": self.ht_size.currentText(),
            "ht_span_length":    self.ht_len.value(),
            "ht_wire_count":     self.ht_wires.currentText(),
            "sd_conductor_size": self.sd_size.currentText(),
            "sd_length":         self.sd_len.value(),
            "sd_phase":          self.sd_phase.currentText(),
        }

    def _save(self) -> None:
        defaults.save(self._collect())
        self.accept()

    def _reset(self) -> None:
        from PyQt6.QtWidgets import QMessageBox
        if QMessageBox.question(
            self, "Reset Defaults",
            "Reset all placement defaults to factory values?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes:
            defaults.reset_to_factory()
            self.reject()   # close; user can reopen to see factory values


# ─────────────────────────────────────────────────────────────────────────────
#  DatabaseManagerDialog
# ─────────────────────────────────────────────────────────────────────────────

class DatabaseManagerDialog(QDialog):
    """View, import, and export the master SQLite database via Excel."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Master DB — Excel Sync")
        self.setGeometry(100, 100, 860, 620)

        lay = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        imp_btn = QPushButton("📥 Import from Excel")
        exp_btn = QPushButton("📤 Export to Excel")
        imp_btn.clicked.connect(self.import_from_excel)
        exp_btn.clicked.connect(self.export_to_excel)
        imp_btn.setStyleSheet("padding:6px; font-weight:bold;")
        exp_btn.setStyleSheet("padding:6px; font-weight:bold;")
        btn_row.addWidget(imp_btn)
        btn_row.addWidget(exp_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        self.tabs          = QTabWidget()
        self.mat_table     = QTableWidget()
        self.labour_table  = QTableWidget()
        self.tabs.addTab(self.mat_table,    "Materials")
        self.tabs.addTab(self.labour_table, "Labour")
        lay.addWidget(self.tabs)

        self._load()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _fill(self, tbl: QTableWidget, data, headers):
        tbl.setRowCount(0)
        tbl.setColumnCount(len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        for r, row in enumerate(data):
            tbl.insertRow(r)
            for c, val in enumerate(row):
                tbl.setItem(r, c, QTableWidgetItem(str(val)))
        tbl.resizeColumnsToContents()
        hdr = tbl.horizontalHeader()
        if hdr is not None:
            hdr.setSectionResizeMode(
                1, QHeaderView.ResizeMode.Stretch
            )

    def _load(self):
        self.mat_table.clear()
        self.labour_table.clear()
        conn   = sqlite3.connect("erp_master.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM materials")
        self._fill(
            self.mat_table, cursor.fetchall(),
            ["Item Code", "Item Name", "Rate", "Unit"]
        )
        cursor.execute("SELECT * FROM labor")
        self._fill(
            self.labour_table, cursor.fetchall(),
            ["Labour Code", "Task Name", "Rate", "Unit"]
        )
        conn.close()

    # ── Actions ───────────────────────────────────────────────────────────────

    def export_to_excel(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export DB", "master_database.xlsx", "Excel (*.xlsx)"
        )
        if not fn:
            return
        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            conn   = sqlite3.connect("erp_master.db")
            cursor = conn.cursor()
            for tbl in ("materials", "labor"):
                ws = wb.create_sheet(tbl)
                cursor.execute(f"PRAGMA table_info({tbl})")
                ws.append([r[1] for r in cursor.fetchall()])
                cursor.execute(f"SELECT * FROM {tbl}")
                for row in cursor.fetchall():
                    ws.append(list(row))
            conn.close()
            wb.save(fn)
            QMessageBox.information(self, "Done", f"Exported to {fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))

    def import_from_excel(self):
        fn, _ = QFileDialog.getOpenFileName(
            self, "Import DB", "", "Excel (*.xlsx)"
        )
        if not fn:
            return
        try:
            wb   = openpyxl.load_workbook(fn)
            conn = sqlite3.connect("erp_master.db")
            cur  = conn.cursor()
            for tbl in ("materials", "labor"):
                if tbl not in wb.sheetnames:
                    continue
                ws = wb[tbl]
                cur.execute(f"DELETE FROM {tbl}")
                hdrs  = [str(c.value) for c in ws[1]]
                ph    = ", ".join(["?"] * len(hdrs))
                query = (
                    f"INSERT OR REPLACE INTO {tbl} "
                    f"({', '.join(hdrs)}) VALUES ({ph})"
                )
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        cur.execute(query, row)
            conn.commit()
            conn.close()
            self._load()
            QMessageBox.information(
                self, "Done",
                "Database imported. Refresh the estimate to see updated rates."
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))


# ─────────────────────────────────────────────────────────────────────────────
#  RulesetManagerDialog
# ─────────────────────────────────────────────────────────────────────────────

class RulesetManagerDialog(QDialog):
    """
    Full rule builder with three panels:

    Left   — hierarchical tree (SmartPole / SmartStructure / SmartSpan /
             SmartConsumer → sub-types) with search and rule counts.

    Centre — filtered card list with:
             • search box + AND/OR logic toggle
             • context-aware filter chips
             • collapsible Simulator strip

    Right  — rule editor:
             • item type / name / code picker
             • condition row builder with dropdowns
             • live condition preview
             • quantity formula input
             • Delete / Save footer

    All tree/chip/simulator data imported from constants.py — no
    hardcoded class-level dicts here.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ruleset Manager")
        self.setGeometry(60, 60, 1440, 880)

        # State
        self.rules               = []
        self.selected_rule_index = -1
        self.selected_result_item = None
        self.condition_widgets   = []
        self.active_tree_filter  = {}
        self.active_obj_type     = "SmartPole"
        self.filter_logic        = "AND"
        self.active_chips        = set()
        self.sim_visible         = False
        self.sim_widgets         = {}

        self._build_ui()
        self.load_rules()
        self._select_tree_root("SmartPole")

    # ═════════════════════════════════════════════════════════════════════════
    #  UI CONSTRUCTION
    # ═════════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        root = QHBoxLayout(self)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)
        root.addWidget(self._build_left())
        root.addWidget(self._build_centre())
        root.addWidget(self._build_right())

    @staticmethod
    def _combo_box_stylesheet() -> str:
        return (
            "QComboBox { color:#222; background:white; border:0.5px solid #ccc; "
            "border-radius:4px; padding:3px 6px; font-size:12px; }"
            "QComboBox QAbstractItemView { color:#222; background:white; "
            "selection-color:#222; selection-background-color:#ddeeff; }"
        )

    # ── LEFT ──────────────────────────────────────────────────────────────────

    def _build_left(self):
        panel = QWidget()
        panel.setFixedWidth(230)
        panel.setStyleSheet("background:#f5f5f5; border-right:1px solid #ddd;")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._tree_search = QLineEdit()
        self._tree_search.setPlaceholderText("Search tree…")
        self._tree_search.setStyleSheet(
            "margin:6px; padding:4px 8px; border:0.5px solid #ccc;"
            "border-radius:4px; font-size:12px;"
        )
        self._tree_search.textChanged.connect(self._filter_tree)
        lay.addWidget(self._tree_search)

        self._tree = QTreeWidget()
        self._tree.setHeaderHidden(True)
        self._tree.setStyleSheet("""
            QTreeWidget { border:none; background:#f5f5f5; font-size:12px; }
            QTreeWidget::item { padding:4px 6px; }
            QTreeWidget::item:selected { background:#ddeeff; color:#0C447C; }
            QTreeWidget::item:hover:!selected { background:#ebebeb; }
        """)
        self._tree.itemClicked.connect(self._on_tree_click)
        lay.addWidget(self._tree)

        self._populate_tree()
        return panel

    def _populate_tree(self):
        self._tree.clear()
        self._tree_items = []   # (QTreeWidgetItem, obj_type, filter_dict)

        def add_node(parent, label, obj_type, fdict, children):
            item = QTreeWidgetItem(
                parent if parent else self._tree, [label]
            )
            item.setData(0, Qt.ItemDataRole.UserRole, (obj_type, fdict))
            self._tree_items.append((item, obj_type, fdict))
            for ch in children:
                add_node(item, ch[0], ch[1], ch[2], ch[3])
            return item

        for entry in TREE_DEF:
            add_node(None, entry[0], entry[1], entry[2], entry[3])

        self._update_tree_counts()
        self._tree.expandToDepth(1)

    def _update_tree_counts(self):
        for item, obj_type, fdict in self._tree_items:
            base  = item.text(0).split("  ")[0]
            count = len(self._get_matching_rules(obj_type, fdict, set()))
            item.setText(0, f"{base}  ({count})" if count else base)

    def _filter_tree(self, text: str):
        text = text.lower()
        for item, *_ in self._tree_items:
            item.setHidden(bool(text) and text not in item.text(0).lower())

    def _select_tree_root(self, obj_type: str):
        for item, ot, fd in self._tree_items:
            if ot == obj_type and not fd:
                self._tree.setCurrentItem(item)
                self._on_tree_click(item, 0)
                return

    def _on_tree_click(self, item, _col):
        obj_type, fdict = item.data(0, Qt.ItemDataRole.UserRole)
        self.active_obj_type    = obj_type
        self.active_tree_filter = fdict
        self.active_chips.clear()
        self.selected_rule_index = -1
        self._rebuild_chips()
        self._refresh_cards()
        self._clear_editor()
        self._update_centre_title()

    # ── CENTRE ────────────────────────────────────────────────────────────────

    def _build_centre(self):
        self._centre = QWidget()
        lay = QVBoxLayout(self._centre)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Top bar
        topbar = QWidget()
        topbar.setStyleSheet("background:white; border-bottom:1px solid #ddd;")
        tl = QHBoxLayout(topbar)
        tl.setContentsMargins(10, 7, 10, 7)
        tl.setSpacing(8)

        self._centre_title = QLabel("")
        self._centre_title.setStyleSheet("font-weight:bold; font-size:13px;")

        self._card_search = QLineEdit()
        self._card_search.setPlaceholderText("Search rules…")
        self._card_search.setStyleSheet(
            "padding:4px 8px; border:0.5px solid #ccc; "
            "border-radius:4px; font-size:12px; max-width:200px;"
        )
        self._card_search.textChanged.connect(self._refresh_cards)

        self._type_filter = QComboBox()
        self._type_filter.addItems(["All", "Material", "Labor"])
        self._type_filter.setFixedWidth(86)
        self._type_filter.setStyleSheet(self._combo_box_stylesheet())
        self._type_filter.currentTextChanged.connect(self._refresh_cards)

        self._logic_btn = QPushButton("AND")
        self._logic_btn.setCheckable(True)
        self._logic_btn.setChecked(True)
        self._logic_btn.setFixedWidth(46)
        self._logic_btn.setStyleSheet(
            "QPushButton{padding:4px; border:1px solid #ccc;"
            "border-radius:4px; font-size:11px; font-weight:bold;}"
            "QPushButton:checked{background:#185FA5; color:white; border-color:#185FA5;}"
        )
        self._logic_btn.clicked.connect(self._toggle_logic)

        new_btn = QPushButton("+ New rule")
        new_btn.setStyleSheet(
            "background:#185FA5; color:white; border:none; "
            "padding:5px 12px; border-radius:4px; font-size:12px;"
        )
        new_btn.clicked.connect(self.create_new_rule)

        tl.addWidget(self._centre_title)
        tl.addStretch()
        tl.addWidget(self._card_search)
        tl.addWidget(self._type_filter)
        tl.addWidget(self._logic_btn)
        tl.addWidget(new_btn)
        lay.addWidget(topbar)

        # Chip bar
        self._chip_bar = QWidget()
        self._chip_bar.setStyleSheet(
            "background:#fafafa; border-bottom:1px solid #eee;"
        )
        self._chip_layout = QHBoxLayout(self._chip_bar)
        self._chip_layout.setContentsMargins(10, 5, 10, 5)
        self._chip_layout.setSpacing(6)
        lay.addWidget(self._chip_bar)

        # Card scroll area
        self._card_container = QWidget()
        self._card_container.setStyleSheet("background:#f8f8f8;")
        self._card_layout = QVBoxLayout(self._card_container)
        self._card_layout.setContentsMargins(8, 8, 8, 8)
        self._card_layout.setSpacing(5)
        self._card_layout.addStretch()

        scroll = QScrollArea()
        scroll.setWidget(self._card_container)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        lay.addWidget(scroll, 1)

        # Simulator (collapsed by default)
        lay.addWidget(self._build_sim_panel())
        return self._centre

    def _toggle_logic(self):
        self.filter_logic = "AND" if self._logic_btn.isChecked() else "OR"
        self._logic_btn.setText(self.filter_logic)
        self._refresh_cards()

    def _rebuild_chips(self):
        while self._chip_layout.count():
            item = self._chip_layout.takeAt(0)
            if item is None:
                continue
            w = item.widget()
            if w is not None:
                w.deleteLater()

        chips = FILTER_CHIPS.get(self.active_obj_type, [])
        self._chip_bar.setVisible(bool(chips))
        if not chips:
            return

        lbl = QLabel("Filter:")
        lbl.setStyleSheet("font-size:11px; color:#888;")
        self._chip_layout.addWidget(lbl)

        self._chip_checks = {}
        for label, key, val in chips:
            cb = QCheckBox(label)
            cb.setStyleSheet(
                "QCheckBox{font-size:11px; padding:2px 6px;"
                "border:0.5px solid #ccc; border-radius:10px; background:white;}"
                "QCheckBox:checked{background:#ddeeff;"
                "border-color:#378ADD; color:#0C447C;}"
            )
            chip_key = (key, str(val))
            cb.stateChanged.connect(
                lambda state, k=chip_key: self._on_chip(k, state)
            )
            self._chip_layout.addWidget(cb)
            self._chip_checks[chip_key] = cb

        self._chip_layout.addStretch()

    def _on_chip(self, chip_key, state):
        if state:
            self.active_chips.add(chip_key)
        else:
            self.active_chips.discard(chip_key)
        self._refresh_cards()

    def _update_centre_title(self):
        visible = self._visible_indices()
        self._centre_title.setText(
            f"{self.active_obj_type.replace('Smart','')}  —  "
            f"{len(visible)} rule(s)"
        )

    # ── Card list ─────────────────────────────────────────────────────────────

    @staticmethod
    def _cond_has(cond: str, key: str, val_str: str) -> bool:
        """Return True when a condition string contains a key/value test."""
        if key == "condition_true":
            return cond.strip() in ("", "True")

        if key.endswith("_gt"):
            base = key[:-3]
            return (f"{base} >" in cond) or (f"{base}>" in cond)
        if key.endswith("_ne"):
            base = key[:-3]
            return (
                f"{base} != '{val_str}'" in cond or
                f"{base} != \"{val_str}\"" in cond or
                f"{base} != {val_str}" in cond
            )

        if val_str.lower() == "false":
            return (
                f"not {key}" in cond or
                f"{key} == False" in cond
            )
        if val_str.lower() == "true":
            return (
                (key in cond and f"not {key}" not in cond)
                or f"{key} == True" in cond
            )

        return (
            f"{key} == '{val_str}'" in cond or
            f"{key} == \"{val_str}\"" in cond or
            f"{key} == {val_str}" in cond
        )

    def _get_matching_rules(self, obj_type, fdict, chips):
        result = []
        for i, rule in enumerate(self.rules):
            if rule.get("object") != obj_type:
                continue
            cond = rule.get("condition", "")

            if fdict:
                ok = True
                for prop, val in fdict.items():
                    vs = str(val)
                    if not self._cond_has(cond, prop, vs):
                        ok = False
                        break
                if not ok:
                    continue

            if chips:
                chip_results = []
                for (key, val_str) in chips:
                    match = self._cond_has(cond, key, val_str)
                    chip_results.append(match)

                if self.filter_logic == "AND" and not all(chip_results):
                    continue
                if self.filter_logic == "OR" and not any(chip_results):
                    continue

            result.append((i, rule))
        return result

    def _visible_indices(self):
        search = (
            self._card_search.text().lower()
            if hasattr(self, "_card_search") else ""
        )
        type_filter = (
            self._type_filter.currentText()
            if hasattr(self, "_type_filter") else "All"
        )
        matched = self._get_matching_rules(
            self.active_obj_type, self.active_tree_filter, self.active_chips
        )
        if type_filter != "All":
            matched = [
                (i, r) for i, r in matched
                if r.get("type", "Material") == type_filter
            ]
        if search:
            matched = [
                (i, r) for i, r in matched
                if search in r.get("item_name", "").lower()
                or search in r.get("condition", "").lower()
            ]
        return matched

    def _refresh_cards(self):
        while self._card_layout.count() > 1:
            item = self._card_layout.takeAt(0)
            if item is None:
                continue
            w = item.widget()
            if w is not None:
                w.deleteLater()

        matched   = self._visible_indices()
        sim_hits  = self._sim_hits() if self.sim_visible else set()

        for orig_idx, rule in matched:
            card = self._make_card(orig_idx, rule, orig_idx in sim_hits)
            self._card_layout.insertWidget(
                self._card_layout.count() - 1, card
            )

        self._update_centre_title()
        self._update_tree_counts()

    def _make_card(self, rule_index, rule, sim_hit=False):
        card     = ClickableCard(lambda idx=rule_index: self._on_card(idx))
        selected = rule_index == self.selected_rule_index

        bc = "#378ADD" if selected else ("#5DCAA5" if sim_hit else "#ddd")
        bw = "1.5px"   if (selected or sim_hit) else "0.5px"
        bg = "#eaf8f4" if sim_hit else "white"
        card.setStyleSheet(
            f"background:{bg}; border:{bw} solid {bc}; border-radius:6px;"
        )
        card.setCursor(Qt.CursorShape.PointingHandCursor)

        lay = QHBoxLayout(card)
        lay.setContentsMargins(9, 7, 9, 7)
        lay.setSpacing(8)

        r_type = rule.get("type", "Material")
        badge  = QLabel("M" if r_type == "Material" else "L")
        badge.setFixedSize(24, 24)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge.setStyleSheet(
            "border-radius:4px; font-size:10px; font-weight:bold; " + (
                "background:#ddeeff; color:#185FA5;" if r_type == "Material"
                else "background:#fff3e0; color:#854F0B;"
            )
        )
        lay.addWidget(badge)

        body = QWidget()
        bl   = QVBoxLayout(body)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(1)

        name_l = QLabel(rule.get("item_name", "Unnamed"))
        name_l.setStyleSheet("font-size:12px; font-weight:bold;")
        cond_l = QLabel(rule.get("condition", "") or "(no condition)")
        cond_l.setStyleSheet("font-size:11px; color:#555; font-family:monospace;")
        form_l = QLabel(f"qty = {rule.get('formula','1')}")
        form_l.setStyleSheet("font-size:10px; color:#999;")
        bl.addWidget(name_l)
        bl.addWidget(cond_l)
        bl.addWidget(form_l)
        lay.addWidget(body, 1)

        return card

    def _on_card(self, rule_index):
        self.selected_rule_index = rule_index
        self._refresh_cards()
        self._build_editor(self.rules[rule_index])

    # ── SIMULATOR ─────────────────────────────────────────────────────────────

    def _build_sim_panel(self):
        self._sim_outer = QWidget()
        self._sim_outer.setStyleSheet(
            "border-top:1px solid #ddd; background:#f5f5f5;"
        )
        ol = QVBoxLayout(self._sim_outer)
        ol.setContentsMargins(0, 0, 0, 0)
        ol.setSpacing(0)

        self._sim_toggle_btn = QPushButton(
            "▲  Simulator — set values and see which rules fire"
        )
        self._sim_toggle_btn.setStyleSheet(
            "text-align:left; padding:6px 12px; border:none; "
            "background:#f0f0f0; font-size:12px; font-weight:bold; color:#333;"
        )
        self._sim_toggle_btn.clicked.connect(self._toggle_sim)
        ol.addWidget(self._sim_toggle_btn)

        self._sim_body = QWidget()
        self._sim_body.setVisible(False)
        sb = QVBoxLayout(self._sim_body)
        sb.setContentsMargins(10, 8, 10, 8)
        sb.setSpacing(6)

        self._sim_inputs_w = QWidget()
        self._sim_inputs_l = QHBoxLayout(self._sim_inputs_w)
        self._sim_inputs_l.setContentsMargins(0, 0, 0, 0)
        self._sim_inputs_l.setSpacing(8)
        sb.addWidget(self._sim_inputs_w)

        run_row = QHBoxLayout()
        run_btn = QPushButton("▶  Run")
        run_btn.setStyleSheet(
            "background:#185FA5; color:white; border:none; "
            "padding:5px 16px; border-radius:4px; font-size:12px;"
        )
        run_btn.clicked.connect(self._run_sim)
        self._sim_count_lbl = QLabel("")
        self._sim_count_lbl.setStyleSheet(
            "font-size:12px; color:#0F6E56; font-weight:bold;"
        )
        run_row.addWidget(run_btn)
        run_row.addWidget(self._sim_count_lbl)
        run_row.addStretch()
        sb.addLayout(run_row)

        self._sim_table = QTableWidget(0, 4)
        self._sim_table.setHorizontalHeaderLabels(
            ["Type", "Item", "Qty", "Formula"]
        )
        sim_hdr = self._sim_table.horizontalHeader()
        if sim_hdr is not None:
            sim_hdr.setSectionResizeMode(
                1, QHeaderView.ResizeMode.Stretch
            )
        self._sim_table.setMaximumHeight(160)
        self._sim_table.setStyleSheet("font-size:11px;")
        self._sim_table.setVisible(False)
        sb.addWidget(self._sim_table)

        ol.addWidget(self._sim_body)
        return self._sim_outer

    def _toggle_sim(self):
        self.sim_visible = not self.sim_visible
        self._sim_body.setVisible(self.sim_visible)
        arrow = "▼" if self.sim_visible else "▲"
        self._sim_toggle_btn.setText(
            f"{arrow}  Simulator — set values and see which rules fire"
        )
        if self.sim_visible:
            self._rebuild_sim_inputs()

    def _rebuild_sim_inputs(self):
        while self._sim_inputs_l.count():
            item = self._sim_inputs_l.takeAt(0)
            if item is None:
                continue
            w = item.widget()
            if w is not None:
                w.deleteLater()
        self.sim_widgets = {}

        defaults = SIM_DEFAULTS.get(self.active_obj_type, {})
        for prop, (wtype, options, default) in defaults.items():
            col = QWidget()
            cl  = QVBoxLayout(col)
            cl.setContentsMargins(0, 0, 0, 0)
            cl.setSpacing(2)
            lbl = QLabel(prop)
            lbl.setStyleSheet("font-size:10px; color:#666;")
            cl.addWidget(lbl)

            if wtype == "combo":
                w = QComboBox()
                w.addItems([str(o) for o in options])
                w.setCurrentText(str(default))
                w.setStyleSheet("font-size:11px; padding:3px;")
            else:  # spin
                w = QSpinBox()
                w.setRange(options[0], options[1])
                w.setValue(default)
                w.setStyleSheet("font-size:11px; padding:3px;")
                w.setFixedWidth(64)

            cl.addWidget(w)
            self.sim_widgets[prop] = w
            self._sim_inputs_l.addWidget(col)
        self._sim_inputs_l.addStretch()

    def _get_sim_ctx(self) -> dict:
        ctx = {
            "use_uh":       False,
            "object_type":  self.active_obj_type,
            "project_type": "NSC",
        }
        for prop, w in self.sim_widgets.items():
            if isinstance(w, QSpinBox):
                ctx[prop] = w.value()
            else:
                val = w.currentText()
                if val == "True":
                    ctx[prop] = True
                elif val == "False":
                    ctx[prop] = False
                else:
                    try:
                        ctx[prop] = int(val)
                    except ValueError:
                        ctx[prop] = val
        return ctx

    def _sim_hits(self) -> set:
        if not self.sim_widgets:
            return set()
        ctx  = self._get_sim_ctx()
        hits = set()
        import math as _math
        for i, rule in enumerate(self.rules):
            if rule.get("object") != self.active_obj_type:
                continue
            cond = rule.get("condition", "True") or "True"
            try:
                if eval(
                    cond,
                    {"__builtins__": {}, "math": _math},
                    ctx
                ):
                    hits.add(i)
            except Exception:
                pass
        return hits

    def _run_sim(self):
        import math as _math
        ctx  = self._get_sim_ctx()
        hits = self._sim_hits()
        self._refresh_cards()

        self._sim_table.setRowCount(0)
        self._sim_table.setVisible(bool(hits))
        mat_n = lab_n = 0

        for i in sorted(hits):
            rule    = self.rules[i]
            r_type  = rule.get("type", "")
            formula = rule.get("formula", "1")
            try:
                qty = eval(
                    formula,
                    {"__builtins__": {"int": int, "round": round},
                     "math": _math},
                    ctx
                )
                qty_s = f"{qty:.3f}".rstrip("0").rstrip(".")
            except Exception:
                qty_s = formula

            r = self._sim_table.rowCount()
            self._sim_table.insertRow(r)
            self._sim_table.setItem(r, 0, QTableWidgetItem(r_type))
            self._sim_table.setItem(r, 1, QTableWidgetItem(rule.get("item_name", "")))
            self._sim_table.setItem(r, 2, QTableWidgetItem(qty_s))
            self._sim_table.setItem(r, 3, QTableWidgetItem(formula))
            if r_type == "Material":
                mat_n += 1
            else:
                lab_n += 1

        total = len(hits)
        self._sim_count_lbl.setText(
            f"{total} rule(s) fire  |  {mat_n} material, {lab_n} labour"
            if total else "No rules matched."
        )

    # ── RIGHT — editor ────────────────────────────────────────────────────────

    def _build_right(self):
        self._editor_outer = QWidget()
        self._editor_outer.setFixedWidth(400)
        self._editor_outer.setStyleSheet(
            "border-left:1px solid #ddd; background:white;"
        )
        lay = QVBoxLayout(self._editor_outer)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._editor_hdr = QLabel("Select a rule to edit")
        self._editor_hdr.setStyleSheet(
            "font-weight:bold; font-size:13px; padding:10px 14px;"
            "border-bottom:1px solid #ddd;"
        )
        lay.addWidget(self._editor_hdr)

        self._editor_body   = QWidget()
        self._editor_body_l = QVBoxLayout(self._editor_body)
        self._editor_body_l.setContentsMargins(14, 12, 14, 12)
        self._editor_body_l.setSpacing(10)
        self._editor_body_l.addStretch()

        scr = QScrollArea()
        scr.setWidget(self._editor_body)
        scr.setWidgetResizable(True)
        scr.setFrameShape(QScrollArea.Shape.NoFrame)
        lay.addWidget(scr, 1)

        # Footer
        footer = QWidget()
        footer.setStyleSheet("border-top:1px solid #ddd; background:white;")
        fl = QHBoxLayout(footer)
        fl.setContentsMargins(12, 8, 12, 8)

        self._del_btn = QPushButton("🗑 Delete")
        self._del_btn.setStyleSheet(
            "color:#c0392b; border:1px solid #c0392b; padding:5px 10px;"
            "border-radius:4px; background:white; font-size:12px;"
        )
        self._del_btn.clicked.connect(self.delete_selected_rule)
        self._del_btn.setEnabled(False)

        self._save_btn = QPushButton("💾 Save rule")
        self._save_btn.setStyleSheet(
            "background:#27ae60; color:white; border:none; padding:5px 14px;"
            "border-radius:4px; font-weight:bold; font-size:12px;"
        )
        self._save_btn.clicked.connect(self.save_rule_changes)
        self._save_btn.setEnabled(False)

        fl.addWidget(self._del_btn)
        fl.addStretch()
        fl.addWidget(self._save_btn)
        lay.addWidget(footer)
        return self._editor_outer

    # ── Editor helpers ────────────────────────────────────────────────────────

    def _sec_lbl(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "font-size:10px; font-weight:bold; color:#888; "
            "text-transform:uppercase; padding-bottom:4px; "
            "border-bottom:1px solid #eee; letter-spacing:.05em;"
        )
        return lbl

    def _field_row(self, label_text: str, widget: QWidget) -> QWidget:
        row = QWidget()
        rl  = QHBoxLayout(row)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)
        lbl = QLabel(label_text)
        lbl.setStyleSheet("font-size:11px; color:#666;")
        lbl.setFixedWidth(54)
        rl.addWidget(lbl)
        rl.addWidget(widget, 1)
        return row

    def _clear_editor(self):
        self._clear_layout(self._editor_body_l)
        self._editor_body_l.addStretch()
        self._editor_hdr.setText("Select a rule to edit")
        self._save_btn.setEnabled(False)
        self._del_btn.setEnabled(False)
        self.condition_widgets    = []
        self.selected_result_item = None

    def _build_editor(self, rule: dict):
        self._clear_layout(self._editor_body_l)
        self.condition_widgets    = []
        self.selected_result_item = None

        self._editor_hdr.setText(f"Editing: {rule.get('item_name','')}")
        self._save_btn.setEnabled(True)
        self._del_btn.setEnabled(True)

        # Item section
        self._editor_body_l.addWidget(self._sec_lbl("Item"))

        self._type_combo = QComboBox()
        self._type_combo.addItems(["Material", "Labor"])
        self._type_combo.setCurrentText(rule.get("type", "Material"))
        self._type_combo.setStyleSheet(self._combo_box_stylesheet())
        self._editor_body_l.addWidget(self._field_row("Type", self._type_combo))

        self._item_display = QLineEdit(rule.get("item_name", ""))
        self._item_display.setReadOnly(True)
        self._item_display.setStyleSheet("background:#f5f5f5; font-size:12px;")
        change_btn = QPushButton("Change…")
        change_btn.setStyleSheet("font-size:11px; padding:3px 8px;")
        change_btn.clicked.connect(self.search_database_for_item)

        item_row = QWidget()
        ir = QHBoxLayout(item_row)
        ir.setContentsMargins(0, 0, 0, 0)
        ir.setSpacing(4)
        lbl = QLabel("Item")
        lbl.setStyleSheet("font-size:11px; color:#666;")
        lbl.setFixedWidth(54)
        ir.addWidget(lbl)
        ir.addWidget(self._item_display, 1)
        ir.addWidget(change_btn)
        self._editor_body_l.addWidget(item_row)

        self._code_display = QLineEdit(rule.get("item_code", ""))
        self._code_display.setReadOnly(True)
        self._code_display.setStyleSheet(
            "background:#f5f5f5; font-size:11px; color:#888;"
        )
        self._editor_body_l.addWidget(
            self._field_row("Code", self._code_display)
        )

        # Condition section
        self._editor_body_l.addWidget(self._sec_lbl("Conditions"))

        self._cond_container = QWidget()
        self._cond_rows_l    = QVBoxLayout(self._cond_container)
        self._cond_rows_l.setContentsMargins(0, 0, 0, 0)
        self._cond_rows_l.setSpacing(3)
        self._editor_body_l.addWidget(self._cond_container)

        add_cond_btn = QPushButton("+ add condition row")
        add_cond_btn.setStyleSheet(
            "color:#185FA5; background:none; border:none; "
            "font-size:11px; text-align:left;"
        )
        add_cond_btn.clicked.connect(self.add_condition_row)
        self._editor_body_l.addWidget(add_cond_btn)

        # Preview
        self._editor_body_l.addWidget(self._sec_lbl("Condition preview"))
        self._preview_lbl = QLabel("")
        self._preview_lbl.setWordWrap(True)
        self._preview_lbl.setStyleSheet(
            "background:#f0f0f0; border-radius:4px; padding:5px 8px;"
            "font-family:monospace; font-size:11px; color:#333;"
        )
        self._editor_body_l.addWidget(self._preview_lbl)

        # Exact raw condition string (authoritative on save)
        self._editor_body_l.addWidget(self._sec_lbl("Raw condition (exact)"))
        self._raw_cond_input = QLineEdit(rule.get("condition", ""))
        self._raw_cond_input.setStyleSheet(
            "font-family:monospace; font-size:12px;"
        )
        self._editor_body_l.addWidget(self._raw_cond_input)
        self._raw_cond_input.textChanged.connect(self._update_preview)

        apply_builder_btn = QPushButton("Use builder preview as raw condition")
        apply_builder_btn.setStyleSheet(
            "font-size:11px; color:#185FA5; background:none; border:none; text-align:left;"
        )
        apply_builder_btn.clicked.connect(
            lambda: self._raw_cond_input.setText(" ".join(self._build_condition_parts()))
        )
        self._editor_body_l.addWidget(apply_builder_btn)

        self._cond_sync_hint = QLabel("")
        self._cond_sync_hint.setWordWrap(True)
        self._cond_sync_hint.setStyleSheet("font-size:10px; color:#a15c00;")
        self._editor_body_l.addWidget(self._cond_sync_hint)

        # Guided raw-condition composer
        self._editor_body_l.addWidget(self._sec_lbl("Guided condition composer"))

        helper_row = QWidget()
        hr = QHBoxLayout(helper_row)
        hr.setContentsMargins(0, 0, 0, 0)
        hr.setSpacing(4)

        obj_props = list(PROPERTY_DATA.get(rule.get("object", self.active_obj_type), {}).keys())
        self._expr_prop_cb = QComboBox()
        self._expr_prop_cb.addItems(obj_props)
        self._expr_prop_cb.setToolTip("Available condition keys for this object")
        self._expr_prop_cb.setStyleSheet(self._combo_box_stylesheet())

        self._expr_op_cb = QComboBox()
        self._expr_op_cb.addItems(["==", "!=", ">", "<", ">=", "<=", "in", "not in"])
        self._expr_op_cb.setFixedWidth(70)
        self._expr_op_cb.setStyleSheet(self._combo_box_stylesheet())

        self._expr_val_w = QLineEdit()
        self._expr_val_w.setPlaceholderText("value")

        ins_clause_btn = QPushButton("Insert clause")
        ins_clause_btn.setStyleSheet("font-size:11px; padding:3px 8px;")
        ins_clause_btn.clicked.connect(self._insert_clause_from_helper)

        hr.addWidget(self._expr_prop_cb, 2)
        hr.addWidget(self._expr_op_cb)
        hr.addWidget(self._expr_val_w, 2)
        hr.addWidget(ins_clause_btn)
        self._editor_body_l.addWidget(helper_row)

        token_row = QWidget()
        tr = QHBoxLayout(token_row)
        tr.setContentsMargins(0, 0, 0, 0)
        tr.setSpacing(4)

        for lbl, token in [
            ("AND", " and "),
            ("OR", " or "),
            ("NOT", "not "),
            ("(", "("),
            (")", ")"),
        ]:
            b = QPushButton(lbl)
            b.setStyleSheet("font-size:10px; padding:2px 6px;")
            b.clicked.connect(lambda _, t=token: self._insert_raw_text(t))
            tr.addWidget(b)

        clear_btn = QPushButton("Clear raw")
        clear_btn.setStyleSheet("font-size:10px; padding:2px 6px; color:#a00;")
        clear_btn.clicked.connect(lambda: self._raw_cond_input.setText(""))
        tr.addWidget(clear_btn)
        tr.addStretch()
        self._editor_body_l.addWidget(token_row)

        self._expr_prop_cb.currentTextChanged.connect(self._sync_expr_value_widget)
        self._sync_expr_value_widget(self._expr_prop_cb.currentText())

        # Formula
        self._editor_body_l.addWidget(self._sec_lbl("Quantity formula"))
        avail = FORMULA_VARS.get(rule.get("object", ""), [])
        hint  = QLabel(
            f"vars: {', '.join(avail)}" if avail else "no numeric vars"
        )
        hint.setStyleSheet("font-size:10px; color:#aaa;")
        self._editor_body_l.addWidget(hint)
        self._formula_input = QLineEdit(rule.get("formula", "1"))
        self._formula_input.setStyleSheet(
            "font-family:monospace; font-size:12px;"
        )
        self._editor_body_l.addWidget(self._formula_input)

        self._editor_body_l.addStretch()
        self._parse_conditions(rule)
        self._update_preview()

    # ── Condition rows ────────────────────────────────────────────────────────

    def _parse_conditions(self, rule: dict):
        cond = rule.get("condition", "")
        if not cond or cond.strip() == "True":
            self.add_condition_row()
            return
        tokens = re.split(r"\s+(and|or)\s+", cond, flags=re.IGNORECASE)
        self.add_condition_row(expression=tokens[0])
        for i in range(1, len(tokens), 2):
            logic = tokens[i].upper()
            expr  = tokens[i + 1] if i + 1 < len(tokens) else ""
            self.add_condition_row(logical_op=logic, expression=expr)

    def add_condition_row(self, logical_op=None, expression=None):
        obj   = self.active_obj_type
        props = list(PROPERTY_DATA.get(obj, {}).keys())

        row_w = QWidget()
        rl    = QHBoxLayout(row_w)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(3)

        logic_cb = QComboBox()
        logic_cb.addItems(["AND", "OR"])
        logic_cb.setFixedWidth(52)
        logic_cb.setVisible(len(self.condition_widgets) > 0)
        if logical_op:
            logic_cb.setCurrentText(logical_op)
        logic_cb.setStyleSheet(self._combo_box_stylesheet())
        logic_cb.currentTextChanged.connect(self._update_preview)

        prop_cb = QComboBox()
        prop_cb.addItems(props)
        prop_cb.setStyleSheet(self._combo_box_stylesheet())

        op_cb = QComboBox()
        op_cb.addItems(["==", "!=", ">", "<", ">=", "<="])
        op_cb.setFixedWidth(50)
        op_cb.setStyleSheet(self._combo_box_stylesheet())
        op_cb.currentTextChanged.connect(self._update_preview)

        val_w = QLineEdit()
        val_w.textChanged.connect(self._update_preview)

        rem_btn = QPushButton("✕")
        rem_btn.setFixedWidth(22)
        rem_btn.setStyleSheet(
            "color:#aaa; border:none; background:none; font-size:11px;"
        )

        rl.addWidget(logic_cb)
        rl.addWidget(prop_cb, 2)
        rl.addWidget(op_cb)
        rl.addWidget(val_w, 2)
        rl.addWidget(rem_btn)

        wm = {
            "widget": row_w, "logic": logic_cb,
            "prop": prop_cb, "op": op_cb, "value": val_w,
        }
        self.condition_widgets.append(wm)
        self._cond_rows_l.addWidget(row_w)

        prop_cb.currentTextChanged.connect(
            lambda t, w=wm: self._on_prop_change(t, w)
        )
        rem_btn.clicked.connect(
            lambda _, w=row_w: self._remove_cond_row(w)
        )

        self._on_prop_change(prop_cb.currentText(), wm)

        if expression:
            self._restore_expr(expression.strip(), wm, op_cb)

        self._update_preview()

    def _on_prop_change(self, prop: str, wm: dict):
        obj       = self.active_obj_type
        prop_info = PROPERTY_DATA.get(obj, {}).get(prop)
        cur       = wm["value"]

        if isinstance(prop_info, list):
            cls = QComboBox
        elif prop_info == "int":
            cls = QSpinBox
        else:
            cls = QLineEdit

        if not isinstance(cur, cls):
            new_w = cls()
            if isinstance(new_w, QSpinBox):
                new_w.setRange(-100000, 100000)
                new_w.valueChanged.connect(self._update_preview)
            elif isinstance(new_w, QComboBox):
                new_w.setStyleSheet(self._combo_box_stylesheet())
                new_w.currentTextChanged.connect(self._update_preview)
            else:
                new_w.textChanged.connect(self._update_preview)
            wm["widget"].layout().replaceWidget(cur, new_w)
            cur.deleteLater()
            wm["value"] = new_w
            cur = new_w

        if isinstance(cur, QComboBox) and isinstance(prop_info, list):
            cur.blockSignals(True)
            cur.clear()
            cur.addItems([str(p) for p in prop_info])
            cur.blockSignals(False)

        self._update_preview()

    def _restore_expr(self, expr: str, wm: dict, op_cb: QComboBox):
        not_m = re.match(r"^not\s+(\w+)$", expr)
        if not_m:
            wm["prop"].setCurrentText(not_m.group(1))
            self._on_prop_change(not_m.group(1), wm)
            op_cb.setCurrentText("==")
            v = wm["value"]
            (v.setCurrentText if isinstance(v, QComboBox) else v.setText)("False")
            return

        plain_m = re.match(r"^(\w+)$", expr)
        if plain_m:
            wm["prop"].setCurrentText(plain_m.group(1))
            self._on_prop_change(plain_m.group(1), wm)
            op_cb.setCurrentText("==")
            v = wm["value"]
            (v.setCurrentText if isinstance(v, QComboBox) else v.setText)("True")
            return

        m = re.match(r"(\w+)\s*([<>=!]+)\s*(.*)", expr)
        if not m:
            return
        prop = m.group(1).strip()
        op   = m.group(2).strip()
        val  = m.group(3).strip().strip("'\"")
        wm["prop"].setCurrentText(prop)
        self._on_prop_change(prop, wm)
        op_cb.setCurrentText(op)
        v = wm["value"]
        if isinstance(v, QComboBox):
            v.setCurrentText(val)
        elif isinstance(v, QSpinBox):
            try:
                v.setValue(int(float(val)))
            except ValueError:
                pass
        else:
            v.setText(val)

    def _remove_cond_row(self, widget: QWidget):
        if len(self.condition_widgets) <= 1:
            return
        self.condition_widgets = [
            w for w in self.condition_widgets if w["widget"] is not widget
        ]
        widget.deleteLater()
        if self.condition_widgets:
            self.condition_widgets[0]["logic"].setVisible(False)
        self._update_preview()

    def _build_condition_parts(self) -> list:
        parts = []
        for i, wm in enumerate(self.condition_widgets):
            prop = wm["prop"].currentText()
            op   = wm["op"].currentText()
            v    = wm["value"]
            if isinstance(v, QSpinBox):
                val = str(v.value())
            elif isinstance(v, QComboBox):
                val = v.currentText()
            else:
                val = v.text().strip()

            if not prop:
                continue
            if i > 0:
                parts.append(wm["logic"].currentText().lower())

            is_numeric = re.match(r"^-?\d+(\.\d+)?$", val)
            is_bool    = val.lower() in ("true", "false")
            if is_numeric or is_bool:
                parts.append(f"{prop} {op} {val}")
            else:
                parts.append(f"{prop} {op} '{val}'")
        return parts

    def _update_preview(self):
        parts = self._build_condition_parts()
        text  = " ".join(parts) if parts else "(no conditions)"
        if hasattr(self, "_preview_lbl"):
            self._preview_lbl.setText(text)
        if hasattr(self, "_cond_sync_hint") and hasattr(self, "_raw_cond_input"):
            raw = self._raw_cond_input.text().strip()
            builder = " ".join(parts).strip()
            # Normalise empty/no-condition variants for comparison
            if raw in ("", "True") and builder in ("", "(no conditions)"):
                self._cond_sync_hint.setText("")
            elif raw == builder:
                self._cond_sync_hint.setText("")
            else:
                self._cond_sync_hint.setText(
                    "Builder preview differs from raw condition. "
                    "For complex rules (groups/in/not), edit Raw condition directly."
                )

    def _insert_raw_text(self, token: str):
        if not hasattr(self, "_raw_cond_input"):
            return
        old = self._raw_cond_input.text()
        pos = self._raw_cond_input.cursorPosition()
        new = old[:pos] + token + old[pos:]
        self._raw_cond_input.setText(new)
        self._raw_cond_input.setCursorPosition(pos + len(token))

    def _sync_expr_value_widget(self, prop: str):
        obj = self.active_obj_type
        pinfo = PROPERTY_DATA.get(obj, {}).get(prop)

        new_w = None
        if isinstance(pinfo, list):
            cb = QComboBox()
            cb.addItems([str(v) for v in pinfo])
            cb.setEditable(False)
            new_w = cb
        elif pinfo == "int":
            sp = QSpinBox()
            sp.setRange(-100000, 100000)
            new_w = sp
        else:
            le = QLineEdit()
            le.setPlaceholderText("value")
            new_w = le

        row = self._expr_val_w.parentWidget()
        lay = row.layout() if row else None
        if lay is not None:
            lay.replaceWidget(self._expr_val_w, new_w)
        self._expr_val_w.deleteLater()
        self._expr_val_w = new_w

    def _insert_clause_from_helper(self):
        if not hasattr(self, "_expr_prop_cb"):
            return
        prop = self._expr_prop_cb.currentText().strip()
        op = self._expr_op_cb.currentText().strip()
        if not prop:
            return

        v = self._expr_val_w
        if isinstance(v, QSpinBox):
            raw_val = str(v.value())
        elif isinstance(v, QComboBox):
            raw_val = v.currentText().strip()
        else:
            raw_val = v.text().strip()

        def _fmt_token(t: str) -> str:
            is_num = re.match(r"^-?\d+(\.\d+)?$", t)
            is_bool = t.lower() in ("true", "false")
            if is_num or is_bool:
                return t
            return f"'{t}'"

        if op in ("in", "not in"):
            vals = [x.strip() for x in raw_val.split(",") if x.strip()]
            if not vals:
                vals = ["value"]
            seq = ", ".join(_fmt_token(x) for x in vals)
            clause = f"{prop} {op} ({seq})"
        else:
            if raw_val == "":
                raw_val = "value"
            clause = f"{prop} {op} {_fmt_token(raw_val)}"

        if self._raw_cond_input.text().strip():
            self._insert_raw_text(" and " + clause)
        else:
            self._insert_raw_text(clause)

    # ── Editor actions ────────────────────────────────────────────────────────

    def search_database_for_item(self):
        db_type, ok = QInputDialog.getItem(
            self, "Select type", "Which database?",
            ["Material", "Labor"], 0, False
        )
        if not (ok and db_type):
            return
        dlg = SearchDialog(db_type, self)
        if dlg.exec():
            item = dlg.get_selected()
            if item:
                self.selected_result_item = item
                self._item_display.setText(item["name"])
                self._code_display.setText(item.get("code", ""))
                self._type_combo.setCurrentText(item["type"])

    def save_rule_changes(self):
        if self.selected_rule_index == -1:
            return
        rule = self.rules[self.selected_rule_index]
        # Save the exact condition text so complex JSON conditions are preserved.
        rule["condition"] = self._raw_cond_input.text().strip()
        if self.selected_result_item:
            rule["type"]      = self.selected_result_item["type"]
            rule["item_code"] = self.selected_result_item.get("code", "")
            rule["item_name"] = (
                self.selected_result_item.get("name")
                or self.selected_result_item.get("item_name", "")
            )
        else:
            rule["type"] = self._type_combo.currentText()
        rule["formula"] = self._formula_input.text().strip() or "1"
        self.save_rules()
        self._update_tree_counts()
        self._refresh_cards()
        self._editor_hdr.setText(f"Editing: {rule.get('item_name','')}")
        QMessageBox.information(self, "Saved", "Rule saved successfully.")

    def create_new_rule(self):
        new_rule = {
            "object":    self.active_obj_type,
            "item_name": "New Rule — edit me",
            "condition": "",
            "type":      "Material",
            "item_code": "N/A",
            "formula":   "1",
        }
        self.rules.append(new_rule)
        self.save_rules()
        self.selected_rule_index = len(self.rules) - 1
        self._update_tree_counts()
        self._refresh_cards()
        self._build_editor(new_rule)

    def delete_selected_rule(self):
        if self.selected_rule_index == -1:
            return
        rule  = self.rules[self.selected_rule_index]
        reply = QMessageBox.question(
            self, "Delete rule",
            f"Delete:\n'{rule.get('item_name')}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            del self.rules[self.selected_rule_index]
            self.selected_rule_index = -1
            self.save_rules()
            self._update_tree_counts()
            self._refresh_cards()
            self._clear_editor()

    # ── Persistence ───────────────────────────────────────────────────────────

    def load_rules(self):
        try:
            with open("rules.json", "r") as f:
                self.rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.rules = []

    def save_rules(self):
        try:
            with open("rules.json", "w") as f:
                json.dump(self.rules, f, indent=2)
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to save rules:\n{exc}")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _clear_layout(self, layout):
        if not layout:
            return
        while layout.count():
            child = layout.takeAt(0)
            if child is None:
                continue
            w = child.widget()
            if w is not None:
                w.deleteLater()
                continue
            l = child.layout()
            if l is not None:
                self._clear_layout(l)
