"""
ui_dialogs.py
=============
All QDialog subclasses for ERP Estimate Generator.

Dialogs
-------
ProjectSetupDialog      NEW — project wizard shown on launch and via
                             "Project Settings" button. Captures project
                             type, subject, lat/long, division, circle,
                             UH toggle. Returns project_meta dict.

SearchDialog            — search materials / labour DB and add to estimate.
                          Unchanged from v4 except minor style tweaks.

SettingsDialog          — gateway to DB manager and Ruleset Manager.
                          Unchanged from v4.

DatabaseManagerDialog   — view, import, export the SQLite master DB.
                          Unchanged from v4.

RulesetManagerDialog    — full rule builder / simulator / editor.
                          Updated: TREE_DEF, FILTER_CHIPS, SIM_DEFAULTS
                          now imported from constants.py instead of being
                          hardcoded in the class body. SmartStructure and
                          SmartConsumer added throughout.
"""

import sqlite3
import json
import re
import openpyxl

from core import defaults
from core import property_catalog
from app_config import APP_DISPLAY_NAME, APP_VERSION

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLineEdit, QListWidget, QPushButton, QCheckBox,
    QTabWidget, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QGroupBox, QComboBox,
    QSpinBox, QDoubleSpinBox, QHeaderView, QInputDialog,
    QWidget, QSplitter, QTreeWidget, QTreeWidgetItem,
    QLabel, QScrollArea, QDialogButtonBox, QFrame,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

from core.constants import (
    PROPERTY_DATA, FORMULA_VARS,
    PROJECT_TYPES, SUPERVISION_RATES,
    SIM_DEFAULTS, TREE_DEF, FILTER_CHIPS,
)


def _runtime_property_data() -> dict:
    return property_catalog.build_property_data(PROPERTY_DATA)


def _runtime_sim_defaults() -> dict:
    return property_catalog.build_sim_defaults(SIM_DEFAULTS)


from ui.dialogs._shared import ClickableCard

class DatabaseManagerDialog(QDialog):
    """View, import, and export the master SQLite database via Excel."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Master DB — Excel Sync")
        self.setGeometry(100, 100, 860, 620)

        lay = QVBoxLayout(self)

        self._loading_table = False
        self._suppress_item_changed = False

        btn_row = QHBoxLayout()
        imp_btn = QPushButton("📥 Import from Excel")
        exp_btn = QPushButton("📤 Export to Excel")
        add_mat_btn = QPushButton("➕ Add Material")
        add_lab_btn = QPushButton("➕ Add Labour")
        del_btn = QPushButton("🗑 Delete Selected")
        imp_btn.clicked.connect(self.import_from_excel)
        exp_btn.clicked.connect(self.export_to_excel)
        add_mat_btn.clicked.connect(lambda: self._open_add_dialog("Material"))
        add_lab_btn.clicked.connect(lambda: self._open_add_dialog("Labour"))
        del_btn.clicked.connect(self._delete_selected)
        imp_btn.setStyleSheet("padding:6px; font-weight:bold;")
        exp_btn.setStyleSheet("padding:6px; font-weight:bold;")
        add_mat_btn.setStyleSheet("padding:6px; font-weight:bold;")
        add_lab_btn.setStyleSheet("padding:6px; font-weight:bold;")
        del_btn.setStyleSheet("padding:6px; font-weight:bold;")
        btn_row.addWidget(imp_btn)
        btn_row.addWidget(exp_btn)
        btn_row.addWidget(add_mat_btn)
        btn_row.addWidget(add_lab_btn)
        btn_row.addWidget(del_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        search_row = QHBoxLayout()
        self._db_search = QLineEdit()
        self._db_search.setPlaceholderText(
            "Search code/name/unit/rate in current tab..."
        )
        self._db_search.textChanged.connect(self._apply_db_filter)
        clear_btn = QPushButton("Clear")
        clear_btn.clicked.connect(lambda: self._db_search.clear())
        self._db_count = QLabel("")
        self._db_count.setStyleSheet("color:#666; font-size:11px;")
        search_row.addWidget(self._db_search, 1)
        search_row.addWidget(clear_btn)
        search_row.addWidget(self._db_count)
        lay.addLayout(search_row)

        self.tabs          = QTabWidget()
        self.mat_table     = QTableWidget()
        self.labour_table  = QTableWidget()
        self.tabs.addTab(self.mat_table,    "Materials")
        self.tabs.addTab(self.labour_table, "Labour")
        self.tabs.currentChanged.connect(lambda _i: self._apply_db_filter())
        self.mat_table.itemChanged.connect(
            lambda item: self._handle_rate_edit(item, "materials")
        )
        self.labour_table.itemChanged.connect(
            lambda item: self._handle_rate_edit(item, "labor")
        )
        lay.addWidget(self.tabs)

        self._load()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _fill(self, tbl: QTableWidget, data, headers):
        self._loading_table = True
        try:
            tbl.setRowCount(0)
            tbl.setColumnCount(len(headers))
            tbl.setHorizontalHeaderLabels(headers)
            for r, row in enumerate(data):
                tbl.insertRow(r)
                for c, val in enumerate(row):
                    it = QTableWidgetItem(str(val))
                    # Only rate column is editable from this dialog.
                    if c == 2:
                        it.setData(Qt.ItemDataRole.UserRole, str(val))
                    else:
                        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    tbl.setItem(r, c, it)
        finally:
            self._loading_table = False
        tbl.resizeColumnsToContents()
        hdr = tbl.horizontalHeader()
        if hdr is not None:
            hdr.setSectionResizeMode(
                1, QHeaderView.ResizeMode.Stretch
            )

    def _refresh_parent_estimate(self) -> None:
        p = self.parent()
        if p is None:
            return
        fn = getattr(p, "refresh_live_estimate", None)
        if callable(fn):
            fn()

    def _handle_rate_edit(self, item: QTableWidgetItem, table_name: str) -> None:
        if self._loading_table or self._suppress_item_changed:
            return
        if item is None or item.column() != 2:
            return

        tbl = item.tableWidget()
        if tbl is None:
            return

        code_item = tbl.item(item.row(), 0)
        old_val = item.data(Qt.ItemDataRole.UserRole)
        if code_item is None or old_val is None:
            return

        try:
            new_rate = float(item.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Invalid Rate", "Rate must be a number.")
            self._suppress_item_changed = True
            item.setText(str(old_val))
            self._suppress_item_changed = False
            return

        if str(new_rate) == str(old_val) or item.text().strip() == str(old_val):
            return

        yn = QMessageBox.question(
            self,
            "Confirm Rate Update",
            (
                "You are updating master DB price.\n\n"
                "This affects all future estimates that use this item.\n"
                f"Code: {code_item.text()}\n"
                f"Old rate: {old_val}\n"
                f"New rate: {new_rate}\n\n"
                "Proceed?"
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if yn != QMessageBox.StandardButton.Yes:
            self._suppress_item_changed = True
            item.setText(str(old_val))
            self._suppress_item_changed = False
            return

        try:
            conn = sqlite3.connect("erp_master.db")
            cur = conn.cursor()
            if table_name == "materials":
                cur.execute(
                    "UPDATE materials SET rate=? WHERE item_code=?",
                    (new_rate, code_item.text().strip()),
                )
            else:
                cur.execute(
                    "UPDATE labor SET rate=? WHERE labor_code=?",
                    (new_rate, code_item.text().strip()),
                )
            conn.commit()
            conn.close()
            item.setData(Qt.ItemDataRole.UserRole, str(new_rate))
            self._refresh_parent_estimate()
        except Exception as exc:
            QMessageBox.critical(self, "Update Failed", str(exc))
            self._suppress_item_changed = True
            item.setText(str(old_val))
            self._suppress_item_changed = False

    def _open_add_dialog(self, db_type: str) -> None:
        dlg = QDialog(self)
        dlg.setModal(True)
        dlg.setWindowTitle(f"Add {db_type}")
        lay = QVBoxLayout(dlg)
        frm = QFormLayout()

        code_ed = QLineEdit()
        name_ed = QLineEdit()
        rate_ed = QDoubleSpinBox()
        unit_ed = QLineEdit()

        rate_ed.setRange(0, 10_000_000)
        rate_ed.setDecimals(3)
        rate_ed.setSingleStep(1.0)

        if db_type == "Material":
            frm.addRow("Item Code:", code_ed)
            frm.addRow("Item Name:", name_ed)
        else:
            frm.addRow("Labour Code:", code_ed)
            frm.addRow("Task Name:", name_ed)
        frm.addRow("Rate:", rate_ed)
        frm.addRow("Unit:", unit_ed)
        lay.addLayout(frm)

        warn = QLabel("Warning: This adds a permanent master DB entry used in future estimates.")
        warn.setStyleSheet("color:#a94442; font-size:11px;")
        lay.addWidget(warn)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        code = code_ed.text().strip()
        name = name_ed.text().strip()
        unit = unit_ed.text().strip()
        rate = float(rate_ed.value())

        if not code or not name or not unit:
            QMessageBox.warning(self, "Missing Data", "Code, name and unit are required.")
            return

        try:
            conn = sqlite3.connect("erp_master.db")
            cur = conn.cursor()
            if db_type == "Material":
                cur.execute(
                    "INSERT INTO materials (item_code, item_name, rate, unit) VALUES (?, ?, ?, ?)",
                    (code, name, rate, unit),
                )
            else:
                cur.execute(
                    "INSERT INTO labor (labor_code, task_name, rate, unit) VALUES (?, ?, ?, ?)",
                    (code, name, rate, unit),
                )
            conn.commit()
            conn.close()
            self._load()
            self._refresh_parent_estimate()
            QMessageBox.information(self, "Added", f"{db_type} added successfully.")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Duplicate Code", "This code already exists in master DB.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))

    def _delete_selected(self) -> None:
        tbl = self._active_table()
        row = tbl.currentRow()
        if row < 0:
            QMessageBox.information(self, "No Selection", "Select one row to delete.")
            return

        code_item = tbl.item(row, 0)
        name_item = tbl.item(row, 1)
        if code_item is None:
            QMessageBox.warning(self, "Invalid Selection", "Could not read selected row.")
            return

        is_material = (self.tabs.currentIndex() == 0)
        table_name = "materials" if is_material else "labor"
        code_col = "item_code" if is_material else "labor_code"
        type_label = "Material" if is_material else "Labour"

        code = code_item.text().strip()
        name = name_item.text().strip() if name_item is not None else ""

        yn = QMessageBox.question(
            self,
            f"Delete {type_label}?",
            (
                f"This will permanently delete the selected {type_label.lower()} from master DB.\n\n"
                f"Code: {code}\n"
                f"Name: {name}\n\n"
                "This action cannot be undone. Continue?"
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if yn != QMessageBox.StandardButton.Yes:
            return

        try:
            conn = sqlite3.connect("erp_master.db")
            cur = conn.cursor()
            cur.execute(
                f"DELETE FROM {table_name} WHERE {code_col}=?",
                (code,),
            )
            conn.commit()
            conn.close()
            self._load()
            self._refresh_parent_estimate()
            QMessageBox.information(self, "Deleted", f"{type_label} deleted successfully.")
        except Exception as exc:
            QMessageBox.critical(self, "Delete Failed", str(exc))

    def _row_matches_filter(self, tbl: QTableWidget, row: int, needle: str) -> bool:
        if not needle:
            return True
        for c in range(tbl.columnCount()):
            cell = tbl.item(row, c)
            if cell is not None and needle in cell.text().lower():
                return True
        return False

    def _active_table(self) -> QTableWidget:
        idx = self.tabs.currentIndex()
        return self.mat_table if idx == 0 else self.labour_table

    def _apply_db_filter(self):
        tbl = self._active_table()
        needle = self._db_search.text().strip().lower()
        visible = 0
        total = tbl.rowCount()
        for r in range(total):
            match = self._row_matches_filter(tbl, r, needle)
            tbl.setRowHidden(r, not match)
            if match:
                visible += 1
        self._db_count.setText(f"{visible}/{total} shown")

    def _load(self):
        self.mat_table.clear()
        self.labour_table.clear()
        conn   = sqlite3.connect("erp_master.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM materials")
        self._fill(
            self.mat_table, cursor.fetchall(),
            ["Item Code", "Item Name", "Rate", "Unit"]
        )
        cursor.execute("SELECT * FROM labor")
        self._fill(
            self.labour_table, cursor.fetchall(),
            ["Labour Code", "Task Name", "Rate", "Unit"]
        )
        conn.close()
        self._apply_db_filter()

    # ── Actions ───────────────────────────────────────────────────────────────

    def export_to_excel(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export DB", "master_database.xlsx", "Excel (*.xlsx)"
        )
        if not fn:
            return
        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            conn   = sqlite3.connect("erp_master.db")
            cursor = conn.cursor()
            for tbl in ("materials", "labor"):
                ws = wb.create_sheet(tbl)
                cursor.execute(f"PRAGMA table_info({tbl})")
                ws.append([r[1] for r in cursor.fetchall()])
                cursor.execute(f"SELECT * FROM {tbl}")
                for row in cursor.fetchall():
                    ws.append(list(row))
            conn.close()
            wb.save(fn)
            QMessageBox.information(self, "Done", f"Exported to {fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))

    def import_from_excel(self):
        fn, _ = QFileDialog.getOpenFileName(
            self, "Import DB", "", "Excel (*.xlsx)"
        )
        if not fn:
            return
        try:
            wb   = openpyxl.load_workbook(fn)
            conn = sqlite3.connect("erp_master.db")
            cur  = conn.cursor()
            for tbl in ("materials", "labor"):
                if tbl not in wb.sheetnames:
                    continue
                ws = wb[tbl]
                cur.execute(f"DELETE FROM {tbl}")
                hdrs  = [str(c.value) for c in ws[1]]
                ph    = ", ".join(["?"] * len(hdrs))
                query = (
                    f"INSERT OR REPLACE INTO {tbl} "
                    f"({', '.join(hdrs)}) VALUES ({ph})"
                )
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        cur.execute(query, row)
            conn.commit()
            conn.close()
            self._load()
            QMessageBox.information(
                self, "Done",
                "Database imported. Refresh the estimate to see updated rates."
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))


# ─────────────────────────────────────────────────────────────────────────────
#  RulesetManagerDialog
# ─────────────────────────────────────────────────────────────────────────────

