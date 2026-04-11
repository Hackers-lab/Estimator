# Temporary script to inspect formulas and values in the uploaded Excel file
import openpyxl

wb = openpyxl.load_workbook('4.xlsx', data_only=False)
ws = wb.active

print('--- OLD (6.xlsx) ---')
wb_old = openpyxl.load_workbook('6.xlsx', data_only=False)
ws_old = wb_old.active
for row in ws_old.iter_rows(min_row=1, max_row=70):
    print([cell.value for cell in row])
    print([cell.coordinate + ':' + str(cell.value) if cell.data_type == 'f' else '' for cell in row])

print('--- NEW (8.xlsx) ---')
wb_new = openpyxl.load_workbook('8.xlsx', data_only=False)
ws_new = wb_new.active
for row in ws_new.iter_rows(min_row=1, max_row=70):
    print([cell.value for cell in row])
    print([cell.coordinate + ':' + str(cell.value) if cell.data_type == 'f' else '' for cell in row])
