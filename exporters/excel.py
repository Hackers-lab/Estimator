"""
excel_exporter.py
=================
ExcelExporter class — all Excel generation logic extracted from app.py.

Usage::

    from excel_exporter import ExcelExporter
    ExcelExporter(app_instance).generate()
"""

from __future__ import annotations

import json as _json
import os
from datetime import datetime
from typing import TYPE_CHECKING, Any

from PyQt6.QtWidgets import QFileDialog, QMessageBox

from canvas import SmartPole, SmartStructure, SmartSpan, SmartConsumer
from app_config import get_data_path

if TYPE_CHECKING:
    import openpyxl as _openpyxl_t
    from app import EstimateApp


def _xl():
    """Lazy-load openpyxl and its styles. Cached after first call."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    return openpyxl, Font, Alignment, PatternFill, Border, Side


class ExcelExporter:
    """Handles all Excel estimate generation for the ERP Estimate Generator."""

    # Unit weights for iron sections (kg/m).  0 = wire (formula gives MT directly).
    _IRON_UNIT_WEIGHTS: dict[str, float] = {
        "0102010611": 6.8,    # CH_75X40
        "0102010911": 9.8,    # CH_100X50
        "0101011311": 5.8,    # ANG_65X65X6
        "0101011011": 4.5,    # ANG_50X50X6
        "0103011511": 3.1,    # FLAT_65X6
        "0503010811": 0,      # GI Wire 5mm (qty already MT)
        "0503010711": 0,      # GI Wire 4mm (qty already MT)
    }

    def __init__(self, app: "EstimateApp") -> None:
        self._app = app

    # ── Main entry point ─────────────────────────────────────────────────────

    def generate(
        self,
        output_path: str | None = None,
        initial_dir: str | None = None,
        show_success: bool = True,
    ) -> str | None:
        app = self._app
        m   = app.project_meta
        subject = m.get("subject", "ERP_Estimate")
        safe    = "".join(c for c in subject if c not in r'\/*?:"<>|')
        default = f"{safe}_Estimate.xlsx" if safe else "ERP_Estimate.xlsx"

        filename = output_path
        if not filename:
            start_path = default
            if initial_dir:
                start_path = os.path.join(initial_dir.rstrip('/\\'), default)
            filename, _ = QFileDialog.getSaveFileName(
                app, "Export ERP Estimate", start_path, "Excel Files (*.xlsx)"
            )
            if not filename:
                return None

        openpyxl, *_ = _xl()
        wb = openpyxl.Workbook()
        self._write_estimate_sheet(wb, m)
        self._write_iron_breakup_sheet(wb)
        wb.save(filename)
        if show_success:
            msg = QMessageBox(app)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Excel Saved")
            msg.setText(f"Excel saved to:\n{filename}")
            open_file_btn = msg.addButton("Open File", QMessageBox.ButtonRole.ActionRole)
            open_folder_btn = msg.addButton("Open Folder", QMessageBox.ButtonRole.ActionRole)
            msg.addButton(QMessageBox.StandardButton.Close)
            msg.exec()
            if msg.clickedButton() == open_file_btn:
                try:
                    os.startfile(filename)  # type: ignore[attr-defined]
                except Exception as exc:
                    QMessageBox.warning(app, "Open File Failed", f"Could not open file.\n\n{exc}")
            elif msg.clickedButton() == open_folder_btn:
                try:
                    os.startfile(os.path.dirname(filename))  # type: ignore[attr-defined]
                except Exception as exc:
                    QMessageBox.warning(app, "Open Folder Failed", f"Could not open folder.\n\n{exc}")
        return filename

    def _write_estimate_sheet(self, wb: Any, m: dict) -> None:
        openpyxl, Font, Alignment, PatternFill, Border, Side = _xl()
        app = self._app
        escalation_count = len(getattr(app, 'escalations', [])) if hasattr(app, 'escalations') else 0

        # Find the actual cell for TOTAL MATERIAL COST (A)
        # This is always the last material summary row, which is after all escalations and sundries
        mat_total_row = 5 + len([x for x in app.live_bom_data if x["type"] == "Material"]) + 1 + escalation_count + 1  # +1 for 'Material Base Total', +escalation_count, +1 for sundries
        mat_total_cell = f'G{mat_total_row}'

        # Find the actual cell for TOTAL LABOR COST (B)
        lab_start_row = mat_total_row + 4  # 3 rows for blank, section header, then labor starts
        lab_total_row = lab_start_row + len([x for x in app.live_bom_data if x["type"] == "Labor"])  # after all labor rows
        lab_total_cell = f'G{lab_total_row}'
        ws  = wb.active
        assert ws is not None
        ws.title = "Estimate"

        sup_rate = m.get("supervision_rate", 0.10)
        sup_pct  = int(sup_rate * 100)

        # Header
        ws.merge_cells("A1:G1")
        ws["A1"] = "AUTOMATED ERP ESTIMATE"
        ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="4F81BD")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = (
            f"Subject: {m.get('subject','')}  |  "
            f"Type: {m.get('project_type','')}  |  "
            f"Date: {datetime.now().strftime('%d-%m-%Y')}"
        )
        ws.merge_cells("A3:G3")
        ws["A3"] = (
            f"Lat: {m.get('lat','')}   Long: {m.get('long','')}   |   "
            f"Materials: {'UH (Readymade)' if m.get('use_uh') else 'Raw Steel'}"
        )

        header_row = ["Sl No.", "Code", "Description", "Qty", "Unit", "Rate", "Amount"]
        ws.append(header_row)
        for cell in ws[4]:
            cell.font = Font(bold=True)
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["B"].width = 15

        row = 5
        mat_items = [x for x in app.live_bom_data if x["type"] == "Material"]
        lab_items = [x for x in app.live_bom_data if x["type"] == "Labor"]

        # ── Materials ──
        ws.cell(row, 3, "A. MATERIALS").font = Font(bold=True)
        row += 1

        mat_start_row = 6
        mat_end_row = mat_start_row + len(mat_items) - 1
        for i, item in enumerate(mat_items, 1):
            qty_val = round(item["qty"], 3)
            qty_is_int = (qty_val == int(qty_val))
            ws.append([
                i, item["code"], item["name"],
                int(qty_val) if qty_is_int else qty_val, item["unit"],
                round(item["rate"], 2), f'=ROUND(D{row}*F{row}, 2)'
            ])
            ws.cell(row, 4).number_format = '0' if qty_is_int else '0.000'
            ws.cell(row, 6).number_format = '0.00'
            ws.cell(row, 7).number_format = '0.00'
            row += 1

        # Calculate mat_base for further calculations, but write formula to Excel
        mat_base = sum(x["amt"] for x in mat_items)
        ws.append(["", "", "Material Base Total", "", "", "", f'=ROUND(SUM(G{mat_start_row}:G{mat_end_row}), 2)'])
        ws.cell(row, 7).number_format = '0.00'
        row += 1


        # Escalation rows (formulas)
        esc_rows = []
        mat_base_cell = f'G{row-1}'
        subtotal_formula = mat_base_cell
        for i, (fy, esc) in enumerate(app.escalations):
            # Each escalation is 5% of (mat_base + all previous escalations)
            if i == 0:
                esc_formula = f'=ROUND(({mat_base_cell})*0.05, 2)'
            else:
                prev_esc_cells = '+'.join(esc_rows)
                esc_formula = f'=ROUND(({mat_base_cell}+{prev_esc_cells})*0.05, 2)'
            ws.append([
                "", "", f"Add: Escalation @ 5% for FY {fy}", "", "", "", esc_formula
            ])
            ws.cell(row, 7).number_format = '0.00'
            row += 1
            esc_cell = f'G{row-1}'
            esc_rows.append(esc_cell)

        # Sundries (formula) - 5% of (mat_base + all escalations)
        if esc_rows:
            subtotal_formula = f'{mat_base_cell}+' + '+'.join(esc_rows)
        else:
            subtotal_formula = mat_base_cell
        sun_formula = f'=ROUND(({subtotal_formula})*0.05, 2)'
        ws.append(["", "", "Add: Sundries @ 5%", "", "", "", sun_formula])
        ws.cell(row, 7).number_format = '0.00'
        row += 1
        sun_row = row-1

        # TOTAL MATERIAL COST (A) (formula)
        # Grand total = mat_base + all escalations + sundries
        if esc_rows:
            grand_total_formula = f'=ROUND({mat_base_cell}+' + '+'.join(esc_rows) + f'+G{sun_row}, 2)'
        else:
            grand_total_formula = f'=ROUND({mat_base_cell}+G{sun_row}, 2)'
        ws.append(["", "", "TOTAL MATERIAL COST (A)", "", "", "", grand_total_formula])
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 7).font = Font(bold=True)
        ws.cell(row, 7).number_format = '0.00'
        mat_total_row = row  # Track the row where TOTAL MATERIAL COST (A) is written
        row += 2

        # ── Labor ──
        ws.cell(row, 3, "B. ERECTION / LABOR").font = Font(bold=True)
        row += 1

        lab_start_row = row
        lab_end_row = lab_start_row + len(lab_items) - 1
        for i, item in enumerate(lab_items, 1):
            qty_val = round(item["qty"], 3)
            qty_is_int = (qty_val == int(qty_val))
            ws.append([
                i, "", item["name"],
                int(qty_val) if qty_is_int else qty_val, item["unit"],
                round(item["rate"], 2), f'=ROUND(D{row}*F{row}, 2)'
            ])
            ws.cell(row, 4).number_format = '0' if qty_is_int else '0.000'
            ws.cell(row, 6).number_format = '0.00'
            ws.cell(row, 7).number_format = '0.00'
            row += 1


        # Formula for labor total
        ws.append(["", "", "TOTAL LABOR COST (B)", "", "", "", f'=ROUND(SUM(G{lab_start_row}:G{lab_end_row}), 2)'])
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 7).font = Font(bold=True)
        ws.cell(row, 7).number_format = '0.00'
        lab_total_row = row  # Track the row where TOTAL LABOR COST (B) is written
        row += 2

        # ── Taxes ──

        # Use the exact rows where totals were written
        mat_total_cell = f'G{mat_total_row}'
        lab_total_cell = f'G{lab_total_row}'

        # Supervision on (A+B)
        ws.cell(row, 3, "C. OVERHEADS & TAXES").font = Font(bold=True)
        row += 1

        # Supervision
        sup_formula = f'=ROUND(({mat_total_cell}+G{lab_total_row})*{sup_rate}, 2)'
        ws.append(["", "", f"Supervision @ {sup_pct}% on (A+B)", "", "", "", sup_formula])
        ws.cell(row, 7).number_format = '0.00'
        sup_row = row
        sup_cell = f'G{sup_row}'
        row += 1

        # GST on labor only
        gst_formula = f'=ROUND(G{lab_total_row}*0.18, 2)'
        ws.append(["", "", "GST @ 18% on Labour only", "", "", "", gst_formula])
        ws.cell(row, 7).number_format = '0.00'
        gst_row = row
        gst_cell = f'G{gst_row}'
        row += 1

        # Sub-Total (A+B+Supervision+GST)
        sub_total_formula = f'=ROUND({mat_total_cell}+G{lab_total_row}+{sup_cell}+{gst_cell}, 2)'
        ws.append(["", "", "Sub-Total", "", "", "", sub_total_formula])
        ws.cell(row, 7).number_format = '0.00'
        sub_total_row = row
        sub_total_cell = f'G{sub_total_row}'
        row += 1

        # Cess on (A+B+Supervision)
        cess_formula = f'=ROUND(({mat_total_cell}+G{lab_total_row}+{sup_cell})*0.01, 2)'
        ws.append(["", "", "Add: Cess @ 1% on (Mat+Lab+Sup)", "", "", "", cess_formula])
        ws.cell(row, 7).number_format = '0.00'
        cess_row = row
        cess_cell = f'G{cess_row}'
        row += 1

        # GRAND TOTAL (Sub-Total + Cess)
        grand_total_formula = f'=ROUND({sub_total_cell}+{cess_cell}, 2)'
        ws.append(["", "", "GRAND TOTAL", "", "", "", grand_total_formula])
        ws.cell(row, 3).font = Font(bold=True, size=12)
        ws.cell(row, 7).font = Font(bold=True, size=12, color="FF0000")
        ws.cell(row, 7).number_format = '0.00'

    # ── Iron breakup sheet ───────────────────────────────────────────────────

    def _compute_canvas_counts(self) -> dict:
        from canvas import SmartPole, SmartStructure, SmartSpan

        scene_items = self._app.scene.items()
        poles    = [i for i in scene_items if isinstance(i, SmartPole)]
        structs  = [i for i in scene_items if isinstance(i, SmartStructure)]
        spans    = [i for i in scene_items if isinstance(i, SmartSpan)]

        new_lt_poles = [p for p in poles if not p.is_existing and p.pole_type == "LT"]
        new_ht_poles = [p for p in poles if not p.is_existing and p.pole_type == "HT"]

        return {
            "lt_pole_count":    len(new_lt_poles),
            "ht_pole_count":    len(new_ht_poles),
            "dp_count":         len([s for s in structs if s.structure_type == "DP"]),
            "tp_count":         len([s for s in structs if s.structure_type == "TP"]),
            "4p_count":         len([s for s in structs if s.structure_type == "4P"]),
            "dtr_count":        len([s for s in structs if s.structure_type == "DTR"]),
            "cg_pole_count":    len([p for p in new_lt_poles if any(
                                    getattr(s, "has_cg", False)
                                    for s in getattr(p, "connected_spans", []))]),
            "pole_ext_count":   len([p for p in poles if getattr(p, "has_extension", False)]),
            "ht_ext_count":     len([p for p in new_ht_poles if getattr(p, "has_extension", False)]),
            "lt_acsr_count":    len([p for p in new_lt_poles if any(
                                    getattr(s, "conductor", "") == "ACSR"
                                    for s in getattr(p, "connected_spans", []))]),
            "ab_cable_count":   len([sp for sp in spans
                                    if getattr(sp, "conductor", "") == "AB Cable"
                                    and not getattr(sp, "is_existing_span", False)]),
        }

    def _write_iron_breakup_sheet(self, wb: Any) -> None:
        """
        Iron Breakup sheet: grouped by Section Type → Object Type → Recipe Items.
        Column F shows Iron (MT) per description row matching the estimate exactly.
        Totals reconcile with the Estimate sheet.
        """
        _, Font, Alignment, PatternFill, Border, Side = _xl()
        ws = wb.create_sheet("Iron Breakup")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 7
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 11
        ws.column_dimensions["F"].width = 13

        KG_PER_METRE = {
            "CH_75X40":    6.8,
            "CH_100X50":   9.8,
            "ANG_65X65X6": 5.8,
            "ANG_50X50X6": 4.5,
            "FLAT_65X6":   3.1,
            "FLAT_50X6":   2.5,
        }
        SECTION_LABELS = {
            "CH_75X40":    "M.S. Channel (75X40mm)",
            "CH_100X50":   "M.S. Channel (100X50mm)",
            "ANG_65X65X6": "M.S. Angle (65X65X6mm)",
            "ANG_50X50X6": "M.S. Angle (50X50X6mm)",
            "FLAT_65X6":   "M.S. Flat (65X6mm)",
            "FLAT_50X6":   "M.S. Flat (50X6mm)",
        }
        SECTION_ORDER = ["CH_100X50", "CH_75X40", "ANG_65X65X6", "ANG_50X50X6", "FLAT_65X6", "FLAT_50X6"]

        try:
            from core import db_gateway as _dbg
            recipes_list = _dbg.get_recipes()
            sections_dict = _dbg.get_sections()
        except Exception as e:
            recipes_list = []
            sections_dict = {}
            print(f"[Excel] Error loading recipes/sections: {e}")

        def find_recipe(rkey):
            return next((r for r in recipes_list if r["recipe_key"] == rkey), None)

        thin = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_row(row_idx, fill_color=None, bold=False, color="000000", size=10, center_cols=()):
            for col in range(1, 7):
                cell = ws.cell(row_idx, col)
                cell.border = thin_border
                if fill_color:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(name="Segoe UI", size=size, bold=bold, color=color)
                if col in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Canvas data ────────────────────────────────────────────────────────
        counts = self._compute_canvas_counts()
        scene_items = self._app.scene.items()
        from canvas import SmartPole
        poles = [i for i in scene_items if isinstance(i, SmartPole)]
        new_lt_poles = [p for p in poles if not p.is_existing and p.pole_type == "LT"]

        lt_recipe_counts: dict = {}
        for p in new_lt_poles:
            rkey = getattr(p, "iron_recipe", "None") or "None"
            if rkey == "None":
                rkey = "POLE_LT_IRON"
            lt_recipe_counts[rkey] = lt_recipe_counts.get(rkey, 0) + 1

        new_ht_poles = [p for p in poles if not p.is_existing and p.pole_type == "HT"]
        ht_recipe_counts: dict = {}
        for p in new_ht_poles:
            rkey = getattr(p, "iron_recipe", "None") or "None"
            if rkey == "None":
                rkey = "POLE_HT_IRON"
            ht_recipe_counts[rkey] = ht_recipe_counts.get(rkey, 0) + 1

        ext_lt = [p for p in poles if not p.is_existing and p.pole_type == "LT"
                  and getattr(p, "has_extension", False)]
        lt_ext_count = len(ext_lt)
        avg_ext_lt = round(
            sum(float(getattr(p, "extension_height", 1.5) or 1.5) for p in ext_lt) / lt_ext_count
            if lt_ext_count else 1.5, 2)

        ext_ht = [p for p in poles if not p.is_existing and p.pole_type == "HT"
                  and getattr(p, "has_extension", False)]
        avg_ext_ht = round(
            sum(float(getattr(p, "extension_height", 3.0) or 3.0) for p in ext_ht) / len(ext_ht)
            if ext_ht else 3.0, 2)

        def _lf(lpp, qpo, stored=""):
            if stored and str(stored).strip():
                s = str(stored).strip()
                return s if s.startswith("=") else f"={s}"
            return f"={qpo}*{lpp}" if qpo > 1 else str(lpp)

        # ── Build object list (title, canvas_count, items per object) ──────────
        # Each item: {description, section, lpp, qpo, lf}
        objects: list[dict] = []

        def add_recipe_obj(title, recipe_key, canvas_count):
            rec = find_recipe(recipe_key)
            if not rec or canvas_count <= 0:
                return
            items = []
            for ri in rec.get("items", []):
                lpp = float(ri.get("length_per_piece", ri.get("length", 0.0)))
                qpo = int(ri.get("qty_per_object", ri.get("qty", 1)))
                if lpp <= 0:
                    continue
                items.append({
                    "description": ri.get("description", ""),
                    "section": ri.get("section", ""),
                    "lpp": lpp, "qpo": qpo,
                    "lf": _lf(lpp, qpo, ri.get("length_formula", "")),
                })
            if items:
                objects.append({"title": title, "canvas_count": canvas_count, "items": items})

        def add_direct_obj(title, canvas_count, direct_items):
            if canvas_count <= 0:
                return
            items = [dict(it) for it in direct_items if it.get("lpp", 0) > 0]
            if items:
                objects.append({"title": title, "canvas_count": canvas_count, "items": items})

        # Build objects in logical order matching the estimate
        for rkey, cnt in lt_recipe_counts.items():
            rec = find_recipe(rkey)
            label = rec["name"] if rec else rkey
            add_recipe_obj(f"LT Pole Iron — {label}", rkey, cnt)

        add_direct_obj(f"LT Pole Extension ({lt_ext_count} nos)", lt_ext_count, [
            {"description": "Single Pole Extension (Angle)", "section": "ANG_65X65X6",
             "lpp": avg_ext_lt, "qpo": 1, "lf": f"={avg_ext_lt}"},
        ])

        add_recipe_obj(f"DP Structure Iron", "DP_IRON", counts["dp_count"])

        add_direct_obj(f"CG Bracket Iron ({counts['cg_pole_count']} poles)", counts["cg_pole_count"], [
            {"description": "CG Cradle Guard Bracket (Angle)", "section": "ANG_65X65X6",
             "lpp": 1.9, "qpo": 1, "lf": "=1.9"},
            {"description": "CG Cradle Guard Bracket (Flat)", "section": "FLAT_65X6",
             "lpp": 0.5, "qpo": 1, "lf": "=0.5"},
        ])

        add_recipe_obj(f"TP Structure Iron", "TP_IRON", counts["tp_count"])
        add_recipe_obj(f"4-Pole Structure Iron", "4P_IRON", counts["4p_count"])
        add_recipe_obj(f"DTR Substation Iron", "DTR_IRON", counts["dtr_count"])

        # Structure extensions — mirrors the STRUCT_HT_EXT_2P/3P/4P recipes used
        # by rules 232-234 (channel ×pole_count, flat ×pole_count). Without this
        # the exported breakup omitted iron the live estimate already counts.
        from canvas import SmartStructure as _SmartStructure
        _all_structs = [i for i in scene_items if isinstance(i, _SmartStructure)]
        for _slabel, _stypes, _ppc in (("DP/DTR", ("DP", "DTR"), 2), ("TP", ("TP",), 3), ("4P", ("4P",), 4)):
            _grp = [s for s in _all_structs
                    if getattr(s, "structure_type", "") in _stypes and getattr(s, "has_extension", False)]
            if not _grp:
                continue
            _cnt = len(_grp)
            _avg = round(sum(float(getattr(s, "extension_height", 3.0) or 3.0) for s in _grp) / _cnt, 2)
            add_direct_obj(f"{_slabel} Structure Extension ({_cnt} nos)", _cnt, [
                {"description": "Structure Extension (Channel)", "section": "CH_75X40",
                 "lpp": round(_avg * 2, 2), "qpo": _ppc, "lf": f"={_ppc}*{_avg}*2"},
                {"description": "Structure Extension (Flat)", "section": "FLAT_65X6",
                 "lpp": 3.0, "qpo": _ppc, "lf": f"={_ppc}*3"},
            ])

        for rkey, cnt in ht_recipe_counts.items():
            rec = find_recipe(rkey)
            label = rec["name"] if rec else rkey
            add_recipe_obj(f"HT Pole Iron — {label}", rkey, cnt)

        add_direct_obj(f"HT Pole Extension ({counts['ht_ext_count']} nos)", counts["ht_ext_count"], [
            {"description": "HT Pole Extension (Channel)", "section": "CH_75X40",
             "lpp": round(avg_ext_ht * 2, 2), "qpo": 1, "lf": f"={avg_ext_ht}*2"},
            {"description": "HT Pole Extension (Flat)", "section": "FLAT_65X6",
             "lpp": 3.0, "qpo": 1, "lf": "=3"},
        ])

        add_direct_obj(f"LT ACSR Bracket ({counts['lt_acsr_count']} poles)", counts["lt_acsr_count"], [
            {"description": "LT ACSR Bracket (Angle)", "section": "ANG_65X65X6",
             "lpp": 1.0, "qpo": 1, "lf": "=1"},
            {"description": "LT ACSR Bracket (Flat)", "section": "FLAT_65X6",
             "lpp": 1.0, "qpo": 1, "lf": "=1"},
        ])

        add_direct_obj(f"AB Cable Clamp ({counts['ab_cable_count']} spans)", counts["ab_cable_count"], [
            {"description": "AB Cable Clamp (Flat)", "section": "FLAT_65X6",
             "lpp": 0.5, "qpo": 1, "lf": "=0.5"},
        ])

        # ── Build section → [objects with items for that section] ─────────────
        from collections import defaultdict
        extra_sections = []
        for obj in objects:
            for it in obj["items"]:
                if it["section"] not in SECTION_ORDER and it["section"] not in extra_sections:
                    extra_sections.append(it["section"])

        write_order = SECTION_ORDER + extra_sections
        section_map: dict = defaultdict(list)  # sec_code → [{title, canvas_count, sec_items}]
        for obj in objects:
            for sec_code in write_order:
                sec_items = [it for it in obj["items"] if it["section"] == sec_code]
                if sec_items:
                    section_map[sec_code].append({
                        "title": obj["title"],
                        "canvas_count": obj["canvas_count"],
                        "sec_items": sec_items,
                    })

        section_totals: dict[str, float] = {}  # sec_code → total MT (Python-side)

        # ── Write sheet ────────────────────────────────────────────────────────
        cr = 1
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=6)
        ws.cell(cr, 1, "IRON CALCULATION BREAKUP")
        ws.cell(cr, 1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(cr, 1).font = Font(name="Segoe UI", bold=True, size=13, color="FFFFFF")
        ws.cell(cr, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[cr].height = 30
        cr += 1

        # Column headers
        for col, hdr in enumerate(["No.", "Description", "No", "Length (m)", "Total (m)", "Iron (MT)"], 1):
            ws.cell(cr, col, hdr)
        style_row(cr, fill_color="2E75B6", bold=True, color="FFFFFF", center_cols=(1, 3, 4, 5, 6))
        ws.row_dimensions[cr].height = 18
        cr += 1

        section_letter_idx = ord('A')
        all_section_iron_f: list[str] = []  # section-total F-cell refs for grand total

        for sec_code in write_order:
            sec_objects = section_map.get(sec_code, [])
            if not sec_objects:
                continue

            kg_m = KG_PER_METRE.get(sec_code)
            if not kg_m and sec_code in sections_dict:
                kg_m = sections_dict[sec_code].get("kg_per_metre", 0.0)
            kg_m = kg_m or 0.0
            sec_label = SECTION_LABELS.get(sec_code, sec_code)
            letter = chr(section_letter_idx)
            section_letter_idx += 1

            # ── Section header ─────────────────────────────────────────────────
            ws.cell(cr, 1, letter)
            ws.cell(cr, 2, f"{sec_label}  —  {kg_m} kg/m")
            ws.cell(cr, 6, "Iron (MT)")
            style_row(cr, fill_color="4472C4", bold=True, color="FFFFFF", size=11,
                      center_cols=(1, 3, 4, 5, 6))
            ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=5)
            ws.row_dimensions[cr].height = 22
            cr += 1

            sec_obj_iron_f: list[str] = []  # object-total F-cell refs for this section

            for obj in sec_objects:
                # Object sub-header
                ws.cell(cr, 2, f"  ↳  {obj['title']}  ({obj['canvas_count']} nos on canvas)")
                style_row(cr, fill_color="BDD7EE", bold=True, size=10, center_cols=())
                ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=6)
                ws.row_dimensions[cr].height = 18
                cr += 1

                # Item rows
                item_f_start = cr
                for row_idx, it in enumerate(obj["sec_items"], 1):
                    ws.cell(cr, 1, row_idx)
                    ws.cell(cr, 2, it["description"])
                    ws.cell(cr, 3, obj["canvas_count"])
                    ws.cell(cr, 4, it["lf"])
                    ws.cell(cr, 5, f"=C{cr}*D{cr}")
                    ws.cell(cr, 6, f"=ROUND(E{cr}*{kg_m}/1000,6)")
                    ws.cell(cr, 6).number_format = '0.000'
                    ws.cell(cr, 5).number_format = '0.00'
                    fill = "FFFFFF" if row_idx % 2 != 0 else "EBF3FB"
                    style_row(cr, fill_color=fill, center_cols=(1, 3))
                    ws.row_dimensions[cr].height = 17
                    mt_val = (it["lpp"] * it["qpo"] * obj["canvas_count"] * kg_m) / 1000.0
                    section_totals[sec_code] = section_totals.get(sec_code, 0.0) + mt_val
                    cr += 1

                item_f_end = cr - 1

                # Object total row
                ws.cell(cr, 2, f"  Iron Total — {obj['title']}")
                ws.cell(cr, 5, f"=SUM(E{item_f_start}:E{item_f_end})")
                ws.cell(cr, 6, f"=SUM(F{item_f_start}:F{item_f_end})")
                ws.cell(cr, 5).number_format = '0.00'
                ws.cell(cr, 6).number_format = '0.000'
                style_row(cr, fill_color="E2EFDA", bold=True, center_cols=(5, 6))
                ws.row_dimensions[cr].height = 18
                sec_obj_iron_f.append(f"F{cr}")
                cr += 1

            # Section total (sum of all object totals for this section)
            sec_formula = "+".join(sec_obj_iron_f) if sec_obj_iron_f else "0"
            ws.cell(cr, 2, f"Section Total — {sec_label}")
            ws.cell(cr, 6, f"=ROUND({sec_formula},3)")
            ws.cell(cr, 6).number_format = '0.000'
            style_row(cr, fill_color="D9E1F2", bold=True, size=11, center_cols=(6,))
            ws.row_dimensions[cr].height = 20
            all_section_iron_f.append(f"F{cr}")
            cr += 2  # blank row between sections

        # ── Iron Summary + Grand Total ─────────────────────────────────────────
        if section_totals:
            cr += 1  # extra blank row before summary block

            # IRON SUMMARY header
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=6)
            ws.cell(cr, 1, "IRON SUMMARY")
            ws.cell(cr, 1).fill = PatternFill("solid", fgColor="1F4E79")
            ws.cell(cr, 1).font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
            ws.cell(cr, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[cr].height = 24
            cr += 1

            # One row per section type with MT > 0
            section_total_cells: list[str] = []
            for sec_code in SECTION_ORDER:
                if section_totals.get(sec_code, 0.0) <= 0:
                    continue
                sec_label = SECTION_LABELS.get(sec_code, sec_code)
                ws.cell(cr, 2, sec_label)
                ws.cell(cr, 6, round(section_totals[sec_code], 3))
                ws.cell(cr, 6).number_format = '0.000'
                style_row(cr, fill_color="D9E1F2", bold=False, center_cols=(6,))
                ws.row_dimensions[cr].height = 17
                section_total_cells.append(f"F{cr}")
                cr += 1

            cr += 1  # blank row before totals

            # Total iron (as per estimate)
            grand_formula = "+".join(section_total_cells) if section_total_cells else "0"
            ws.cell(cr, 2, "TOTAL IRON (as per Estimate):")
            ws.cell(cr, 6, f"=ROUND({grand_formula},3)")
            ws.cell(cr, 6).number_format = '0.000'
            style_row(cr, fill_color="FFF2CC", bold=True, size=11, center_cols=(6,))
            ws.row_dimensions[cr].height = 22
            grand_row = cr
            cr += 1

            # Wastage @ 3%
            ws.cell(cr, 2, "  Add: Fabrication Wastage @ 3%:")
            ws.cell(cr, 6, f"=ROUND(F{grand_row}*0.03,3)")
            ws.cell(cr, 6).number_format = '0.000'
            style_row(cr, fill_color="FFF2CC", bold=True, center_cols=(6,))
            ws.row_dimensions[cr].height = 18
            wastage_row = cr
            cr += 1

            # Final total — no "procurement" language
            ws.cell(cr, 2, "TOTAL IRON (incl. wastage):")
            ws.cell(cr, 6, f"=ROUND(F{grand_row}+F{wastage_row},3)")
            ws.cell(cr, 6).number_format = '0.000 "MT"'
            style_row(cr, fill_color="C6EFCE", bold=True, size=12, center_cols=(6,))
            ws.row_dimensions[cr].height = 26

