"""
Main application module for the ERP Estimate Generator.
Version 5.0 — Redesigned with:
  - Project Setup Wizard (project type, UH toggle, supervision rate)
  - SmartPole with pole_type2 (PCC/STP/H-BEAM) + cascading heights
  - SmartStructure as separate canvas object (DP/TP/4P/DTR)
  - SmartSpan with unified conductor_size + voltage auto-detection
  - SmartConsumer (replaces SmartHome) with phase + agency supply
  - Iron Breakup sheet in Excel export
  - Detail View toggle for canvas symbols
  - Full backward compatibility with v4 saved JSON files
"""

import sys
import math
import json
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QComboBox, QGraphicsScene,
    QFormLayout, QGroupBox, QSpinBox, QLineEdit,
    QFileDialog, QMessageBox, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QSplitter, QGraphicsView,
    QDialog, QDialogButtonBox, QDoubleSpinBox, QScrollArea,
    QFrame, QMenu, QTextBrowser, QInputDialog
)
from PyQt6.QtGui import (
    QPen, QBrush, QColor, QPainter, QPageLayout, QPageSize, QFont,
    QAction, QKeySequence, QIcon, QPixmap
)
from PyQt6.QtCore import Qt, QTimer, QRectF, QPointF, QMarginsF, QEvent, pyqtSignal
from PyQt6.QtPrintSupport import QPrinter

from constants import TOOLS, PROJECT_TYPES, SUPERVISION_RATES
from database import setup_database
from rule_engine import DynamicRuleEngine
from ui_components import InteractiveView, DraggableLabel
from canvas_objects import SmartPole, SmartStructure, SmartSpan, SmartConsumer
from ui_dialogs import (
    SearchDialog, SettingsDialog, DatabaseManagerDialog,
    RulesetManagerDialog, ProjectSetupDialog
)


# ─────────────────────────────────────────────────────────────────────────────
#  RESOURCE PATH HELPER (PyInstaller-compatible)
# ─────────────────────────────────────────────────────────────────────────────
def resource_path(relative_path):
    """Return absolute path to a bundled resource, works for dev and PyInstaller."""
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)


# ─────────────────────────────────────────────────────────────────────────────
#  PROJECT META DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_PROJECT_META = {
    "subject":          "",
    "lat":              "",
    "long":             "",
    "project_type":     "NSC",
    "use_uh":           False,
    "supervision_rate": 0.10,
}


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN APPLICATION CLASS
# ─────────────────────────────────────────────────────────────────────────────
class EstimateApp(QMainWindow):
    refresh_signal = pyqtSignal()

    def __init__(self):
        super().__init__()

        # Expiry guard
        if date.today() >= date(2026, 4, 30):
            QMessageBox.critical(
                self, "Application Expired",
                "This version has expired. Please obtain the latest release."
            )
            sys.exit()

        setup_database()

        # ── Project-level state ────────────────────────────────────────────
        self.project_meta   = dict(DEFAULT_PROJECT_META)
        self.bom_overrides  = {}
        self.live_bom_data  = []
        self.escalations    = []
        self.detail_view    = True          # show stay/earth/CG symbols
        self.span_start_pole = None
        self.last_placed_node = None        # for auto-span chain when placing nodes
        self.autosave_file  = "autosave_erp.json"
        self.current_tool   = "SELECT"

        # ── Page grid state ────────────────────────────────────────────────
        # 17.5 scene units ≈ 1 real-world metre  (calibrated: 40m span = ~700 units)
        self.pdf_scale  = 200   # default print scale
        self.pdf_show_project_name = True
        self.pdf_show_legend = True
        self.pdf_orientation_mode = "Auto + Overrides"
        self.pdf_auto_gain_threshold = 1.08
        self.pdf_page_overrides: dict[int, str] = {}
        self.show_page_grid  = True
        self.show_crosshatch = True

        # Rule engine (lazy-init on first refresh)
        self.rule_engine = DynamicRuleEngine()

        # ── History state (Undo/Redo) ──────────────────────────────────────
        self.history = []
        self.history_index = -1
        self._is_undoing = False
        self._refreshing_live = False
        self._history_timer = QTimer(self)
        self._history_timer.setSingleShot(True)
        self._history_timer.timeout.connect(self.push_history)

        # ── Build UI ───────────────────────────────────────────────────────
        self.setWindowTitle("ERP Estimate Generator — v5.0")
        self.setGeometry(50, 50, 1650, 930)
        logo_path = resource_path("logo.svg")
        if os.path.exists(logo_path):
            self.setWindowIcon(QIcon(logo_path))
        self._build_menu_bar()
        self._build_ui()

        app_inst = QApplication.instance()
        if app_inst is not None:
            app_inst.installEventFilter(self)

        # ── Wire signals ───────────────────────────────────────────────────
        self.refresh_signal.connect(self.refresh_live_estimate)
        self.scene.selectionChanged.connect(self.on_selection_changed)

        # ── Load autosave ──────────────────────────────────────────────────
        self.set_tool("SELECT")
        self.load_autosave()
        self.on_selection_changed()

    # =========================================================================
    #  MENU BAR
    # =========================================================================

    def _build_menu_bar(self):
        mb = self.menuBar()
        assert mb is not None
        mb.setStyleSheet(
            "QMenuBar { background:#f5f5f5; font-size:12px; }"
            "QMenuBar::item:selected { background:#d0d0d0; }"
            "QMenu { font-size:12px; }"
            "QMenu::item:selected { background:#3498db; color:white; }"
        )

        # ── File ──────────────────────────────────────────────────────────
        file_menu = mb.addMenu("&File")
        assert file_menu is not None

        act_new = QAction("📄  New Drawing", self)
        act_new.setShortcut(QKeySequence("Ctrl+N"))
        act_new.triggered.connect(self.new_drawing)
        file_menu.addAction(act_new)

        act_open = QAction("📂  Open…", self)
        act_open.setShortcut(QKeySequence("Ctrl+O"))
        act_open.triggered.connect(self.load_from_file)
        file_menu.addAction(act_open)

        act_save = QAction("💾  Save…", self)
        act_save.setShortcut(QKeySequence("Ctrl+S"))
        act_save.triggered.connect(self.save_to_file)
        file_menu.addAction(act_save)

        file_menu.addSeparator()

        self.act_undo = QAction("↶ Undo", self)
        self.act_undo.setShortcut(QKeySequence("Ctrl+Z"))
        self.act_undo.triggered.connect(self.undo)
        file_menu.addAction(self.act_undo)

        self.act_redo = QAction("↷ Redo", self)
        self.act_redo.setShortcut(QKeySequence("Ctrl+Y"))
        self.act_redo.triggered.connect(self.redo)
        file_menu.addAction(self.act_redo)

        file_menu.addSeparator()

        act_exit = QAction("Exit", self)
        act_exit.setShortcut(QKeySequence("Ctrl+Q"))
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        # ── Export ────────────────────────────────────────────────────────
        export_menu = mb.addMenu("E&xport")
        assert export_menu is not None

        act_pdf = QAction("🗺️  Export PDF Drawing", self)
        act_pdf.triggered.connect(self.export_pdf)
        export_menu.addAction(act_pdf)

        act_xl = QAction("📊  Generate Excel Estimate", self)
        act_xl.triggered.connect(self.generate_excel)
        export_menu.addAction(act_xl)

        # ── Settings ─────────────────────────────────────────────────────
        settings_menu = mb.addMenu("&Settings")
        assert settings_menu is not None

        act_proj = QAction("🗂  Project Settings", self)
        act_proj.triggered.connect(lambda: self._run_project_wizard(first_run=False))
        settings_menu.addAction(act_proj)

        settings_menu.addSeparator()

        act_db = QAction("🗃️  Master Database (Excel Sync)", self)
        act_db.triggered.connect(self.open_db_manager)
        settings_menu.addAction(act_db)

        act_rules = QAction("🧠  Ruleset Manager", self)
        act_rules.triggered.connect(self.open_rule_manager)
        settings_menu.addAction(act_rules)

        # ── Help ──────────────────────────────────────────────────────────
        help_menu = mb.addMenu("&Help")
        assert help_menu is not None

        act_help = QAction("📖  User Guide", self)
        act_help.setShortcut(QKeySequence("F1"))
        act_help.triggered.connect(self.show_help)
        help_menu.addAction(act_help)

        help_menu.addSeparator()

        act_credits = QAction("🏆  Credits", self)
        act_credits.triggered.connect(self.show_credits)
        help_menu.addAction(act_credits)

        act_about = QAction("ℹ️  About", self)
        act_about.triggered.connect(self.show_about_dialog)
        help_menu.addAction(act_about)

    # =========================================================================
    #  UI CONSTRUCTION
    # =========================================================================

    def _build_ui(self):
        central = QWidget()
        root_layout = QHBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        self.setCentralWidget(central)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        root_layout.addWidget(self.splitter)

        # Left: canvas area
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(4, 4, 4, 4)
        left_layout.setSpacing(4)
        self.splitter.addWidget(left_panel)

        left_layout.addLayout(self._build_icon_ribbon())
        left_layout.addLayout(self._build_draw_toolbar())

        self.scene = QGraphicsScene()
        self.view  = InteractiveView(self.scene, self)
        left_layout.addWidget(self.view)

        # ── Bottom canvas control bar ──────────────────────────────────────
        bottom_bar = QHBoxLayout()
        bottom_bar.setSpacing(8)
        bottom_bar.setContentsMargins(2, 0, 2, 0)

        # Show Symbols checkbox
        self.detail_chk = QCheckBox("Show Symbols")
        self.detail_chk.setChecked(True)
        self.detail_chk.setStyleSheet(
            "font-size:11px; font-weight:bold; color:#555; spacing:4px;"
        )
        self.detail_chk.toggled.connect(self._toggle_detail_view)
        bottom_bar.addWidget(self.detail_chk)

        # Separator
        sep1 = QLabel("|");
        sep1.setStyleSheet("color:#ccc; font-size:14px;")
        bottom_bar.addWidget(sep1)

        # Page Grid toggle
        self.grid_chk = QCheckBox("Page Grid")
        self.grid_chk.setChecked(True)
        self.grid_chk.setStyleSheet(
            "font-size:11px; font-weight:bold; color:#3a7bd5; spacing:4px;"
        )
        self.grid_chk.toggled.connect(self._toggle_page_grid)
        bottom_bar.addWidget(self.grid_chk)

        # Crosshatch toggle
        self.hatch_chk = QCheckBox("Crosshatch")
        self.hatch_chk.setChecked(True)
        self.hatch_chk.setStyleSheet(
            "font-size:11px; color:#555; spacing:4px;"
        )
        self.hatch_chk.toggled.connect(self._toggle_crosshatch)
        bottom_bar.addWidget(self.hatch_chk)

        # Separator
        sep2 = QLabel("|");
        sep2.setStyleSheet("color:#ccc; font-size:14px;")
        bottom_bar.addWidget(sep2)

        # Scale label + dropdown
        bottom_bar.addWidget(QLabel("Print Scale:"))
        self.scale_cb = QComboBox()
        self.scale_cb.addItems([
            "1:150", "1:200", "1:300"
        ])
        self.scale_cb.setCurrentText("1:200")
        self.scale_cb.setToolTip(
            "Sets how much drawing area fits on one A4 page.\n"
            "1:150 = more detailed  |  1:300 = wider area."
        )
        self.scale_cb.currentTextChanged.connect(self._on_scale_changed)
        self.scale_cb.setStyleSheet(
            "font-size:11px; font-weight:bold; color:#1a5276;"
        )
        bottom_bar.addWidget(self.scale_cb)

        # Orientation mode
        bottom_bar.addWidget(QLabel("Orientation:"))
        self.orient_cb = QComboBox()
        self.orient_cb.addItems([
            "Landscape (All)",
            "Portrait (All)",
            "Auto (Global Best)",
            "Auto + Overrides",
        ])
        self.orient_cb.setCurrentText(self.pdf_orientation_mode)
        self.orient_cb.setToolTip(
            "Landscape/Portrait force all pages.\n"
            "Auto (Global Best) picks one best orientation for the full drawing.\n"
            "Auto + Overrides allows manual page overrides."
        )
        self.orient_cb.currentTextChanged.connect(self._on_orientation_mode_changed)
        self.orient_cb.setStyleSheet("font-size:11px; color:#1a5276;")
        bottom_bar.addWidget(self.orient_cb)

        self.page_override_btn = QPushButton("Page Overrides")
        self.page_override_btn.setToolTip(
            "Set manual orientation for specific pages, e.g. 2:P, 5:L.\n"
            "Note: overrides that conflict with current grid geometry are ignored."
        )
        self.page_override_btn.setStyleSheet(
            "font-size:10px; padding:3px 6px;"
        )
        self.page_override_btn.clicked.connect(self._edit_page_overrides)
        bottom_bar.addWidget(self.page_override_btn)

        # Separator
        sep3 = QLabel("|")
        sep3.setStyleSheet("color:#ccc; font-size:14px;")
        bottom_bar.addWidget(sep3)

        # PDF title toggle
        self.pdf_title_chk = QCheckBox("Project Name")
        self.pdf_title_chk.setChecked(True)
        self.pdf_title_chk.setToolTip("Show project name in PDF drawing title strip")
        self.pdf_title_chk.setStyleSheet(
            "font-size:11px; color:#555; spacing:4px;"
        )
        self.pdf_title_chk.toggled.connect(self._toggle_pdf_project_name)
        bottom_bar.addWidget(self.pdf_title_chk)

        # PDF legend toggle
        self.pdf_legend_chk = QCheckBox("Legend")
        self.pdf_legend_chk.setChecked(True)
        self.pdf_legend_chk.setToolTip("Show legend on the last PDF drawing page")
        self.pdf_legend_chk.setStyleSheet(
            "font-size:11px; color:#555; spacing:4px;"
        )
        self.pdf_legend_chk.toggled.connect(self._toggle_pdf_legend)
        bottom_bar.addWidget(self.pdf_legend_chk)

        bottom_bar.addStretch()
        left_layout.addLayout(bottom_bar)

        # Right: properties + estimate table
        right_splitter = QSplitter(Qt.Orientation.Vertical)
        self.splitter.addWidget(right_splitter)
        self.splitter.setSizes([1000, 650])

        right_splitter.addWidget(self._build_properties_panel())
        right_splitter.addWidget(self._build_estimate_panel())
        right_splitter.setSizes([320, 680])

    def _build_icon_ribbon(self):
        """Small icon-only ribbon above the main drawing tools."""
        bar = QHBoxLayout()
        bar.setSpacing(6)
        bar.setContentsMargins(0, 0, 0, 2)
        
        def _make_icon_btn(text, tooltip, bg, color="#000"):
            btn = QPushButton(text)
            btn.setToolTip(tooltip)
            btn.setFixedSize(28, 28)
            btn.setStyleSheet(f"font-size:16px; font-weight:bold; background:{bg}; color:{color}; border-radius:3px; border:1px solid #ccc;")
            return btn

        # Undo / Redo
        self.undo_btn = _make_icon_btn("↶", "Undo (Ctrl+Z)", "#e6e6e6")
        self.undo_btn.clicked.connect(self.undo)
        bar.addWidget(self.undo_btn)
        
        self.redo_btn = _make_icon_btn("↷", "Redo (Ctrl+Y)", "#e6e6e6")
        self.redo_btn.clicked.connect(self.redo)
        bar.addWidget(self.redo_btn)
        
        sep = QLabel("|")
        sep.setStyleSheet("color:#bbb; padding:0 2px;")
        bar.addWidget(sep)
        
        # PDF / Excel
        btn_pdf = _make_icon_btn("📑", "Export PDF Drawing", "#f9ebea", "#78281f")
        btn_pdf.clicked.connect(self.export_pdf)
        bar.addWidget(btn_pdf)
        
        btn_xl = _make_icon_btn("📊", "Export Excel Estimate", "#eaf2f8", "#154360")
        btn_xl.clicked.connect(self.generate_excel)
        bar.addWidget(btn_xl)

        bar.addStretch()
        return bar

    def _build_draw_toolbar(self):
        bar = QHBoxLayout()
        bar.setSpacing(3)
        self.tools_btns = {}
        for key, txt in TOOLS.items():
            btn = QPushButton(txt)
            btn.clicked.connect(lambda checked, t=key: self.set_tool(t))
            btn.setStyleSheet(
                "padding:7px 5px; font-weight:bold; background:lightgray;"
            )
            bar.addWidget(btn)
            self.tools_btns[key] = btn

        # Thin visual separator
        sep = QLabel("|")
        sep.setStyleSheet("color:#bbb; font-size:16px; padding:0 4px;")
        bar.addWidget(sep)

        # Fit-View button — also triggered by F key on the canvas
        fit_btn = QPushButton("⬡ Fit View")
        fit_btn.setToolTip(
            "Fit all drawing content in view  [F or Ctrl+0]\n"
            "Useful after zooming out too far or after loading a project."
        )
        fit_btn.setStyleSheet(
            "padding:7px 10px; font-weight:bold;"
            "background:#d5e8f7; color:#1a5276; border-radius:3px;"
        )
        fit_btn.clicked.connect(self._fit_view)
        bar.addWidget(fit_btn)
        return bar

    def _build_properties_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(6, 6, 6, 0)
        lay.setSpacing(4)

        # Project info strip with edit button
        info_row = QHBoxLayout()
        info_row.setSpacing(0)
        self.proj_info_label = QLabel()
        self.proj_info_label.setWordWrap(True)
        self.proj_info_label.setMinimumWidth(0)
        self.proj_info_label.setStyleSheet(
            "font-size:11px; color:#555; padding:3px 5px;"
            "background:#f0f0f0; border-radius:3px 0 0 3px;"
        )
        info_row.addWidget(self.proj_info_label, 1)

        edit_btn = QPushButton("✏️")
        edit_btn.setToolTip("Edit Project Settings")
        edit_btn.setFixedSize(28, 24)
        edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        edit_btn.setStyleSheet(
            "QPushButton { background:#f0f0f0; border:1px solid #ccc;"
            "  border-left:none; border-radius:0 3px 3px 0; font-size:13px; }"
            "QPushButton:hover { background:#d5e8f7; }"
        )
        edit_btn.clicked.connect(lambda: self._run_project_wizard(first_run=False))
        info_row.addWidget(edit_btn)
        lay.addLayout(info_row)
        self._refresh_proj_label()

        # Object property editor
        self.editor_group = QGroupBox("Object Properties")
        self.editor_layout = QFormLayout()
        self.editor_layout.setSpacing(5)
        self.editor_group.setLayout(self.editor_layout)

        scroll = QScrollArea()
        scroll.setWidget(self.editor_group)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        lay.addWidget(scroll)

        return w

    def _build_estimate_panel(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(6, 0, 6, 6)
        lay.setSpacing(4)

        lay.addWidget(QLabel("<b>Live Estimate</b> (double-click Qty to edit)"))

        self.live_table = QTableWidget(0, 6)
        self.live_table.setHorizontalHeaderLabels(
            ["Type", "Code", "Name", "Qty", "Unit", "Total (Rs)"]
        )
        live_hdr = self.live_table.horizontalHeader()
        assert live_hdr is not None
        live_hdr.setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self.live_table.setColumnWidth(0, 65)
        self.live_table.setColumnWidth(1, 85)
        self.live_table.setColumnWidth(3, 65)
        self.live_table.itemChanged.connect(self.on_table_edit)
        lay.addWidget(self.live_table)

        # Custom item buttons
        btn_row = QHBoxLayout()
        add_mat = QPushButton("+ Add Material")
        add_lab = QPushButton("+ Add Labor")
        add_mat.clicked.connect(lambda: self.open_search("Material"))
        add_lab.clicked.connect(lambda: self.open_search("Labor"))
        add_mat.setStyleSheet(
            "background:#3498db; color:white; font-weight:bold; padding:5px;"
        )
        add_lab.setStyleSheet(
            "background:#e67e22; color:white; font-weight:bold; padding:5px;"
        )
        btn_row.addWidget(add_mat)
        btn_row.addWidget(add_lab)
        lay.addLayout(btn_row)

        self.grand_total_label = QLabel("<b>Grand Total: Rs. 0.00</b>")
        self.grand_total_label.setStyleSheet(
            "font-size:15px; color:#d32f2f; margin-top:4px;"
        )
        lay.addWidget(self.grand_total_label)

        return w

    # =========================================================================
    #  PROJECT WIZARD
    # =========================================================================

    def _run_project_wizard(self, first_run=False):
        dlg = ProjectSetupDialog(self.project_meta, self, first_run=first_run)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.project_meta = dlg.get_meta()
            self._refresh_proj_label()
            self.refresh_live_estimate()

    def _refresh_proj_label(self):
        m = self.project_meta
        sup_pct = int(m.get("supervision_rate", 0.10) * 100)
        uh_txt  = "UH Materials" if m.get("use_uh") else "Raw Steel"
        self.proj_info_label.setText(
            f"📌 {m.get('subject','(no subject)')}   |   "
            f"Type: {m.get('project_type','NSC')}   |   "
            f"Sup: {sup_pct}%   |   "
            f"Materials: {uh_txt}"
        )

    # =========================================================================
    #  TOOL MANAGEMENT
    # =========================================================================

    # Drawing tools — switching between these keeps the auto-span chain alive
    _DRAWING_TOOLS = frozenset({"ADD_LT", "ADD_HT", "ADD_EXISTING", "ADD_STRUCTURE", "ADD_CONSUMER"})

    def set_tool(self, tool_name):
        prev_tool = self.current_tool
        self.current_tool = tool_name
        if self.span_start_pole:
            self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1))
        self.span_start_pole = None
        # Clear auto-connect chain ONLY when leaving drawing mode
        # (i.e. switching to SELECT or ADD_SPAN — not between placement tools)
        leaving_drawing = tool_name not in self._DRAWING_TOOLS
        if leaving_drawing and self.last_placed_node is not None:
            try:
                self.last_placed_node.setPen(QPen(Qt.GlobalColor.black, 1))
            except RuntimeError:
                pass
            self.last_placed_node = None
        for key, btn in self.tools_btns.items():
            active = key == tool_name
            btn.setStyleSheet(
                "padding:7px 5px; font-weight:bold; background:"
                + ("lightblue;" if active else "lightgray;")
            )
        self.update_view_drag_mode()

    def update_view_drag_mode(self):
        """Delegate canvas interaction state to InteractiveView.

        SELECT mode now uses dynamic drag mode/cursor behavior:
        empty-space pan, hover-select pointer, selected-object drag pointer.
        """
        if self.current_tool == "SELECT":
            self.view.refresh_interaction_state()
        else:
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag)
            self.view.setCursor(Qt.CursorShape.ArrowCursor)

    def _fit_view(self):
        """Fit all drawing content in view. Called by toolbar button and F key."""
        bounds = self.scene.itemsBoundingRect()
        if not bounds.isNull():
            self.view.fitInView(
                bounds.adjusted(-60, -60, 60, 60),
                Qt.AspectRatioMode.KeepAspectRatio
            )

    def _toggle_detail_view(self, checked=None):
        self.detail_view = self.detail_chk.isChecked()
        # Redraw all canvas items
        for item in self.scene.items():
            if isinstance(item, (SmartPole, SmartStructure, SmartSpan, SmartConsumer)):
                item.detail_view = self.detail_view
                item.update_visuals()

    def _toggle_page_grid(self, checked):
        self.show_page_grid = checked
        self.view.grid_show = checked
        vp = self.view.viewport()
        assert vp is not None
        vp.update()

    def _toggle_crosshatch(self, checked):
        self.show_crosshatch = checked
        self.view.grid_crosshatch = checked
        vp = self.view.viewport()
        assert vp is not None
        vp.update()

    def _toggle_pdf_project_name(self, checked):
        self.pdf_show_project_name = checked

    def _toggle_pdf_legend(self, checked):
        self.pdf_show_legend = checked

    def _on_scale_changed(self, text):
        """Called when user picks a new print scale from the dropdown."""
        try:
            self.pdf_scale = int(text.split(":")[1])
        except (IndexError, ValueError):
            self.pdf_scale = 200
        self._refresh_page_grid()

    def _on_orientation_mode_changed(self, text):
        self.pdf_orientation_mode = text
        use_override = (text == "Auto + Overrides")
        self.page_override_btn.setEnabled(use_override)
        self._refresh_page_grid()

    def _edit_page_overrides(self):
        if self.pdf_orientation_mode != "Auto + Overrides":
            QMessageBox.information(
                self,
                "Orientation Overrides",
                "Switch Orientation mode to 'Auto + Overrides' to edit page overrides.",
            )
            return

        self._refresh_page_grid()
        tiles = self.view.grid_tiles
        if not tiles:
            QMessageBox.information(self, "Orientation Overrides", "No pages available.")
            return

        total_pages = tiles[0].get("total", len(tiles))
        cur_parts = [f"{k}:{v}" for k, v in sorted(self.pdf_page_overrides.items())]
        cur_text = ", ".join(cur_parts)

        text, ok = QInputDialog.getText(
            self,
            "Page Orientation Overrides",
            (
                f"Enter overrides as page:orientation (L or P).\n"
                f"Example: 2:P, 5:L\n"
                f"Pages available: 1..{total_pages}\n"
                f"Conflicting overrides may be ignored to keep non-overlapping page grid."
            ),
            text=cur_text,
        )
        if not ok:
            return

        raw = text.strip()
        if not raw:
            self.pdf_page_overrides = {}
            self._refresh_page_grid()
            return

        parsed: dict[int, str] = {}
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for p in parts:
            if ":" not in p:
                QMessageBox.warning(self, "Invalid Override", f"Invalid token: {p}")
                return
            page_str, orient_str = p.split(":", 1)
            try:
                page_no = int(page_str.strip())
            except ValueError:
                QMessageBox.warning(self, "Invalid Override", f"Invalid page number: {page_str}")
                return
            orient = orient_str.strip().upper()
            if orient not in ("L", "P"):
                QMessageBox.warning(self, "Invalid Override", f"Invalid orientation for page {page_no}: {orient_str}")
                return
            if page_no < 1 or page_no > total_pages:
                QMessageBox.warning(
                    self,
                    "Invalid Override",
                    f"Page {page_no} is out of range (1..{total_pages}).",
                )
                return
            parsed[page_no] = orient

        self.pdf_page_overrides = parsed
        self._refresh_page_grid()

    # =========================================================================
    #  PAGE GRID COMPUTATION
    # =========================================================================

    # Physical A4 dimensions in millimetres
    _A4_W_MM = 297.0   # landscape width
    _A4_H_MM = 210.0   # landscape height
    # Calibration: 40m real span ≈ 700 scene units → 17.5 units per metre
    _SCENE_UNITS_PER_M = 17.5

    def _a4_scene_dims(self, scale):
        """
        Return (page_w_scene, page_h_scene) — size of an A4 landscape sheet
        in scene units at the given 1:scale ratio.
        """
        m_per_mm = scale / 1000.0          # 1mm paper → this many real metres
        w = self._A4_W_MM * m_per_mm * self._SCENE_UNITS_PER_M
        h = self._A4_H_MM * m_per_mm * self._SCENE_UNITS_PER_M
        return w, h

    def _a4_scene_dims_oriented(self, scale, orient: str):
        w_l, h_l = self._a4_scene_dims(scale)
        if orient == "P":
            return h_l, w_l
        return w_l, h_l

    def _auto_tile_orientation(self, union_rect: QRectF, pad: float = 10.0) -> str:
        """Pick L/P by comparing best fit scale of content in both orientations."""
        if union_rect.isNull() or union_rect.isEmpty():
            return "L"

        content_w = max(1.0, union_rect.width())
        content_h = max(1.0, union_rect.height())

        pw_l, ph_l = self._a4_scene_dims_oriented(self.pdf_scale, "L")
        pw_p, ph_p = self._a4_scene_dims_oriented(self.pdf_scale, "P")

        s_l = min((pw_l - 2 * pad) / content_w, (ph_l - 2 * pad) / content_h)
        s_p = min((pw_p - 2 * pad) / content_w, (ph_p - 2 * pad) / content_h)

        if s_p > s_l * self.pdf_auto_gain_threshold:
            return "P"
        return "L"

    def _auto_global_orientation(self, bounds: QRectF) -> str:
        """Pick one stable orientation for the whole drawing bounds."""
        return self._auto_tile_orientation(bounds)

    def _resolve_orientation(self, auto_orient: str, page_num: int) -> tuple[str, bool]:
        mode = self.pdf_orientation_mode
        if mode == "Landscape (All)":
            return "L", False
        if mode == "Portrait (All)":
            return "P", False
        if mode == "Auto + Overrides":
            ov = self.pdf_page_overrides.get(page_num)
            if ov in ("L", "P"):
                return ov, True
            return auto_orient, False
        return auto_orient, False

    def _refresh_page_grid(self):
        """
        Recompute the A4 page tiles in scene coordinates and push them
        to the InteractiveView for background rendering.
        This uses a Sparse Grid algorithm -- only pages containing items are generated.
        Tile positions are computed on a fixed landscape grid for stable row/col order,
        while each tile is tagged L/P individually based on local drawing geometry.
        Called after any placement, deletion, scale change, or load.
        """
        items = [
            i for i in self.scene.items()
            if isinstance(i, (SmartPole, SmartStructure, SmartConsumer, SmartSpan))
        ]
        if not items:
            if self.pdf_orientation_mode == "Portrait (All)":
                blank_orient = "P"
            else:
                blank_orient = "L"
            pw, ph = self._a4_scene_dims_oriented(self.pdf_scale, blank_orient)
            blank_rect = QRectF(-pw / 2.0, -ph / 2.0, pw, ph)
            self.view.grid_tiles = [{
                "rect": blank_rect,
                "orient": blank_orient,
                "auto_orient": blank_orient,
                "is_override": False,
                "items_count": 0,
                "row": 0,
                "col": 0,
                "page_num": 1,
                "total": 1,
            }]
            margin = max(pw, ph)
            self.scene.setSceneRect(blank_rect.adjusted(-margin, -margin, margin, margin))
            vp = self.view.viewport()
            assert vp is not None
            vp.update()
            return

        PAD = 50  # Increased padding for better object spacing
        bounds = self.scene.itemsBoundingRect().adjusted(-PAD, -PAD, PAD, PAD)

        if self.pdf_orientation_mode == "Landscape (All)":
            base_orient = "L"
        elif self.pdf_orientation_mode == "Portrait (All)":
            base_orient = "P"
        else:
            # Use global auto orientation for stable non-overlapping page grid.
            base_orient = self._auto_global_orientation(bounds)
        pw_L, ph_L = self._a4_scene_dims_oriented(self.pdf_scale, base_orient)

        # Edge margin: Minimum space between objects and page boundary (scene units)
        # This ensures objects don't appear at the very edge of the page
        EDGE_MARGIN = 30
        
        # Keep page 1 centered while the drawing still fits on a single page with safety margins.
        # Only switch to top-left packed tiling when overflow requires more pages.
        # Check if bounds fit within the page with additional edge protection
        if (bounds.width() <= pw_L - 2 * EDGE_MARGIN and 
            bounds.height() <= ph_L - 2 * EDGE_MARGIN):
            single_rect = QRectF(-pw_L / 2.0, -ph_L / 2.0, pw_L, ph_L)
            
            # Ensure the centered page actually contains all items with margin
            centered_bounds_check = bounds.adjusted(0, 0, 0, 0)  # Use bounds as-is
            page_left = single_rect.left()
            page_right = single_rect.right()
            page_top = single_rect.top()
            page_bottom = single_rect.bottom()
            
            # Verify bounds fit within centered page with margins
            if (centered_bounds_check.left() >= page_left + EDGE_MARGIN and
                centered_bounds_check.right() <= page_right - EDGE_MARGIN and
                centered_bounds_check.top() >= page_top + EDGE_MARGIN and
                centered_bounds_check.bottom() <= page_bottom - EDGE_MARGIN):
                
                self.view.grid_tiles = [{
                    "rect": single_rect,
                    "auto_orient": base_orient,
                    "orient": base_orient,
                    "is_override": False,
                    "items_count": len(items),
                    "row": 0,
                    "col": 0,
                    "page_num": 1,
                    "total": 1,
                }]
                margin = max(pw_L, ph_L)
                self.scene.setSceneRect(single_rect.adjusted(-margin, -margin, margin, margin))
                vp = self.view.viewport()
                assert vp is not None
                vp.update()
                return

        # A tile is occupied when any drawable item's geometry touches it.
        def items_in(rect):
            return [i for i in items if rect.intersects(i.sceneBoundingRect())]

        # Pack pages from the drawing bounds origin so page-1 starts where
        # content starts, instead of snapping to world-origin multiples.
        eps = 1e-6
        cols = max(1, int(math.ceil(max(0.0, bounds.width() - eps) / pw_L)))
        rows = max(1, int(math.ceil(max(0.0, bounds.height() - eps) / ph_L)))
        base_left = bounds.left()
        base_top = bounds.top()

        occupied_tiles = []
        # Generate occupied tiles strictly top-to-bottom, left-to-right
        for r in range(rows):
            for c in range(cols):
                rect = QRectF(base_left + c * pw_L, base_top + r * ph_L, pw_L, ph_L)
                ins = items_in(rect)
                if not ins:
                    continue

                # Local orientation based on item spread inside this tile.
                local_union = None
                for item in ins:
                    inter = rect.intersected(item.sceneBoundingRect())
                    if inter.isNull() or inter.isEmpty():
                        continue
                    local_union = inter if local_union is None else local_union.united(inter)

                occupied_tiles.append({
                    "rect": rect,
                    "auto_orient": base_orient,
                    "items_count": len(ins),
                    "row": r,
                    "col": c,
                })

        # Number occupied tiles sequentially 1..N with no gaps.
        total = len(occupied_tiles)
        final_tiles = []
        for i, t in enumerate(occupied_tiles):
            page_num = i + 1
            orient, is_override = self._resolve_orientation(t.get("auto_orient", base_orient), page_num)

            # Keep page grid geometry stable/non-overlapping. If an override
            # disagrees with the global geometry orientation, keep geometry
            # orientation and drop override for this layout pass.
            if orient != base_orient:
                orient = base_orient
                is_override = False

            t["page_num"] = page_num
            t["total"] = total
            t["orient"] = orient
            t["is_override"] = is_override
            final_tiles.append(t)

        # Drop overrides that no longer map to current page count.
        if self.pdf_page_overrides:
            self.pdf_page_overrides = {
                k: v for k, v in self.pdf_page_overrides.items() if 1 <= k <= total
            }

        self.view.grid_tiles = final_tiles

        # Constrain the scene rect to exactly the generated tiles + margin
        if final_tiles:
            full_rect = final_tiles[0]["rect"]
            for t in final_tiles[1:]:
                full_rect = full_rect.united(t["rect"])
            
            w_land, h_land = self._a4_scene_dims_oriented(self.pdf_scale, "L")
            w_port, h_port = self._a4_scene_dims_oriented(self.pdf_scale, "P")
            margin = max(w_land, h_land, w_port, h_port)
            self.scene.setSceneRect(full_rect.adjusted(-margin, -margin, margin, margin))
        else:
             self.scene.setSceneRect(QRectF(-1500, -1500, 3000, 3000))

        vp = self.view.viewport()
        assert vp is not None
        vp.update()




    # =========================================================================
    #  CANVAS CLICK HANDLER
    # =========================================================================

    def handle_canvas_click(self, event, view):
        if event.button() == Qt.MouseButton.RightButton:
            self.set_tool("SELECT")
            return
        if self.current_tool == "SELECT":
            return

        pos = view.mapToScene(event.pos())
        item_at = self.scene.itemAt(pos, view.transform())

        # ── Pole placement ────────────────────────────────────────────────
        if self.current_tool in ("ADD_LT", "ADD_HT", "ADD_EXISTING"):
            p_type    = "LT" if self.current_tool in ("ADD_LT", "ADD_EXISTING") else "HT"
            is_exist  = self.current_tool == "ADD_EXISTING"
            pole = SmartPole(
                pos.x(), pos.y(), self.refresh_signal,
                p_type, is_exist,
                detail_view=self.detail_view
            )
            self.scene.addItem(pole)
            self._auto_connect_span(pole)
            self.refresh_live_estimate()

        # ── Structure placement ───────────────────────────────────────────
        elif self.current_tool == "ADD_STRUCTURE":
            struct = SmartStructure(
                pos.x(), pos.y(), self.refresh_signal,
                detail_view=self.detail_view
            )
            self.scene.addItem(struct)
            self._auto_connect_span(struct)
            self.refresh_live_estimate()

        # ── Consumer placement ────────────────────────────────────────────
        elif self.current_tool == "ADD_CONSUMER":
            consumer = SmartConsumer(
                pos.x(), pos.y(), self.refresh_signal,
                detail_view=self.detail_view
            )
            self.scene.addItem(consumer)
            self._auto_connect_span(consumer)
            self.refresh_live_estimate()

        # ── Span drawing ──────────────────────────────────────────────────
        elif self.current_tool == "ADD_SPAN":
            if not isinstance(item_at, (SmartPole, SmartStructure, SmartConsumer)):
                return
            if not self.span_start_pole:
                self.span_start_pole = item_at
                item_at.setPen(QPen(Qt.GlobalColor.yellow, 3))
            elif self.span_start_pole != item_at:
                # Warn on HT↔LT cross-connection
                p1, p2 = self.span_start_pole, item_at
                if (isinstance(p1, SmartPole) and isinstance(p2, SmartPole)):
                    eff1 = p1.existing_subtype if p1.is_existing else p1.pole_type
                    eff2 = p2.existing_subtype if p2.is_existing else p2.pole_type
                    if (eff1 == "HT") != (eff2 == "HT"):
                        ans = QMessageBox.question(
                            self, "Warning",
                            "Connect HT pole to LT pole?",
                            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                        )
                        if ans == QMessageBox.StandardButton.No:
                            return

                span = SmartSpan(p1, p2, detail_view=self.detail_view)
                p1.connected_spans.append(span)
                p2.connected_spans.append(span)
                self.scene.addItem(span)
                self.scene.addItem(span.label)
                self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1))
                self.span_start_pole = None
                self.refresh_live_estimate()

    # =========================================================================
    #  AUTO-CONNECT SPAN HELPER
    # =========================================================================

    # HT-class node types: includes HT poles and all existing structure subtypes
    _HT_SUBTYPES = frozenset({"HT", "DP", "TP", "4P", "DTR"})

    @staticmethod
    def _is_ht_node(node) -> bool:
        """
        Returns True if a node is effectively HT.
        SmartStructure is always HT.
        SmartPole: check existing_subtype (DP/TP/4P/DTR/HT) or pole_type.
        SmartConsumer: always LT.
        """
        if isinstance(node, SmartStructure):
            return True
        if isinstance(node, SmartPole):
            eff = node.existing_subtype if node.is_existing else node.pole_type
            return eff in EstimateApp._HT_SUBTYPES
        return False  # SmartConsumer = LT

    def _auto_connect_span(self, new_node):
        """
        Called every time a new pole / structure / consumer is placed.
        If a previous node exists (last_placed_node), automatically draw
        a span connecting it to the new node.
        The new node then becomes last_placed_node for the next placement.
        Right-click / switching to SELECT clears the chain.
        """
        prev = self.last_placed_node

        # Remove yellow highlight from previous node
        if prev is not None:
            try:
                prev.setPen(QPen(Qt.GlobalColor.black, 1))
            except RuntimeError:
                prev = None

        if prev is not None and prev is not new_node:
            p1, p2 = prev, new_node

            # HT ↔ LT cross-connection warning
            # DP/TP/4P/DTR existing poles are treated as HT
            if self._is_ht_node(p1) != self._is_ht_node(p2):
                ans = QMessageBox.question(
                    self, "Warning",
                    "Connect HT to LT node?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if ans == QMessageBox.StandardButton.No:
                    # Break chain — don't connect, but still update chain start
                    self.last_placed_node = new_node
                    new_node.setPen(QPen(Qt.GlobalColor.yellow, 3))
                    return

            span = SmartSpan(p1, p2, detail_view=self.detail_view)
            p1.connected_spans.append(span)
            p2.connected_spans.append(span)
            self.scene.addItem(span)
            self.scene.addItem(span.label)

        # New node becomes the chain anchor — highlighted yellow
        self.last_placed_node = new_node
        new_node.setPen(QPen(Qt.GlobalColor.yellow, 3))

    # =========================================================================
    #  SELECTION / PROPERTY EDITOR
    # =========================================================================

    def on_selection_changed(self):
        try:
            if not self.scene.views():
                return
        except RuntimeError:
            return

        # Clear editor
        while self.editor_layout.count():
            child = self.editor_layout.takeAt(0)
            if child is None:
                continue
            w = child.widget()
            if w is not None:
                w.deleteLater()

        sel = self.scene.selectedItems()
        if not sel:
            self.editor_group.setTitle("Canvas Shortcuts")
            self._build_empty_editor_hint()
            return
        if len(sel) > 1:
            self.editor_group.setTitle(f"{len(sel)} items selected")
            return

        item = sel[0]
        if isinstance(item, DraggableLabel):
            self.editor_group.setTitle("Text label")
            return

        if isinstance(item, SmartPole):
            self._build_pole_editor(item)
        elif isinstance(item, SmartStructure):
            self._build_structure_editor(item)
        elif isinstance(item, SmartSpan):
            self._build_span_editor(item)
        elif isinstance(item, SmartConsumer):
            self._build_consumer_editor(item)

    def _build_empty_editor_hint(self):
        hint = QLabel(
            "<b>Mouse</b><br>"
            "• Left drag on blank space: Pan canvas<br>"
            "• Shift + Left drag: Rubber-band multi-select<br>"
            "• Hover on object: Select cursor<br>"
            "• Middle drag / Ctrl + Left drag: Pan canvas<br>"
            "• Wheel: Zoom, Ctrl + Wheel: Fine zoom<br><br>"
            "<b>Keyboard</b><br>"
            "• Esc: Select tool<br>"
            "• F or Ctrl+0: Fit drawing to view<br>"
            "• Ctrl+A: Select all objects"
        )
        hint.setWordWrap(True)
        hint.setStyleSheet(
            "QLabel {"
            "  background:#f7fbff;"
            "  border:1px solid #d7e8f6;"
            "  border-radius:6px;"
            "  padding:10px;"
            "  color:#2b3d4f;"
            "  line-height:1.35;"
            "}"
        )
        self.editor_layout.addRow(hint)

    # ── Pole editor ───────────────────────────────────────────────────────────

    def _build_pole_editor(self, item):
        subtype = getattr(item, "existing_subtype", item.pole_type)
        if item.is_existing:
            self.editor_group.setTitle(f"Existing — {subtype}")
        else:
            self.editor_group.setTitle(f"{item.pole_type} Pole")

        # Existing subtype picker — only shown for existing poles
        if item.is_existing:
            type_cb = QComboBox()
            type_cb.addItems(["LT", "HT", "DP", "TP", "4P", "DTR"])
            type_cb.setCurrentText(subtype)
            type_cb.currentTextChanged.connect(
                lambda t, i=item: self._update_existing_subtype(i, t)
            )
            self.editor_layout.addRow("Existing Type:", type_cb)

            if subtype == "DTR":
                dtr_cb = QComboBox()
                dtr_cb.addItems(
                    ["None", "10KVA", "16KVA", "25KVA", "63KVA", "100KVA", "160KVA"]
                )
                dtr_cb.setCurrentText(getattr(item, "existing_dtr_size", "None"))
                dtr_cb.currentTextChanged.connect(
                    lambda t, i=item: self._update_pole(i, "existing_dtr_size", t)
                )
                self.editor_layout.addRow("DTR Size:", dtr_cb)

        # Pole type 2 (material)
        pt2_cb = QComboBox()
        pt2_cb.addItems(["PCC", "STP", "H-BEAM"])
        pt2_cb.setCurrentText(item.pole_type2)
        pt2_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_pole_type2(i, t)
        )
        self.editor_layout.addRow("Material:", pt2_cb)

        # Height (cascading)
        ht_cb = QComboBox()
        ht_cb.addItems(self._height_options(item.pole_type2))
        ht_cb.setCurrentText(item.height)
        ht_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_pole(i, "height", t)
        )
        self.editor_layout.addRow("Height:", ht_cb)

        # Extension
        ext_chk = QCheckBox("Extension required")
        ext_chk.setChecked(item.has_extension)
        ext_chk.stateChanged.connect(
            lambda v, i=item: self._toggle_pole_extension(i, v == 2)
        )
        self.editor_layout.addRow(ext_chk)

        if item.has_extension:
            ext_ht = QDoubleSpinBox()
            ext_ht.setRange(1.0, 10.0)
            ext_ht.setSingleStep(0.5)
            ext_ht.setSuffix(" m")
            ext_ht.setValue(item.extension_height)
            ext_ht.valueChanged.connect(
                lambda v, i=item: self._update_pole(i, "extension_height", v)
            )
            self.editor_layout.addRow("Ext. Height:", ext_ht)

        # Earth count
        earth_sp = QSpinBox()
        earth_sp.setRange(0, 10)
        earth_sp.setValue(item.earth_count)
        earth_sp.valueChanged.connect(
            lambda v, i=item: self._update_pole(i, "earth_count", v)
        )
        self.editor_layout.addRow("Earthing Sets:", earth_sp)

        # Stay count + override indicator
        stay_row = QHBoxLayout()
        stay_sp = QSpinBox()
        stay_sp.setRange(0, 10)
        stay_sp.setValue(item.stay_count)
        stay_sp.valueChanged.connect(
            lambda v, i=item: self._manual_stay(i, v)
        )
        stay_row.addWidget(stay_sp)
        if item.override_auto_stay:
            lock_lbl = QLabel("🔒 Manual")
            lock_lbl.setStyleSheet("color:#e67e22; font-size:10px;")
            stay_row.addWidget(lock_lbl)
            reset_btn = QPushButton("Reset")
            reset_btn.setFixedWidth(48)
            reset_btn.setStyleSheet("font-size:10px; padding:2px;")
            reset_btn.clicked.connect(
                lambda _, i=item: self._reset_auto_stay(i)
            )
            stay_row.addWidget(reset_btn)
        stay_w = QWidget()
        stay_w.setLayout(stay_row)
        self.editor_layout.addRow("Stay Sets:", stay_w)

        # ── Stay angle rotation (manual override) ─────────────────────────
        def _make_angle_row(label_text, angle_val, rotate_fn, reset_fn):
            row_w   = QWidget()
            row_lay = QHBoxLayout(row_w)
            row_lay.setContentsMargins(0, 0, 0, 0)
            row_lay.setSpacing(4)
            angle_lbl = QLabel(f"{int(angle_val) if angle_val is not None else 'Auto'}°")
            angle_lbl.setFixedWidth(40)
            angle_lbl.setStyleSheet("color:#555; font-size:10px;")
            ccw_btn = QPushButton("↺ −15°")
            ccw_btn.setFixedWidth(55)
            ccw_btn.setStyleSheet("font-size:10px; padding:2px;")
            ccw_btn.clicked.connect(lambda _, fn=rotate_fn: fn(-15))
            cw_btn  = QPushButton("↻ +15°")
            cw_btn.setFixedWidth(55)
            cw_btn.setStyleSheet("font-size:10px; padding:2px;")
            cw_btn.clicked.connect(lambda _, fn=rotate_fn: fn(+15))
            rst_btn = QPushButton("Auto")
            rst_btn.setFixedWidth(40)
            rst_btn.setStyleSheet("font-size:10px; padding:2px;")
            rst_btn.clicked.connect(reset_fn)
            row_lay.addWidget(angle_lbl)
            row_lay.addWidget(ccw_btn)
            row_lay.addWidget(cw_btn)
            row_lay.addWidget(rst_btn)
            return row_w

        self.editor_layout.addRow(
            "Stay dir.:",
            _make_angle_row(
                "Stay dir.",
                item.stay_angle_override,
                lambda delta, i=item: self._rotate_stay(i, delta),
                lambda _, i=item: self._reset_stay_angle(i),
            )
        )
        self.editor_layout.addRow(
            "Earth dir.:",
            _make_angle_row(
                "Earth dir.",
                item.earth_angle_override,
                lambda delta, i=item: self._rotate_earth(i, delta),
                lambda _, i=item: self._reset_earth_angle(i),
            )
        )

        # Note
        note = QLineEdit(getattr(item, "custom_note", ""))
        note.setPlaceholderText("Custom note...")
        note.textChanged.connect(
            lambda t, i=item: self._update_note(i, t)
        )
        self.editor_layout.addRow("Note:", note)

        self._add_delete_btn(item)

    # ── Structure editor ──────────────────────────────────────────────────────

    def _build_structure_editor(self, item):
        self.editor_group.setTitle(f"Structure — {item.structure_type}")

        # Structure type
        st_cb = QComboBox()
        st_cb.addItems(["DP", "TP", "4P", "DTR"])
        st_cb.setCurrentText(item.structure_type)
        st_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_structure_type(i, t)
        )
        self.editor_layout.addRow("Structure Type:", st_cb)

        # DTR size (only when DTR)
        if item.structure_type == "DTR":
            dtr_cb = QComboBox()
            dtr_cb.addItems(
                ["None", "10KVA", "16KVA", "25KVA", "63KVA", "100KVA", "160KVA"]
            )
            dtr_cb.setCurrentText(item.dtr_size)
            dtr_cb.currentTextChanged.connect(
                lambda t, i=item: self._update_structure(i, "dtr_size", t)
            )
            self.editor_layout.addRow("DTR Size:", dtr_cb)

        # Pole material
        pt2_cb = QComboBox()
        pt2_cb.addItems(["PCC", "STP", "H-BEAM"])
        pt2_cb.setCurrentText(item.pole_type2)
        pt2_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_struct_type2(i, t)
        )
        self.editor_layout.addRow("Pole Material:", pt2_cb)

        # Height (cascading)
        ht_cb = QComboBox()
        ht_cb.addItems(self._height_options(item.pole_type2))
        ht_cb.setCurrentText(item.height)
        ht_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_structure(i, "height", t)
        )
        self.editor_layout.addRow("Height:", ht_cb)

        # Extension
        ext_chk = QCheckBox("Extension required")
        ext_chk.setChecked(item.has_extension)
        ext_chk.stateChanged.connect(
            lambda v, i=item: self._toggle_struct_extension(i, v == 2)
        )
        self.editor_layout.addRow(ext_chk)

        if item.has_extension:
            ext_ht = QDoubleSpinBox()
            ext_ht.setRange(1.0, 10.0)
            ext_ht.setSingleStep(0.5)
            ext_ht.setSuffix(" m")
            ext_ht.setValue(item.extension_height)
            ext_ht.valueChanged.connect(
                lambda v, i=item: self._update_structure(i, "extension_height", v)
            )
            self.editor_layout.addRow("Ext. Height:", ext_ht)

        # Earth count
        earth_sp = QSpinBox()
        earth_sp.setRange(0, 20)
        earth_sp.setValue(item.earth_count)
        earth_sp.valueChanged.connect(
            lambda v, i=item: self._update_structure(i, "earth_count", v)
        )
        self.editor_layout.addRow("Earthing Sets:", earth_sp)

        # Stay count
        stay_sp = QSpinBox()
        stay_sp.setRange(0, 20)
        stay_sp.setValue(item.stay_count)
        stay_sp.valueChanged.connect(
            lambda v, i=item: self._update_structure(i, "stay_count", v)
        )
        self.editor_layout.addRow("Stay Sets:", stay_sp)

        # Note
        note = QLineEdit(getattr(item, "custom_note", ""))
        note.setPlaceholderText("Custom note...")
        note.textChanged.connect(
            lambda t, i=item: self._update_note(i, t)
        )
        self.editor_layout.addRow("Note:", note)

        self._add_delete_btn(item)

    # ── Span editor ───────────────────────────────────────────────────────────

    def _build_span_editor(self, item):
        if item.is_service_drop:
            self.editor_group.setTitle("Service Connection")
            self._build_service_drop_editor(item)
        else:
            self.editor_group.setTitle("Span")
            self._build_line_span_editor(item)

        note = QLineEdit(getattr(item, "custom_note", ""))
        note.setPlaceholderText("Custom note...")
        note.textChanged.connect(
            lambda t, i=item: self._update_note(i, t)
        )
        self.editor_layout.addRow("Note:", note)
        self._add_delete_btn(item)

    def _build_service_drop_editor(self, item):
        len_sp = QSpinBox()
        len_sp.setRange(1, 150)
        len_sp.setValue(int(item.length))
        len_sp.valueChanged.connect(
            lambda v, i=item: self._update_span(i, "length", v)
        )
        self.editor_layout.addRow("Length (m):", len_sp)

    def _build_line_span_editor(self, item):
        # Voltage level (read-only, auto-detected)
        vl_lbl = QLabel(
            f"{'LT' if item.is_lt_span else 'HT'} (auto-detected)"
        )
        vl_lbl.setStyleSheet("color:#555; font-style:italic;")
        self.editor_layout.addRow("Voltage Level:", vl_lbl)

        len_sp = QSpinBox()
        len_sp.setRange(1, 500)
        len_sp.setValue(int(item.length))
        len_sp.valueChanged.connect(
            lambda v, i=item: self._update_span(i, "length", v)
        )
        self.editor_layout.addRow("Length (m):", len_sp)

        # Conductor type
        cond_cb = QComboBox()
        cond_cb.addItems(["ACSR", "AB Cable", "PVC Cable"])
        cond_cb.setCurrentText(item.conductor)
        cond_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_conductor(i, t)
        )
        self.editor_layout.addRow("Conductor:", cond_cb)

        # Conductor size (cascading)
        sz_cb = QComboBox()
        sz_cb.addItems(self._conductor_sizes(item.conductor, item.is_lt_span))
        sz_cb.setCurrentText(item.conductor_size)
        sz_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_span(i, "conductor_size", t)
        )
        self.editor_layout.addRow("Size:", sz_cb)

        # Wire count (ACSR only)
        if item.conductor == "ACSR":
            wc_cb = QComboBox()
            wc_cb.addItems(["2", "3", "4"])
            wc_cb.setCurrentText(str(item.wire_count))
            wc_cb.currentTextChanged.connect(
                lambda t, i=item: self._update_span(i, "wire_count", t)
            )
            self.editor_layout.addRow("Wire Count:", wc_cb)

        # Work nature
        aug_cb = QComboBox()
        aug_cb.addItems(["New", "Replace 2W->4W", "Add-on 2W"])
        aug_cb.setCurrentText(item.aug_type)
        aug_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_span(i, "aug_type", t)
        )
        self.editor_layout.addRow("Work Nature:", aug_cb)

        # CG
        cg_chk = QCheckBox("Cattle Guard required")
        cg_chk.setChecked(item.has_cg)
        cg_chk.stateChanged.connect(
            lambda v, i=item: self._update_span_refresh(i, "has_cg", v == 2)
        )
        self.editor_layout.addRow(cg_chk)

    # ── Consumer editor ───────────────────────────────────────────────────────

    def _build_consumer_editor(self, item):
        self.editor_group.setTitle("Consumer")

        phase_cb = QComboBox()
        phase_cb.addItems(["1 Phase", "3 Phase"])
        phase_cb.setCurrentText(item.phase)
        phase_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_consumer(i, "phase", t)
        )
        self.editor_layout.addRow("Phase:", phase_cb)

        sz_cb = QComboBox()
        sz_cb.addItems(self._service_cable_sizes(item.phase))
        sz_cb.setCurrentText(item.cable_size)
        sz_cb.currentTextChanged.connect(
            lambda t, i=item: self._update_consumer(i, "cable_size", t)
        )
        self.editor_layout.addRow("Cable Size:", sz_cb)

        agency_chk = QCheckBox("Agency Supplied (not WBSEDCL)")
        agency_chk.setChecked(item.agency_supply)
        agency_chk.stateChanged.connect(
            lambda v, i=item: self._update_consumer(i, "agency_supply", v == 2)
        )
        self.editor_layout.addRow(agency_chk)

        cons_chk = QCheckBox("Include cable in estimate (FDS only)")
        cons_chk.setChecked(getattr(item, "consider_cable", False))
        cons_chk.stateChanged.connect(
            lambda v, i=item: self._update_consumer(i, "consider_cable", v == 2)
        )
        self.editor_layout.addRow(cons_chk)

        note = QLineEdit(getattr(item, "custom_note", ""))
        note.setPlaceholderText("Custom note...")
        note.textChanged.connect(
            lambda t, i=item: self._update_note(i, t)
        )
        self.editor_layout.addRow("Note:", note)

        self._add_delete_btn(item)

    # ── Editor helpers ────────────────────────────────────────────────────────

    def _height_options(self, pole_type2):
        return {
            "PCC":    ["8MTR", "9MTR"],
            "STP":    ["9MTR", "9.5MTR", "11MTR"],
            "H-BEAM": ["13MTR"],
        }.get(pole_type2, ["8MTR", "9MTR"])

    def _conductor_sizes(self, conductor, is_lt):
        if conductor == "ACSR":
            return ["30SQMM", "50SQMM"]
        if conductor == "AB Cable":
            if is_lt:
                return [
                    "3CX50+1CX35",
                    "3CX50+1CX16+1CX35",
                    "3CX70+1CX16+1CX50",
                ]
            else:
                return ["3CX50+1CX150", "3CX95+1CX70"]
        if conductor == "PVC Cable":
            return ["10 SQMM", "16 SQMM", "25 SQMM",
                    "50 SQMM", "95 SQMM", "120 SQMM"]
        return ["10 SQMM"]

    def _service_cable_sizes(self, phase):
        if phase == "1 Phase":
            return ["10 SQMM", "16 SQMM"]
        return ["10 SQMM", "16 SQMM", "25 SQMM", "50 SQMM"]

    def _add_delete_btn(self, item):
        del_btn = QPushButton("🗑 Delete Selected")
        del_btn.setStyleSheet(
            "background:#ff4c4c; color:white; padding:5px; font-weight:bold;"
        )
        del_btn.clicked.connect(lambda: self.delete_item(item))
        self.editor_layout.addRow(del_btn)

    # =========================================================================
    #  UPDATE CALLBACKS
    # =========================================================================

    def _convert_node(self, item, target: str):
        """Convert a SmartPole to a different type or to a SmartStructure."""
        if target.startswith("—"):
            return

        x, y = item.x(), item.y()
        spans = list(item.connected_spans)

        structure_targets = {"DP Structure": "DP", "TP Structure": "TP",
                             "4P Structure": "4P", "DTR": "DTR"}
        pole_targets      = {"LT Pole": "LT", "HT Pole": "HT"}

        if target in structure_targets:
            # ── Pole → Structure ──────────────────────────────────────────
            st = structure_targets[target]
            new_item = SmartStructure(x, y, self.refresh_signal,
                                     detail_view=self.detail_view)
            new_item.structure_type = st
            new_item.earth_count    = SmartStructure._EARTH_DEFAULTS.get(st, 2)
            new_item.stay_count     = getattr(item, "stay_count", 4)
            new_item.update_visuals()

        elif target in pole_targets:
            # ── Pole type change (LT↔HT or toggle existing) ───────────────
            new_item = SmartPole(x, y, self.refresh_signal,
                                 pole_type=pole_targets[target],
                                 is_existing=item.is_existing,
                                 detail_view=self.detail_view)
            new_item.pole_type2       = item.pole_type2
            new_item.height           = item.height
            new_item.has_extension    = item.has_extension
            new_item.extension_height = item.extension_height
            new_item.earth_count      = item.earth_count
            new_item.stay_count       = item.stay_count
            new_item.override_auto_stay    = item.override_auto_stay
            new_item.stay_angle_override   = item.stay_angle_override
            new_item.earth_angle_override  = item.earth_angle_override
            new_item.custom_note      = item.custom_note
            new_item.update_visuals()
        else:
            return

        # ── Re-wire spans ─────────────────────────────────────────────────
        self.scene.addItem(new_item)
        self.scene.addItem(new_item.label)
        new_item.label.setPos(
            -(new_item.label.boundingRect().width() / 2), 14
        )

        for span in spans:
            if span.p1 is item:
                span.p1 = new_item
            if span.p2 is item:
                span.p2 = new_item
            new_item.connected_spans.append(span)
            span.update_position()

        # ── Remove old item ───────────────────────────────────────────────
        self.scene.removeItem(item.label)
        self.scene.removeItem(item)

        # ── Select new item and refresh ───────────────────────────────────
        self.scene.clearSelection()
        new_item.setSelected(True)
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_pole(self, item, prop, value):
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_pole_type2(self, item, value):
        item.pole_type2 = value
        # Reset height to first valid option
        options = self._height_options(value)
        if item.height not in options:
            item.height = options[0]
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _toggle_pole_extension(self, item, value):
        item.has_extension = value
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _manual_stay(self, item, value):
        item.override_auto_stay = True
        item.stay_count = value
        item.update_visuals()
        self.refresh_live_estimate()

    def _reset_auto_stay(self, item):
        item.override_auto_stay = False
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _rotate_stay(self, item, delta: float):
        """Rotate stay symbol ±delta degrees; initialise from auto if not overridden."""
        if item.stay_angle_override is None:
            item.stay_angle_override = item._calc_stay_angle()
        item.stay_angle_override = (item.stay_angle_override + delta) % 360
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _reset_stay_angle(self, item):
        """Clear stay angle override — revert to auto-calculated direction."""
        item.stay_angle_override = None
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _rotate_earth(self, item, delta: float):
        """Rotate earth symbol ±delta degrees; initialise from auto if not overridden."""
        if item.earth_angle_override is None:
            auto_stay = (item.stay_angle_override
                         if item.stay_angle_override is not None
                         else item._calc_stay_angle())
            item.earth_angle_override = (auto_stay + 180) % 360
        item.earth_angle_override = (item.earth_angle_override + delta) % 360
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _reset_earth_angle(self, item):
        """Clear earth angle override — revert to auto (opposite of stay)."""
        item.earth_angle_override = None
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_structure(self, item, prop, value):
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_existing_subtype(self, item, value):
        item.existing_subtype = value
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_structure_type(self, item, value):
        item.structure_type = value
        # Reset earth defaults
        defaults = {"DP": 2, "TP": 3, "4P": 4, "DTR": 5}
        item.earth_count = defaults.get(value, 2)
        if value != "DTR":
            item.dtr_size = "None"
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_struct_type2(self, item, value):
        item.pole_type2 = value
        options = self._height_options(value)
        if item.height not in options:
            item.height = options[0]
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _toggle_struct_extension(self, item, value):
        item.has_extension = value
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_span(self, item, prop, value):
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()

    def _update_span_refresh(self, item, prop, value):
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_conductor(self, item, conductor):
        item.conductor = conductor
        # Reset size to first valid option
        sizes = self._conductor_sizes(conductor, item.is_lt_span)
        item.conductor_size = sizes[0]
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(50, self.on_selection_changed)

    def _update_consumer(self, item, prop, value):
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed)

    def _update_note(self, item, text):
        item.custom_note = text
        item.update_visuals()

    # =========================================================================
    #  DELETION
    # =========================================================================

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Shift:
            if self.current_tool != "SELECT":
                self.set_tool("SELECT")
            event.accept()
            return
        # Ctrl+A — select all canvas items
        if (event.key() == Qt.Key.Key_A and
                event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            for item in self.scene.items():
                item.setSelected(True)
            event.accept()
            return
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            self.delete_selected_items()
        super().keyPressEvent(event)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.KeyPress and event.key() == Qt.Key.Key_Shift:
            if self.current_tool != "SELECT":
                self.set_tool("SELECT")
        return super().eventFilter(obj, event)

    def delete_selected_items(self):
        items = self.scene.selectedItems()
        for item in items:
            if isinstance(item, SmartSpan):
                self.delete_item(item)
        for item in items:
            if isinstance(item, (SmartPole, SmartStructure, SmartConsumer)):
                self.delete_item(item)

    def delete_item(self, item):
        if not item or not item.scene():
            return
        if hasattr(item, "connected_spans"):
            for span in list(item.connected_spans):
                if span.label and span.label.scene():
                    self.scene.removeItem(span.label)
                if span.scene():
                    self.scene.removeItem(span)
                for endpoint in (span.p1, span.p2):
                    if hasattr(endpoint, "connected_spans") and span in endpoint.connected_spans:
                        endpoint.connected_spans.remove(span)
        if isinstance(item, SmartSpan) and item.label and item.label.scene():
            self.scene.removeItem(item.label)
        if item.scene():
            self.scene.removeItem(item)
        self.refresh_live_estimate()

    # =========================================================================
    #  LIVE ESTIMATE ENGINE
    # =========================================================================

    def recalculate_all_span_types(self):
        """
        Propagation logic: spans between two effectively-existing endpoints
        become existing spans (no BOM contribution).
        """
        all_poles = [
            i for i in self.scene.items()
            if isinstance(i, (SmartPole, SmartStructure))
        ]
        existing_set = {p for p in all_poles if getattr(p, "is_existing", False)}

        while True:
            promoted = set()
            for pole in all_poles:
                if pole in existing_set:
                    continue
                existing_connections = sum(
                    1 for s in pole.connected_spans
                    if (s.p1 in existing_set or s.p2 in existing_set)
                    and (s.p1 != pole and s.p2 != pole
                         or (s.p1 in existing_set and s.p2 in existing_set))
                )
                neighbours_existing = sum(
                    1 for s in pole.connected_spans
                    if (s.p1 if s.p2 == pole else s.p2) in existing_set
                )
                if neighbours_existing >= 2:
                    promoted.add(pole)
            if not promoted:
                break
            existing_set.update(promoted)

        for span in self.scene.items():
            if not isinstance(span, SmartSpan):
                continue
            both_existing = (
                span.p1 in existing_set and span.p2 in existing_set
            )
            new_val = both_existing and not span.is_service_drop
            if span.is_existing_span != new_val:
                span.is_existing_span = new_val
                # When a HT ACSR span first becomes existing, default wire_count to 3
                if new_val and not span.is_lt_span and span.conductor == "ACSR":
                    span.wire_count = "3"
                span.update_visuals()

    def _auto_stay_update(self):
        """Auto-update stay counts based on span angles."""
        for pole in self.scene.items():
            if not isinstance(pole, SmartPole):
                continue
            if pole.override_auto_stay:
                continue
            if pole.pole_type == "DTR":
                continue

            active_spans = [
                s for s in pole.connected_spans
                if not s.is_service_drop and not s.is_existing_span
            ]
            n = len(active_spans)
            should_stay = False

            if n == 1:
                should_stay = True
            elif n == 2:
                s1, s2 = active_spans
                other1 = s1.p1 if s1.p2 == pole else s1.p2
                other2 = s2.p1 if s2.p2 == pole else s2.p2
                v1 = (other1.x() - pole.x(), other1.y() - pole.y())
                v2 = (other2.x() - pole.x(), other2.y() - pole.y())
                mag1 = math.hypot(*v1)
                mag2 = math.hypot(*v2)
                if mag1 > 0 and mag2 > 0:
                    dot = v1[0] * v2[0] + v1[1] * v2[1]
                    angle = math.degrees(
                        math.acos(min(1.0, max(-1.0, dot / (mag1 * mag2))))
                    )
                    if (180 - angle) > 20:
                        should_stay = True

            target = 1 if should_stay else 0
            needs_visual_refresh = (pole.stay_count != target)
            if pole.stay_count != target:
                pole.stay_count = target
            if needs_visual_refresh or pole.connected_spans:
                pole.update_visuals()

    def refresh_live_estimate(self):
        if self._refreshing_live:
            return
        self._refreshing_live = True

        try:
            self.recalculate_all_span_types()
            self._auto_stay_update()
            self._refresh_page_grid()   # keep page grid in sync with canvas content

            if not self._is_undoing:
                # Debounce: Capture state 500ms after the last activity/refresh.
                # This cleanly avoids saving history frames 60x/sec during a drag.
                self._history_timer.start(500)

            use_uh        = self.project_meta.get("use_uh", False)
            project_type  = self.project_meta.get("project_type", "NSC")
            sup_rate      = self.project_meta.get("supervision_rate", 0.10)

            # Load rules
            rules = []
            try:
                with open("rules.json", "r") as f:
                    rules = json.load(f)
            except (FileNotFoundError, json.JSONDecodeError):
                pass

            canvas_items = [
                i for i in self.scene.items()
                if isinstance(i, (SmartPole, SmartStructure, SmartSpan, SmartConsumer))
            ]
            raw_bom, raw_lab = self.rule_engine.process(
                canvas_items, rules, use_uh, project_type
            )

            # Apply 3% wastage + sag to steel & conductor material quantities
            _SAG_ITEMS = {
                "M.S Channel 75X40 mm", "M.S Angle 65X65X6mm",
                "M.S Angle 50X50X6mm", "M.S Flat 65X6 mm",
                "M.S Channel 100X50 mm",
                "G.I. Wire 5 MM (6 SWG)", "G.I. Wire 4 MM (8 SWG)",
                "ACSR Conductor 50SQMM (Rabbit)",
                "ACSR Conductor 30SQMM (Weasel)",
                "CABLE (PVC 1.1KV GRADE) 4CORE X10SQMM",
                "CABLE (PVC 1.1KV GRADE) 4CX16SQMM",
                "CABLE (PVC 1.1KV GRADE) 4CX25SQMM",
                "LT AB CABLE 1.1KV 3CX50+1CX16+1CX35SQMM",
            }
            for name in list(raw_bom):
                if name in _SAG_ITEMS:
                    raw_bom[name] = raw_bom[name] * 1.03

            # Build live_bom_data
            self.live_bom_data = []
            conn   = sqlite3.connect("erp_master.db")
            cursor = conn.cursor()
            processed = set()

            combined = (
                [("Material", n, q) for n, q in raw_bom.items()] +
                [("Labor",    n, q) for n, q in raw_lab.items()]
            )

            for item_type, name, qty in combined:
                if name in self.bom_overrides and self.bom_overrides[name]["type"] == item_type:
                    qty = self.bom_overrides[name]["qty"]

                row = self._db_lookup(cursor, item_type, name)
                if row:
                    code, rate, unit = row
                    self.live_bom_data.append({
                        "type": item_type, "code": code, "name": name,
                        "qty": qty, "unit": unit, "rate": rate,
                        "amt": qty * rate
                    })
                processed.add(name)

            # Custom overrides not in auto-BOM
            for name, override in self.bom_overrides.items():
                if name not in processed:
                    row = self._db_lookup(cursor, override["type"], name)
                    if row:
                        code, rate, unit = row
                        qty = override["qty"]
                        self.live_bom_data.append({
                            "type": override["type"], "code": code, "name": name,
                            "qty": qty, "unit": unit, "rate": rate,
                            "amt": qty * rate
                        })

            conn.close()
            self._refresh_table()
            self._recalculate_totals(sup_rate)
        finally:
            self._refreshing_live = False

    def _db_lookup(self, cursor, item_type, name):
        if item_type == "Material":
            cursor.execute(
                "SELECT item_code, rate, unit FROM materials WHERE item_name=?", (name,)
            )
        else:
            cursor.execute(
                "SELECT labor_code, rate, unit FROM labor WHERE task_name=?", (name,)
            )
        return cursor.fetchone()

    def _refresh_table(self):
        try:
            self.live_table.itemChanged.disconnect(self.on_table_edit)
        except TypeError:
            pass

        self.live_table.setRowCount(0)
        for i, item in enumerate(self.live_bom_data):
            self.live_table.insertRow(i)
            self.live_table.setItem(i, 0, QTableWidgetItem(item["type"]))
            self.live_table.setItem(i, 1, QTableWidgetItem(item["code"]))
            self.live_table.setItem(i, 2, QTableWidgetItem(item["name"]))
            qty_item = QTableWidgetItem(f"{item['qty']:.3f}")
            qty_item.setBackground(QColor("#fff3cd"))
            self.live_table.setItem(i, 3, qty_item)
            self.live_table.setItem(i, 4, QTableWidgetItem(item["unit"]))
            self.live_table.setItem(i, 5, QTableWidgetItem(f"{item['amt']:.2f}"))
            for col in (0, 1, 2, 4, 5):
                t = self.live_table.item(i, col)
                if t:
                    t.setFlags(t.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self.live_table.itemChanged.connect(self.on_table_edit)

    def _recalculate_totals(self, sup_rate):
        mat_base = sum(x["amt"] for x in self.live_bom_data if x["type"] == "Material")
        lab_sub  = sum(x["amt"] for x in self.live_bom_data if x["type"] == "Labor")

        now = datetime.now()
        fy_start = now.year if now.month >= 4 else now.year - 1

        self.escalations = []
        cur = mat_base
        for yr in range(2024, fy_start + 1):
            esc = cur * 0.05
            self.escalations.append((f"{str(yr)[-2:]}-{str(yr+1)[-2:]}", esc))
            cur += esc

        sun      = cur * 0.05
        mat_sub  = cur + sun
        sup      = (mat_sub + lab_sub) * sup_rate
        gst      = lab_sub * 0.18
        cess     = (mat_sub + lab_sub + sup) * 0.01
        final    = mat_sub + lab_sub + sup + gst + cess

        self.grand_total_label.setText(
            f"<b>Estimated Cost (incl. taxes): Rs. {final:,.2f}</b>"
        )

    def on_table_edit(self, item):
        if item.column() != 3:
            return
        try:
            new_qty   = float(item.text())
            name_item = self.live_table.item(item.row(), 2)
            type_item = self.live_table.item(item.row(), 0)
            if name_item is None or type_item is None:
                return
            name      = name_item.text()
            row_type  = type_item.text()
            self.bom_overrides[name] = {"qty": new_qty, "type": row_type}
            self.refresh_live_estimate()
        except (ValueError, RuntimeError):
            pass

    # =========================================================================
    #  SEARCH / CUSTOM ITEMS
    # =========================================================================

    def open_search(self, item_type):
        dlg = SearchDialog(item_type, self)
        if dlg.exec():
            sel = dlg.get_selected()
            if sel:
                self.bom_overrides[sel["name"]] = {
                    "qty": 1, "type": sel["type"]
                }
                self.refresh_live_estimate()

    def open_settings_dialog(self):
        SettingsDialog(self).exec()

    def open_db_manager(self):
        DatabaseManagerDialog(self).exec()

    def open_rule_manager(self):
        RulesetManagerDialog(self).exec()

    # =========================================================================
    #  EXCEL EXPORT
    # =========================================================================

    def generate_excel(self):
        m = self.project_meta
        subject = m.get("subject", "ERP_Estimate")
        safe    = "".join(c for c in subject if c not in r'\/*?:"<>|')
        default = f"{safe}_Estimate.xlsx" if safe else "ERP_Estimate.xlsx"

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export ERP Estimate", default, "Excel Files (*.xlsx)"
        )
        if not filename:
            return

        wb = openpyxl.Workbook()
        self._write_estimate_sheet(wb, m)
        self._write_iron_breakup_sheet(wb)
        wb.save(filename)
        QMessageBox.information(self, "Success", f"Excel saved to:\n{filename}")

    def _write_estimate_sheet(self, wb, m):
        ws = wb.active
        ws.title = "Estimate"

        sup_rate = m.get("supervision_rate", 0.10)
        sup_pct  = int(sup_rate * 100)

        # Header
        ws.merge_cells("A1:G1")
        ws["A1"] = "AUTOMATED ERP ESTIMATE"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4F81BD")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = (
            f"Subject: {m.get('subject','')}  |  "
            f"Type: {m.get('project_type','')}  |  "
            f"Date: {datetime.now().strftime('%d-%m-%Y')}"
        )
        ws.merge_cells("A3:G3")
        ws["A3"] = (
            f"Lat: {m.get('lat','')}   Long: {m.get('long','')}   |   "
            f"Materials: {'UH (Readymade)' if m.get('use_uh') else 'Raw Steel'}"
        )

        header_row = ["Sl No.", "Code", "Description", "Qty", "Unit", "Rate", "Amount"]
        ws.append(header_row)
        for cell in ws[4]:
            cell.font = Font(bold=True)
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["B"].width = 15

        row = 5
        mat_items = [x for x in self.live_bom_data if x["type"] == "Material"]
        lab_items = [x for x in self.live_bom_data if x["type"] == "Labor"]

        # ── Materials ──
        ws.cell(row, 3, "A. MATERIALS").font = Font(bold=True)
        row += 1
        for i, item in enumerate(mat_items, 1):
            ws.append([
                i, item["code"], item["name"],
                round(item["qty"], 3), item["unit"],
                item["rate"], round(item["amt"], 2)
            ])
            row += 1

        mat_base = sum(x["amt"] for x in mat_items)
        ws.append(["", "", "Material Base Total", "", "", "", round(mat_base, 2)])
        row += 1

        cur = mat_base
        for fy, esc in self.escalations:
            ws.append([
                "", "", f"Add: Escalation @ 5% for FY {fy}",
                "", "", "", round(esc, 2)
            ])
            row += 1
            cur += esc

        sun     = cur * 0.05
        mat_sub = cur + sun
        ws.append(["", "", "Add: Sundries @ 5%", "", "", "", round(sun, 2)])
        row += 1
        ws.append(["", "", "TOTAL MATERIAL COST (A)", "", "", "", round(mat_sub, 2)])
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 7).font = Font(bold=True)
        row += 2

        # ── Labor ──
        ws.cell(row, 3, "B. ERECTION / LABOR").font = Font(bold=True)
        row += 1
        for i, item in enumerate(lab_items, 1):
            ws.append([
                i, "", item["name"],
                round(item["qty"], 3), item["unit"],
                item["rate"], round(item["amt"], 2)
            ])
            row += 1

        lab_sub = sum(x["amt"] for x in lab_items)
        ws.append(["", "", "TOTAL LABOR COST (B)", "", "", "", round(lab_sub, 2)])
        ws.cell(row, 3).font = Font(bold=True)
        ws.cell(row, 7).font = Font(bold=True)
        row += 2

        # ── Taxes ──
        sup  = (mat_sub + lab_sub) * sup_rate
        gst  = lab_sub * 0.18
        cess = (mat_sub + lab_sub + sup) * 0.01
        sub_c = mat_sub + lab_sub + sup + gst
        g_tot = sub_c + cess

        ws.cell(row, 3, "C. OVERHEADS & TAXES").font = Font(bold=True)
        row += 1
        for label, val in [
            (f"Supervision @ {sup_pct}% on (A+B)", sup),
            ("GST @ 18% on Labour only",            gst),
            ("Sub-Total",                           sub_c),
            ("Add: Cess @ 1% on (Mat+Lab+Sup)",     cess),
            ("GRAND TOTAL",                         g_tot),
        ]:
            ws.append(["", "", label, "", "", "", round(val, 2)])
            row += 1
        ws.cell(row - 1, 3).font = Font(bold=True, size=12)
        ws.cell(row - 1, 7).font = Font(bold=True, size=12, color="FF0000")

    def _write_iron_breakup_sheet(self, wb):
        """
        Generates a detailed Iron Breakup sheet showing where each steel
        item comes from (HT bracket, extension, DTR structure, CG, tee-off
        etc.) with per-source metre/kg rows, plus 3% wastage + sag.
        Quantities are derived from the same rule-engine that drives the
        Estimate sheet so the two sheets always agree.
        """
        ws = wb.create_sheet("Iron Breakup")

        wastage_sag_pct = 0.03

        # Section definitions — same unit-weights as rule_engine formulas
        #   kg_m > 0  →  steel section (metres / kg display)
        #   kg_m == 0 →  wire section  (kg-only display, formula already gives MT)
        sections = [
            ("B", "M.S. Channel (75X40mm)",   "0102010611", 6.8),
            ("B2","M.S. Channel (100X50mm)",  "0102010911", 9.8),
            ("C", "M.S. Angle (65X65X6mm)",   "0101011311", 5.8),
            ("D", "M.S. Angle (50X50X6mm)",   "0101011011", 4.5),
            ("E", "M.S. Flat (65X6mm)",       "0103011511", 3.1),
            ("F", "G.I. Wire 5 MM (6 SWG)",   "0503010811", 0),
            ("G", "G.I. Wire 4 MM (8 SWG)",   "0503010711", 0),
        ]

        detail = self._collect_iron_detail()

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12

        header_fill  = PatternFill("solid", fgColor="4F81BD")
        section_fill = PatternFill("solid", fgColor="D9E1F2")
        total_fill   = PatternFill("solid", fgColor="EBF1DE")
        thin   = Side(border_style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        cr = 1
        ws.cell(cr, 1, "IRON CALCULATION BREAKUP").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=6)
        ws.cell(cr, 1).fill = header_fill
        ws.cell(cr, 1).font = Font(bold=True, size=13, color="FFFFFF")
        ws.cell(cr, 1).alignment = Alignment(horizontal="center")
        cr += 1

        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=6)
        ws.cell(cr, 1, "Steel quantities from rule-engine + 3% wastage & sag")
        ws.cell(cr, 1).alignment = Alignment(horizontal="center")
        ws.cell(cr, 1).font = Font(bold=True, color="2F5597")
        cr += 1

        for sec_key, sec_title, item_code, kg_m in sections:
            rows = detail.get(item_code, [])
            if not rows:
                continue   # skip sections with zero quantity

            # Section header
            ws.cell(cr, 1, sec_key).font = Font(bold=True)
            ws.cell(cr, 2, sec_title).font = Font(bold=True)
            ws.cell(cr, 3, "No").font = Font(bold=True)
            if kg_m:
                ws.cell(cr, 4, "Lgth(m)").font = Font(bold=True)
                ws.cell(cr, 5, "Total(m)").font = Font(bold=True)
            else:
                ws.cell(cr, 4, ""); ws.cell(cr, 5, "")
            ws.cell(cr, 6, "Wt(kg)").font = Font(bold=True)
            for col in range(1, 7):
                ws.cell(cr, col).fill = section_fill
                ws.cell(cr, col).border = border
            cr += 1

            is_wire = (kg_m == 0)  # wire sections show kg only

            subtotal_m  = 0.0
            subtotal_kg = 0.0
            for i, (desc, count, length_each, total_m, wt_kg) in enumerate(rows, 1):
                ws.cell(cr, 1, i)
                ws.cell(cr, 2, desc)
                ws.cell(cr, 3, count if count else "")
                if is_wire:
                    ws.cell(cr, 4, "")
                    ws.cell(cr, 5, "")
                else:
                    ws.cell(cr, 4, round(length_each, 2) if length_each else "")
                    ws.cell(cr, 5, round(total_m, 3))
                ws.cell(cr, 6, round(wt_kg, 2))
                for col in range(1, 7):
                    ws.cell(cr, col).border = border
                subtotal_m  += total_m
                subtotal_kg += wt_kg
                cr += 1

            if is_wire:
                extra_kg = subtotal_kg * wastage_sag_pct
                # Wastage row
                ws.cell(cr, 1, len(rows) + 1)
                ws.cell(cr, 2, "Add: Wastage + Sag @ 3%")
                ws.cell(cr, 3, ""); ws.cell(cr, 4, ""); ws.cell(cr, 5, "")
                ws.cell(cr, 6, round(extra_kg, 2))
                for col in range(1, 7):
                    ws.cell(cr, col).border = border
                cr += 1
                # Total row
                ws.cell(cr, 2, "Total (incl. 3% wastage & sag)").font = Font(bold=True)
                ws.cell(cr, 5, "")
                ws.cell(cr, 6, round(subtotal_kg + extra_kg, 2)).font = Font(bold=True)
            else:
                base_kg  = subtotal_m * kg_m
                extra_m  = subtotal_m * wastage_sag_pct
                extra_kg = base_kg * wastage_sag_pct
                # Wastage row
                ws.cell(cr, 1, len(rows) + 1)
                ws.cell(cr, 2, "Add: Wastage + Sag @ 3%")
                ws.cell(cr, 3, ""); ws.cell(cr, 4, "")
                ws.cell(cr, 5, round(extra_m, 3))
                ws.cell(cr, 6, round(extra_kg, 2))
                for col in range(1, 7):
                    ws.cell(cr, col).border = border
                cr += 1
                # Total row
                ws.cell(cr, 2, "Total (incl. 3% wastage & sag)").font = Font(bold=True)
                ws.cell(cr, 5, round(subtotal_m + extra_m, 3)).font = Font(bold=True)
                ws.cell(cr, 6, round(base_kg + extra_kg, 2)).font = Font(bold=True)

            for col in range(1, 7):
                ws.cell(cr, col).border = border
                ws.cell(cr, col).fill = total_fill
            cr += 2

    def _collect_iron_detail(self):
        """
        Re-evaluate each steel rule per canvas item and return a
        per-item_code list of (description, count, length_each, total_m, wt_kg)
        rows suitable for the detailed Iron Breakup sheet.

        Returns dict  { item_code: [(desc, count, len_each, tot_m, wt_kg), ...] }
        """
        import json as _json

        # unit weights — same as rule_engine.calculate_qty
        # 0 = wire (formula already gives MT, no m→kg conversion)
        UW = {
            "0102010611": 6.8,   # CH_75X40
            "0102010911": 9.8,   # CH_100X50
            "0101011311": 5.8,   # ANG_65X65X6
            "0101011011": 4.5,   # ANG_50X50X6
            "0103011511": 3.1,   # FLAT_65X6
            "0503010811": 0,     # GI Wire 5mm (qty already MT)
            "0503010711": 0,     # GI Wire 4mm (qty already MT)
        }

        # Load rules
        try:
            with open("rules.json", "r") as f:
                rules = _json.load(f)
        except (FileNotFoundError, _json.JSONDecodeError):
            rules = []

        steel_rules = [
            r for r in rules
            if r.get("type") == "Material" and r.get("item_code") in UW
        ]

        use_uh       = self.project_meta.get("use_uh", False)
        project_type = self.project_meta.get("project_type", "NSC")

        # Accum: item_code → { (source_label, len_each) → [count, total_m, total_kg] }
        accum: dict[str, dict[tuple, list]] = {code: {} for code in UW}

        for item in self.scene.items():
            if isinstance(item, SmartPole):
                ctx = self.rule_engine._build_pole_context(item, use_uh, project_type)
            elif isinstance(item, SmartStructure):
                ctx = self.rule_engine._build_structure_context(item, use_uh, project_type)
            elif isinstance(item, SmartSpan):
                ctx = self.rule_engine._build_span_context(item, use_uh, project_type)
            elif isinstance(item, SmartConsumer):
                ctx = self.rule_engine._build_consumer_context(item, use_uh, project_type)
            else:
                continue

            obj_type = ctx.get("object_type", "")

            for rule in steel_rules:
                target = rule.get("object", "")
                if target == "SmartHome" and obj_type == "SmartConsumer":
                    pass
                elif target != obj_type:
                    continue

                if not self.rule_engine.evaluate_rule(ctx, rule.get("condition", "")):
                    continue

                qty_mt = self.rule_engine.calculate_qty(ctx, rule.get("formula", "1"))
                if qty_mt <= 0:
                    continue

                code = rule["item_code"]
                kg_m  = UW[code]
                wt_kg = qty_mt * 1000.0
                tot_m = wt_kg / kg_m if kg_m else 0.0  # 0 for wire

                # Build a human-readable source label
                label, len_each = self._iron_source_label(ctx, rule, tot_m)

                key = (label, len_each)
                if key not in accum[code]:
                    accum[code][key] = [0, 0.0, 0.0]
                accum[code][key][0] += 1
                accum[code][key][1] += tot_m
                accum[code][key][2] += wt_kg

        # Flatten to sorted list
        result: dict[str, list] = {}
        for code, entries in accum.items():
            rows = []
            for (label, len_each), (cnt, tot_m, wt_kg) in sorted(entries.items()):
                rows.append((label, cnt, len_each, round(tot_m, 4), round(wt_kg, 4)))
            if rows:
                result[code] = rows

        return result

    @staticmethod
    def _iron_source_label(ctx, rule, total_m):
        """
        Return (description_string, length_per_unit) for iron breakup row.
        """
        obj = ctx.get("object_type", "")
        cond = rule.get("condition", "")

        if obj == "SmartPole":
            pole_type = ctx.get("pole_type", "")
            is_existing = ctx.get("is_existing", False)
            prefix = "Existing" if is_existing else "New"

            if "has_extension" in cond:
                ext_h = ctx.get("extension_height", 3.0)
                formula = rule.get("formula", "")
                if "FLAT" in formula:
                    return (f"{prefix} {pole_type} Pole Extension Flat ({ext_h}m)", round(total_m, 2))
                return (f"{prefix} {pole_type} Pole Extension ({ext_h}m)", round(total_m, 2))
            elif "has_cg" in cond:
                return ("Cradle Guard (CG) Bracket on Pole", 1.9 if "ANG" in rule.get("formula", "") else 0.5)
            elif "lt_acsr_count" in cond:
                return (f"LT Bracket on {prefix} LT Pole", 1.0)
            elif "ht_spans_count" in cond:
                return (f"Tee-off Bracket on {prefix} HT Pole", round(total_m, 2))
            elif "earth_count" in cond:
                ec = ctx.get("earth_count", 1)
                return (f"Earthing on {prefix} {pole_type} Pole ({ec} nos)", round(total_m, 2))
            else:
                return (f"{prefix} {pole_type} Pole Iron", round(total_m, 2))

        elif obj == "SmartStructure":
            st = ctx.get("structure_type", "")
            # Check earth_count first — applies to any structure type
            if "earth_count" in cond:
                ec = ctx.get("earth_count", 1)
                return (f"Earthing on {st} Structure ({ec} nos)", round(total_m, 2))

            if st == "DTR":
                # Identify specific DTR component from formula
                formula = rule.get("formula", "")
                if "CH_75X40" in formula or "CH_100X50" in formula:
                    return ("DTR Sub-Stn (Channel — Top + Isolator + Base + Bolt)", round(total_m, 2))
                elif "ANG_65X65X6" in formula:
                    return ("DTR Sub-Stn (Angle — Fuse + Switch + Support + FootRest)", round(total_m, 2))
                elif "ANG_50X50X6" in formula:
                    return ("DTR Sub-Stn (Angle 50 — Main Switch)", round(total_m, 2))
                elif "FLAT_65X6" in formula:
                    return ("DTR Sub-Stn (Flat — HT Clamp)", round(total_m, 2))
                else:
                    return (f"DTR Sub-Stn Iron", round(total_m, 2))
            elif st == "DP":
                formula = rule.get("formula", "")
                if "CH_75X40" in formula:
                    return ("DP Structure (Channel)", round(total_m, 2))
                else:
                    return ("DP Structure (Flat)", round(total_m, 2))
            elif st == "TP":
                formula = rule.get("formula", "")
                if "CH_75X40" in formula:
                    return ("TP Structure (Channel)", round(total_m, 2))
                elif "ANG_65X65X6" in formula:
                    return ("TP Structure (Angle)", round(total_m, 2))
                else:
                    return ("TP Structure (Flat)", round(total_m, 2))
            elif st == "4P":
                formula = rule.get("formula", "")
                if "CH_75X40" in formula:
                    return ("4P Structure (Channel)", round(total_m, 2))
                elif "ANG_65X65X6" in formula:
                    return ("4P Structure (Angle)", round(total_m, 2))
                else:
                    return ("4P Structure (Flat)", round(total_m, 2))
            else:
                return (f"{st} Structure Iron", round(total_m, 2))

        elif obj == "SmartSpan":
            formula = rule.get("formula", "")
            if "0503010811" == rule.get("item_code") or "0503010711" == rule.get("item_code"):
                length = ctx.get("length", 0)
                return (f"CG Earthing Wire on Span ({length}m)", round(total_m, 2))
            return ("AB Cable Span (Flat)", 0.5)

        return ("Other Iron", round(total_m, 2))

    # =========================================================================
    #  PDF EXPORT
    # =========================================================================

    def export_pdf(self):
        """
        Export a multi-page PDF whose page layout matches the canvas page grid.
        Each tile visible in the grid becomes one page in the PDF.
        """
        m       = self.project_meta
        subject = m.get("subject", "Project_Drawing")
        safe    = "".join(c for c in subject if c not in r'\/*?:"<>|')
        default = f"{safe}.pdf" if safe else "Project_Drawing.pdf"

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export PDF Drawing", default, "PDF Files (*.pdf)"
        )
        if not filename:
            return

        if self.scene.itemsBoundingRect().isNull():
            QMessageBox.warning(self, "Empty Canvas", "Nothing to export.")
            return

        # Ensure grid tiles are current
        self._refresh_page_grid()
        tiles = self.view.grid_tiles
        if not tiles:
            QMessageBox.warning(self, "Empty Canvas", "Nothing to export.")
            return

        total_pages = tiles[0]["total"]

        # ── UI margins in device pixels ────────────────────────────────────
        MARG_T, MARG_B, MARG_L, MARG_R = 10, 10, 10, 10
        TITLE_H  = 32   # title strip height (device px) — tall enough for 2-line wrap
        FOOTER_H = 16   # footer strip height

        # ScreenResolution keeps device-pixel coords the same as screen coords,
        # so TITLE_H/FOOTER_H sizes and font pt values render correctly in the PDF.
        printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)
        printer.setFullPage(True)

        def _layout_for(orient: str) -> QPageLayout:
            qt_orient = (
                QPageLayout.Orientation.Portrait
                if orient == "P"
                else QPageLayout.Orientation.Landscape
            )
            return QPageLayout(
                QPageSize(QPageSize.PageSizeId.A4),
                qt_orient,
                QMarginsF(0, 0, 0, 0)
            )

        painter = None   # opened on first page

        for page_idx, tile in enumerate(tiles):
            src_rect  = tile["rect"]         # scene-coordinate source
            page_num  = tile["page_num"]
            is_last   = (page_idx == len(tiles) - 1)
            orient    = tile.get("orient", "L")

            if painter is None:
                # First page must set layout before opening painter.
                printer.setPageLayout(_layout_for(orient))
                painter = QPainter(printer)
            else:
                # Set next-page layout then advance to next page.
                printer.setPageLayout(_layout_for(orient))
                printer.newPage()

            paper   = printer.paperRect(QPrinter.Unit.DevicePixel)
            page_w  = paper.width()  - MARG_L - MARG_R
            page_h  = paper.height() - MARG_T  - MARG_B
            ox, oy  = MARG_L, MARG_T

            # ── Border ───────────────────────────────────────────────────
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.setPen(QPen(QColor(180, 180, 180), 0.8))
            painter.drawRect(QRectF(ox, oy, page_w, page_h))

            # ── Title strip ──────────────────────────────────────────────
            painter.fillRect(QRectF(ox, oy, page_w, TITLE_H), QColor(240, 244, 250))
            painter.setPen(Qt.GlobalColor.black)
            painter.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            if self.pdf_show_project_name:
                title_text = m.get("subject") or "ERP PROJECT DRAWING"
                painter.drawText(
                    QRectF(ox + 4, oy, page_w * 0.70, TITLE_H),
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                    | Qt.TextFlag.TextWordWrap,
                    title_text
                )
            # Page number (right side of title)
            painter.setFont(QFont("Arial", 8))
            painter.drawText(
                QRectF(ox + page_w * 0.70, oy, page_w * 0.30 - 4, TITLE_H),
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
                f"Page {page_num} / {total_pages}   [Scale 1:{self.pdf_scale}]"
            )
            # Title separator line
            painter.setPen(QPen(QColor(180, 180, 180), 0.5))
            painter.drawLine(
                QPointF(ox, oy + TITLE_H), QPointF(ox + page_w, oy + TITLE_H)
            )

            # ── Footer strip ─────────────────────────────────────────────
            footer_y = oy + page_h - FOOTER_H
            painter.fillRect(QRectF(ox, footer_y, page_w, FOOTER_H), QColor(240, 244, 250))
            painter.setPen(QPen(QColor(180, 180, 180), 0.5))
            painter.drawLine(QPointF(ox, footer_y), QPointF(ox + page_w, footer_y))
            painter.setPen(Qt.GlobalColor.black)
            painter.setFont(QFont("Arial", 7))
            date_str   = datetime.now().strftime("%d-%m-%Y")
            lat_str    = m.get("lat", "")
            long_str   = m.get("long", "")
            footer_txt = (
                f"{m.get('project_type','')}  |  "
                f"{date_str}  |  "
                f"Lat: {lat_str}   Long: {long_str}  |  "
                f"ERP Estimate Generator v5.0"
            )
            painter.drawText(
                QRectF(ox + 4, footer_y, page_w - 8, FOOTER_H),
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                footer_txt
            )

            # ── Drawing area ─────────────────────────────────────────────
            draw_top  = oy + TITLE_H + 2
            draw_h    = page_h - TITLE_H - FOOTER_H - 4
            draw_rect = QRectF(ox, draw_top, page_w, draw_h)

            self.scene.render(
                painter, draw_rect, src_rect,
                Qt.AspectRatioMode.IgnoreAspectRatio
            )

            # ── Legend: only on the last page ────────────────────────────
            if is_last and self.pdf_show_legend:
                self._draw_pdf_legend(painter, draw_rect)

            # ── Continuation markers at tile edges (Sparse Grid) ─────────────
            _col = tile.get("col", 0)
            _row = tile.get("row", 0)

            def get_neighbor(r, c):
                for t in tiles:
                    if t.get("row") == r and t.get("col") == c:
                        return t.get("page_num")
                return None

            cut_pen = QPen(QColor(200, 80, 80, 160), 0.7)
            cut_pen.setStyle(Qt.PenStyle.DashLine)
            painter.setPen(cut_pen)
            painter.setFont(QFont("Arial", 7, QFont.Weight.Bold))
            painter.setPen(QColor(200, 60, 60, 200))

            # Right edge → next column
            next_right = get_neighbor(_row, _col + 1)
            if next_right:
                painter.drawText(
                    QRectF(ox + page_w - 60, draw_top + 4, 56, 14),
                    Qt.AlignmentFlag.AlignRight, f"Page {next_right}→"
                )
            
            # Left edge ← previous column
            next_left = get_neighbor(_row, _col - 1)
            if next_left:
                painter.drawText(
                    QRectF(ox + 4, draw_top + 4, 56, 14),
                    Qt.AlignmentFlag.AlignLeft, f"←Page {next_left}"
                )
            
            # Bottom edge ↓ next row
            next_down = get_neighbor(_row + 1, _col)
            if next_down:
                painter.drawText(
                    QRectF(ox + page_w / 2 - 30, footer_y - 16, 60, 14),
                    Qt.AlignmentFlag.AlignCenter, f"Page {next_down}↓"
                )
            
            # Top edge ↑ previous row
            next_up = get_neighbor(_row - 1, _col)
            if next_up:
                painter.drawText(
                    QRectF(ox + page_w / 2 - 30, draw_top + 4, 60, 14),
                    Qt.AlignmentFlag.AlignCenter, f"↑Page {next_up}"
                )

        if painter:
            painter.end()

        l_count = sum(1 for t in tiles if t.get("orient") == "L")
        p_count = sum(1 for t in tiles if t.get("orient") == "P")
        override_count = sum(1 for t in tiles if t.get("is_override"))
        orient_summary = f"L:{l_count}  P:{p_count}"
        if override_count:
            orient_summary += f"  |  Overrides:{override_count}"

        QMessageBox.information(
            self, "PDF Exported",
            f"Saved {total_pages} page(s) to:\n{filename}\n\n"
            f"Scale: 1:{self.pdf_scale}  |  "
            f"Orientation: {self.pdf_orientation_mode} ({orient_summary})"
        )

    def _draw_pdf_legend(self, painter, border):
        legend_data = {
            "New LT Pole":      {"s": "🔵", "q": 0},
            "New HT Pole":      {"s": "🔴", "q": 0},
            "DP Structure":     {"s": "🟩", "q": 0},
            "TP Structure":     {"s": "🟩", "q": 0},
            "4P Structure":     {"s": "🟩", "q": 0},
            "DTR":              {"s": "🟠", "q": 0},
            "Existing Pole":    {"s": "⚪", "q": 0},
            "Extension":        {"s": "[E]", "q": 0},
            "Consumer":         {"s": "🏠", "q": 0},
            "Earthing":         {"s": "⏚",  "q": 0},
            "Stay":             {"s": "S→",  "q": 0},
            "CG (SP)":          {"s": "[CG]", "q": 0},
            "CG (DP)":          {"s": "[CG]", "q": 0},
            "New ACSR":         {"s": "---",  "l": 0},
            "New AB Cable":     {"s": "~~~",  "l": 0},
            "New PVC Cable":    {"s": "...",  "l": 0},
            "Existing Span":    {"s": "———",  "l": 0},
            "Service Drop":     {"s": "--s",  "l": 0},
        }

        for item in self.scene.items():
            if isinstance(item, SmartPole):
                legend_data["Earthing"]["q"] += item.earth_count
                legend_data["Stay"]["q"]     += item.stay_count
                if item.has_extension:
                    legend_data["Extension"]["q"] += 1
                if item.is_existing:
                    legend_data["Existing Pole"]["q"] += 1
                elif item.pole_type == "LT":
                    legend_data["New LT Pole"]["q"] += 1
                else:
                    legend_data["New HT Pole"]["q"] += 1
            elif isinstance(item, SmartStructure):
                st_key = item.structure_type if item.structure_type == "DTR" else f"{item.structure_type} Structure"
                if st_key in legend_data:
                    legend_data[st_key]["q"] += 1
                legend_data["Earthing"]["q"]      += item.earth_count
                legend_data["Stay"]["q"]          += item.stay_count
                if item.has_extension:
                    legend_data["Extension"]["q"] += 1
            elif isinstance(item, SmartConsumer):
                legend_data["Consumer"]["q"] += 1
            elif isinstance(item, SmartSpan):
                if item.has_cg:
                    is_dp = isinstance(item.p1, SmartStructure) or isinstance(item.p2, SmartStructure)
                    legend_data["CG (DP)" if is_dp else "CG (SP)"]["q"] += 1
                key = "Service Drop" if item.is_service_drop else (
                    "Existing Span" if item.is_existing_span else
                    f"New {item.conductor}"
                )
                if key in legend_data:
                    if "l" in legend_data[key]:
                        legend_data[key]["l"] += item.length
                    else:
                        legend_data[key]["q"] = legend_data[key].get("q", 0) + 1

        used = []
        for desc, d in legend_data.items():
            q = d.get("q", 0)
            l = d.get("l", 0)
            if q > 0 or l > 0:
                val = str(q) if "q" in d else f"{int(l)}m"
                used.append({"desc": desc, "sym": d["s"], "val": val})

        if not used:
            return

        # ── Layout constants ──────────────────────────────────────────────
        # Two side-by-side mini-tables to halve the vertical footprint.
        # Each half: sym(22) + desc(90) + qty(36) = 148px wide, plus sl(20)
        cw  = {"sl": 18, "sym": 22, "desc": 90, "qty": 34}   # per-column widths
        ckeys = list(cw.keys())
        half_w  = sum(cw.values())          # width of one sub-table
        gap     = 6                          # gap between the two sub-tables
        total_w = half_w * 2 + gap

        row_h  = 14
        hdr_h  = 15
        ll_h   = 16

        # Split entries into two columns (left half / right half)
        mid      = (len(used) + 1) // 2
        left_col  = used[:mid]
        right_col = used[mid:]
        rows     = max(len(left_col), len(right_col))
        total_h  = hdr_h + rows * row_h + ll_h

        # Anchor bottom-right inside the page border
        leg_left = border.right() - total_w - 5
        leg_top  = border.bottom() - total_h - 5
        leg_rect = QRectF(leg_left, leg_top, total_w, total_h)

        painter.save()
        painter.setOpacity(0.82)            # semi-transparent

        # Background
        painter.setBrush(QBrush(QColor(255, 255, 255, 210)))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRect(leg_rect)
        painter.setOpacity(1.0)             # restore for text/lines

        grid_pen   = QPen(QColor(170, 170, 170), 0.4)
        border_pen = QPen(Qt.GlobalColor.black, 0.7)

        def _sub_table(entries, left_x, number_offset):
            """Draw one half-table at left_x, starting index at number_offset."""
            cy = leg_top

            # Header
            painter.setBrush(QBrush(QColor(200, 200, 200, 200)))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRect(QRectF(left_x, cy, half_w, hdr_h))
            painter.setPen(QPen(Qt.GlobalColor.black))
            painter.setFont(QFont("Arial", 6, QFont.Weight.Bold))
            cx = left_x
            for k in ckeys:
                lbl = {"sl": "#", "sym": "Sym", "desc": "Description", "qty": "Qty"}[k]
                painter.drawText(QRectF(cx, cy, cw[k], hdr_h),
                                 Qt.AlignmentFlag.AlignCenter, lbl)
                cx += cw[k]
            cy += hdr_h

            # Header bottom line
            painter.setPen(border_pen)
            painter.drawLine(QPointF(left_x, cy), QPointF(left_x + half_w, cy))

            # Column separators (full height)
            painter.setPen(grid_pen)
            sx = left_x
            for k in ckeys[:-1]:
                sx += cw[k]
                painter.drawLine(QPointF(sx, leg_top), QPointF(sx, leg_top + total_h - ll_h))

            # Data rows
            painter.setFont(QFont("Arial", 6))
            for i, entry in enumerate(entries):
                bg = QColor(248, 248, 248, 200) if i % 2 == 0 else QColor(255, 255, 255, 180)
                painter.setBrush(QBrush(bg))
                painter.setPen(Qt.PenStyle.NoPen)
                painter.drawRect(QRectF(left_x, cy, half_w, row_h))
                painter.setPen(QPen(Qt.GlobalColor.black))
                cx = left_x
                painter.drawText(QRectF(cx, cy, cw["sl"], row_h),
                                 Qt.AlignmentFlag.AlignCenter, str(i + 1 + number_offset))
                cx += cw["sl"]
                painter.drawText(QRectF(cx, cy, cw["sym"], row_h),
                                 Qt.AlignmentFlag.AlignCenter, entry["sym"])
                cx += cw["sym"]
                painter.drawText(QRectF(cx + 2, cy, cw["desc"] - 2, row_h),
                                 Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                                 entry["desc"])
                cx += cw["desc"]
                painter.drawText(QRectF(cx, cy, cw["qty"], row_h),
                                 Qt.AlignmentFlag.AlignCenter, entry["val"])
                cy += row_h
                painter.setPen(grid_pen)
                painter.drawLine(QPointF(left_x, cy), QPointF(left_x + half_w, cy))

            # Outer border for this sub-table
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.setPen(border_pen)
            painter.drawRect(QRectF(left_x, leg_top, half_w, total_h - ll_h))

        _sub_table(left_col,  leg_left,            0)
        _sub_table(right_col, leg_left + half_w + gap, len(left_col))

        # ── Footer (coordinates) — full-width ────────────────────────────
        cy = leg_top + total_h - ll_h
        painter.setBrush(QBrush(QColor(220, 220, 220, 200)))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRect(QRectF(leg_left, cy, total_w, ll_h))
        painter.setPen(QPen(Qt.GlobalColor.black))
        painter.setFont(QFont("Arial", 6, QFont.Weight.Normal, True))
        painter.drawText(
            QRectF(leg_left, cy, total_w, ll_h),
            Qt.AlignmentFlag.AlignCenter,
            f"Lat: {self.project_meta.get('lat', '')}   Long: {self.project_meta.get('long', '')}"
        )
        # Final outer border
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(border_pen)
        painter.drawRect(leg_rect)
        painter.restore()

    # =========================================================================
    #  SAVE / LOAD / AUTOSAVE
    # =========================================================================

    def compile_save_data(self):
        state = {
            "version":       5,
            "project_meta":  self.project_meta,
            "overrides":     self.bom_overrides,
            "nodes":         [],
            "spans":         [],
        }
        node_map = {}
        node_id_by_obj = {}
        for i, item in enumerate(self.scene.items()):
            if isinstance(item, (SmartPole, SmartStructure, SmartConsumer)):
                node_map[i]   = item
                node_id_by_obj[id(item)] = i
                nd = {
                    "id":      i,
                    "type":    (
                        "Pole"      if isinstance(item, SmartPole)      else
                        "Structure" if isinstance(item, SmartStructure) else
                        "Consumer"
                    ),
                    "x":       item.x(),
                    "y":       item.y(),
                    "label_x": item.label.pos().x(),
                    "label_y": item.label.pos().y(),
                    "label_text": item.label.toPlainText(),
                    "custom_note": getattr(item, "custom_note", ""),
                }
                if isinstance(item, SmartPole):
                    nd.update({
                        "pole_type":         item.pole_type,
                        "pole_type2":        item.pole_type2,
                        "is_existing":       item.is_existing,
                        "existing_subtype":   item.existing_subtype,
                        "existing_dtr_size":  getattr(item, "existing_dtr_size", "None"),
                        "height":            item.height,
                        "has_extension":     item.has_extension,
                        "extension_height":  item.extension_height,
                        "earth_count":        item.earth_count,
                        "stay_count":         item.stay_count,
                        "override_auto_stay":  item.override_auto_stay,
                        "stay_angle_override":  item.stay_angle_override,
                        "earth_angle_override": item.earth_angle_override,
                    })
                elif isinstance(item, SmartStructure):
                    nd.update({
                        "structure_type":    item.structure_type,
                        "pole_type2":        item.pole_type2,
                        "height":            item.height,
                        "has_extension":     item.has_extension,
                        "extension_height":  item.extension_height,
                        "earth_count":       item.earth_count,
                        "stay_count":        item.stay_count,
                        "dtr_size":          item.dtr_size,
                    })
                elif isinstance(item, SmartConsumer):
                    nd.update({
                        "phase":           item.phase,
                        "cable_size":      item.cable_size,
                        "agency_supply":   item.agency_supply,
                        "consider_cable":  getattr(item, "consider_cable", False),
                    })
                state["nodes"].append(nd)

        for item in self.scene.items():
            if isinstance(item, SmartSpan):
                p1_id = node_id_by_obj.get(id(item.p1))
                p2_id = node_id_by_obj.get(id(item.p2))
                if p1_id is None or p2_id is None:
                    continue
                state["spans"].append({
                    "p1_id":          p1_id,
                    "p2_id":          p2_id,
                    "length":         item.length,
                    "conductor":      item.conductor,
                    "conductor_size": item.conductor_size,
                    "wire_count":     item.wire_count,
                    "aug_type":       item.aug_type,
                    "has_cg":         item.has_cg,
                    "is_service_drop": item.is_service_drop,
                    "consider_cable": item.consider_cable,
                    "phase":          item.phase,
                    "custom_note":    getattr(item, "custom_note", ""),
                    "label_x":        item.label.pos().x(),
                    "label_y":        item.label.pos().y(),
                    "label_text":     item.label.toPlainText(),
                })

        return state

    def parse_load_data(self, state, fit_view=True):
        self.scene.clear()

        # Support v4 files
        version = state.get("version", 4)

        if version >= 5:
            saved_meta = state.get("project_meta", {})
            self.project_meta = {**DEFAULT_PROJECT_META, **saved_meta}
        else:
            # v4 backward compat
            self.project_meta = dict(DEFAULT_PROJECT_META)
            self.project_meta["subject"] = state.get("subject", "")
            self.project_meta["lat"]     = state.get("lat", "")
            self.project_meta["long"]    = state.get("long", "")
            self.project_meta["use_uh"]  = state.get("uh_toggle", False)

        self._refresh_proj_label()
        self.bom_overrides = state.get("overrides", {})
        node_map = {}

        for nd in state.get("nodes", []):
            ntype = nd.get("type", "Pole")
            x, y  = nd["x"], nd["y"]

            if ntype == "Pole":
                # v4 compat: old DTR poles become SmartStructure
                old_pole_type = nd.get("pole_type", "LT")
                if old_pole_type == "DTR":
                    struct = SmartStructure(
                        x, y, self.refresh_signal, detail_view=self.detail_view
                    )
                    struct.structure_type   = "DTR"
                    struct.dtr_size         = nd.get("dtr_size", "None")
                    struct.earth_count      = nd.get("earth_count", 5)
                    struct.stay_count       = nd.get("stay_count", 4)
                    struct.height           = nd.get("height", "9MTR")
                    struct.update_visuals()
                    struct.label.setPos(nd["label_x"], nd["label_y"])
                    struct.label.setPlainText(nd["label_text"])
                    self.scene.addItem(struct)
                    node_map[nd["id"]] = struct
                else:
                    pole = SmartPole(
                        x, y, self.refresh_signal,
                        old_pole_type,
                        nd.get("is_existing", False),
                        detail_view=self.detail_view
                    )
                    pole.pole_type2       = nd.get("pole_type2", "PCC")
                    pole.height           = nd.get("height", "8MTR")
                    pole.has_extension    = nd.get("has_extension", False)
                    pole.extension_height = nd.get("extension_height", 3.0)
                    pole.earth_count          = nd.get("earth_count", 1)
                    pole.stay_count            = nd.get("stay_count", 0)
                    pole.override_auto_stay    = nd.get("override_auto_stay", False)
                    pole.stay_angle_override   = nd.get("stay_angle_override", None)
                    pole.earth_angle_override  = nd.get("earth_angle_override", None)
                    pole.custom_note           = nd.get("custom_note", "")
                    pole.existing_subtype      = nd.get("existing_subtype", nd.get("pole_type", "LT"))
                    pole.existing_dtr_size     = nd.get("existing_dtr_size", "None")
                    pole.update_visuals()
                    pole.label.setPos(nd["label_x"], nd["label_y"])
                    pole.label.setPlainText(nd["label_text"])
                    self.scene.addItem(pole)
                    node_map[nd["id"]] = pole

            elif ntype == "Structure":
                struct = SmartStructure(
                    x, y, self.refresh_signal, detail_view=self.detail_view
                )
                struct.structure_type   = nd.get("structure_type", "DP")
                struct.pole_type2       = nd.get("pole_type2", "PCC")
                struct.height           = nd.get("height", "9MTR")
                struct.has_extension    = nd.get("has_extension", False)
                struct.extension_height = nd.get("extension_height", 3.0)
                struct.earth_count      = nd.get("earth_count", 2)
                struct.stay_count       = nd.get("stay_count", 4)
                struct.dtr_size         = nd.get("dtr_size", "None")
                struct.custom_note      = nd.get("custom_note", "")
                struct.update_visuals()
                struct.label.setPos(nd["label_x"], nd["label_y"])
                struct.label.setPlainText(nd["label_text"])
                self.scene.addItem(struct)
                node_map[nd["id"]] = struct

            elif ntype in ("Consumer", "Home"):  # "Home" for v4 compat
                consumer = SmartConsumer(
                    x, y, self.refresh_signal, detail_view=self.detail_view
                )
                consumer.phase          = nd.get("phase", "3 Phase")
                consumer.cable_size     = nd.get("cable_size", "10 SQMM")
                consumer.agency_supply  = nd.get("agency_supply", False)
                consumer.consider_cable = nd.get("consider_cable", False)
                consumer.custom_note   = nd.get("custom_note", "")
                consumer.update_visuals()
                consumer.label.setPos(nd["label_x"], nd["label_y"])
                consumer.label.setPlainText(nd["label_text"])
                self.scene.addItem(consumer)
                node_map[nd["id"]] = consumer

        for sd in state.get("spans", []):
            p1 = node_map.get(sd["p1_id"])
            p2 = node_map.get(sd["p2_id"])
            if not (p1 and p2):
                continue
            span = SmartSpan(p1, p2, detail_view=self.detail_view)
            span.length         = sd.get("length", 40)
            span.conductor      = sd.get("conductor", "ACSR")
            # v4 compat: merge wire_size/cable_size into conductor_size
            span.conductor_size = sd.get(
                "conductor_size",
                sd.get("wire_size", sd.get("cable_size", "50SQMM"))
            )
            span.wire_count     = sd.get("wire_count", "3")
            span.aug_type       = sd.get("aug_type", "New")
            span.has_cg         = sd.get("has_cg", False)
            span.is_service_drop = sd.get("is_service_drop", False)
            span.consider_cable  = sd.get("consider_cable", False)
            span.phase           = sd.get("phase", "3 Phase")
            span.custom_note     = sd.get("custom_note", "")
            span.update_visuals()
            span.label.setPos(sd["label_x"], sd["label_y"])
            span.label.setPlainText(sd["label_text"])
            p1.connected_spans.append(span)
            p2.connected_spans.append(span)
            self.scene.addItem(span)
            self.scene.addItem(span.label)

        self.refresh_live_estimate()

        # After loading, optionally fit the view
        if fit_view:
            def _fit_after_load():
                b = self.scene.itemsBoundingRect()
                if not b.isNull():
                    self.view.fitInView(
                        b.adjusted(-60, -60, 60, 60),
                        Qt.AspectRatioMode.KeepAspectRatio
                    )
            QTimer.singleShot(80, _fit_after_load)

    # =========================================================================
    #  UNDO / REDO
    # =========================================================================

    def push_history(self):
        """Capture state and push to undo stack."""
        if self._is_undoing:
            return
        
        state = self.compile_save_data()
        
        # Don't push if nothing structurally changed
        if self.history and self.history_index >= 0:
            if state == self.history[self.history_index]:
                return

        # If user did something after undoing, truncate the 'redo' future
        self.history = self.history[:self.history_index + 1]
        self.history.append(state)
        
        if len(self.history) > 50:
            self.history.pop(0)
        else:
            self.history_index += 1

    def undo(self):
        if self.history_index > 0:
            self.history_index -= 1
            self._is_undoing = True
            
            self.parse_load_data(self.history[self.history_index], fit_view=False)
            
            self._is_undoing = False

    def redo(self):
        if self.history_index < len(self.history) - 1:
            self.history_index += 1
            self._is_undoing = True
            
            self.parse_load_data(self.history[self.history_index], fit_view=False)
            
            self._is_undoing = False

    def _show_blank_start_page(self):
        """Show a clean blank A4 page on startup/new drawing with no auto-added objects."""
        self.last_placed_node = None
        self.span_start_pole  = None
        self.refresh_live_estimate()

        def _fit_blank_page():
            tiles = self.view.grid_tiles
            if tiles:
                self.view.fitInView(
                    tiles[0]["rect"].adjusted(-60, -60, 60, 60),
                    Qt.AspectRatioMode.KeepAspectRatio
                )

        QTimer.singleShot(80, _fit_blank_page)

    def new_drawing(self):
        ans = QMessageBox.question(
            self, "New Canvas", "Clear canvas and start fresh?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if ans == QMessageBox.StandardButton.Yes:
            self.scene.clear()
            self.span_start_pole  = None
            self.last_placed_node = None
            self.bom_overrides.clear()
            self._show_blank_start_page()

    def load_from_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Open Project", "", "JSON Files (*.json)"
        )
        if filename:
            with open(filename, "r") as f:
                self.parse_load_data(json.load(f))

    def save_to_file(self):
        m = self.project_meta
        safe = "".join(
            c for c in m.get("subject", "") if c not in r'\/*?:"<>|'
        )
        default = f"{safe}.json" if safe else "project.json"
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Project", default, "JSON Files (*.json)"
        )
        if filename:
            with open(filename, "w") as f:
                json.dump(self.compile_save_data(), f, indent=2)

    def load_autosave(self):
        loaded_any = False
        if os.path.exists(self.autosave_file):
            try:
                if os.path.getsize(self.autosave_file) > 0:
                    with open(self.autosave_file, "r") as f:
                        data = json.load(f)
                    # Only load if there is actual canvas content
                    if data.get("nodes") or data.get("spans"):
                        self.parse_load_data(data)
                        loaded_any = True
            except (json.JSONDecodeError, KeyError):
                pass
        # Blank canvas — show a blank A4 page only
        if not loaded_any:
            QTimer.singleShot(100, self._show_blank_start_page)

    def closeEvent(self, event):
        with open(self.autosave_file, "w") as f:
            json.dump(self.compile_save_data(), f)
        super().closeEvent(event)

    # =========================================================================
    #  INFO DIALOGS
    # =========================================================================

    def show_about_dialog(self):
        dlg = QMessageBox(self)
        dlg.setWindowTitle("About")
        logo_path = resource_path("logo.svg")
        if os.path.exists(logo_path):
            pix = QPixmap(logo_path).scaled(96, 96, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            dlg.setIconPixmap(pix)
        dlg.setText("""
        <h2>ERP Estimate Generator v5.0</h2>
        <p>Interactive electrical network estimation tool for WBSEDCL projects.</p>
        <ul>
            <li>Project type-based supervision rates</li>
            <li>SmartPole, SmartStructure, SmartSpan, SmartConsumer objects</li>
            <li>Dynamic rule engine with JSON ruleset</li>
            <li>Iron breakup sheet in Excel export</li>
            <li>PDF drawings with legend</li>
        </ul>
        <p><b>Developed by: Pramod Verma</b></p>
        """)
        dlg.exec()

    def show_credits(self):
        QMessageBox.information(self, "Credits", """
        <h2 style='color:#3498db;'>Contributors</h2>
        <ul>
            <li><b>Praful Singh</b> — Visual improvements, PDF legend</li>
            <li><b>Rajsekhar Gorai</b> — 8mtr HT pole extension logic</li>
            <li><b>Amit Karmakar</b> — DTR properties, Lat/Long fields</li>
            <li><b>Santanu Das</b> — Providing data, manuals, circulars for proper integration</li>
            <li><b>Sourabh Jaiswal</b> — Suggesting HT LT restrictions</li>
        </ul>
        <p style='font-style:italic;'>Thanks to all who provided feedback!</p>
        """)

    def show_help(self):
        help_path = resource_path("HELP.html")
        if os.path.exists(help_path):
            with open(help_path, "r", encoding="utf-8") as f:
                html = f.read()
        else:
            html = "<h2>Help file not found</h2><p>HELP.html is missing.</p>"

        dlg = QDialog(self)
        dlg.setWindowTitle("User Guide — ERP Estimate Generator")
        dlg.resize(820, 650)
        lay = QVBoxLayout(dlg)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(True)
        browser.setHtml(html)
        lay.addWidget(browser)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        close_btn.setStyleSheet(
            "padding:6px 20px; font-weight:bold; background:#3498db; color:white;"
        )
        lay.addWidget(close_btn)
        dlg.exec()


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = EstimateApp()
    win.showMaximized()
    sys.exit(app.exec())
